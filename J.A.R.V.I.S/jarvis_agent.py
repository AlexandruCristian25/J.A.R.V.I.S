import os
import re
import csv
from pathlib import Path

from llm_local import ask_llm

try:
    from llm_local import ensure_ollama_ready, ollama_status_report
except Exception:
    ensure_ollama_ready = None
    ollama_status_report = None
from pdf_reader import read_pdf, analyze_pdf, open_pdf

# ==========================================================
# EMBEDDED DEVELOPER ASSISTANT
# Source integrated from developer_assistant.py.
# This removes the need for an external developer_assistant.py file.
# ==========================================================
_EMBEDDED_DEVELOPER_ASSISTANT_CODE = 'import os\nimport re\nimport sys\nimport json\nimport shutil\nimport subprocess\nimport webbrowser\nfrom pathlib import Path\nfrom datetime import datetime\n\n# ==========================================================\n# J.A.R.V.I.S Developer Assistant\n# Steps 1-6:\n# 1. IDE Manager\n# 2. Code Navigator\n# 3. Code Copier / Line Replacement\n# 4. AI Code Reviewer / Improvements\n# 5. Report Generator: MD / DOCX / PDF / PPTX\n# 6. Refactoring Helpers\n# ==========================================================\n\nSKIP_DIRS = {\n    "node_modules", "venv", ".venv", "jarvis-env", "__pycache__",\n    ".git", ".idea", ".vscode", "dist", "build", ".next",\n    ".cache", "site-packages", "file_backups",\n}\n\nTEXT_EXTENSIONS = {\n    ".py", ".js", ".jsx", ".ts", ".tsx", ".html", ".css", ".scss",\n    ".sass", ".json", ".md", ".txt", ".yml", ".yaml", ".env",\n    ".ini", ".cfg", ".toml", ".xml", ".java", ".c", ".cpp",\n    ".h", ".hpp", ".cs", ".php", ".rb", ".go", ".rs", ".sql",\n    ".bat", ".ps1", ".sh",\n}\n\nREPORT_DIR = "developer_reports"\nBACKUP_DIR = "developer_backups"\n\n# ==========================================================\n# BASIC HELPERS\n# ==========================================================\ndef clean_text(text):\n    text = str(text).strip()\n    return re.sub(r"\\s+", " ", text)\n\n\ndef normalize_lower(text):\n    return clean_text(text).lower()\n\n\ndef ensure_dir(path):\n    os.makedirs(path, exist_ok=True)\n\n\ndef timestamp():\n    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")\n\n\ndef safe_read_text(path):\n    try:\n        with open(path, "r", encoding="utf-8", errors="ignore") as f:\n            return f.read()\n    except Exception as e:\n        return f"[READ ERROR] {e}"\n\n\ndef safe_write_text(path, content):\n    with open(path, "w", encoding="utf-8", errors="ignore") as f:\n        f.write(str(content))\n\n\ndef numbered_lines(text, start_line=1):\n    lines = str(text).splitlines()\n    width = len(str(start_line + len(lines)))\n    return "\\n".join(f"{str(index).rjust(width)} | {line}" for index, line in enumerate(lines, start=start_line))\n\n\ndef short(text, limit=900):\n    text = str(text).strip()\n    return text if len(text) <= limit else text[:limit - 3] + "..."\n\n\ndef open_path(path):\n    try:\n        os.startfile(str(path))\n        return True\n    except Exception:\n        pass\n    try:\n        subprocess.Popen(["cmd", "/c", "start", "", str(path)], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, shell=False)\n        return True\n    except Exception:\n        return False\n\n\ndef create_backup(file_path):\n    ensure_dir(BACKUP_DIR)\n    file_path = Path(file_path)\n    if not file_path.exists():\n        return None\n    backup_path = Path(BACKUP_DIR) / f"{file_path.stem}_{timestamp()}{file_path.suffix}.bak"\n    shutil.copy2(file_path, backup_path)\n    return str(backup_path)\n\n# ==========================================================\n# PROJECT RESOLUTION\n# ==========================================================\ndef find_project(project_name):\n    try:\n        from tools import find_project as tools_find_project\n        project = tools_find_project(project_name)\n        if project and isinstance(project, dict):\n            path = project.get("path")\n            if path and os.path.exists(path):\n                return project\n    except Exception:\n        pass\n\n    query = normalize_lower(project_name).replace(" ", "").replace("_", "").replace("-", "")\n    roots = [os.getcwd(), str(Path.home())]\n    for letter in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":\n        drive = f"{letter}:\\\\"\n        if os.path.exists(drive):\n            roots.append(drive)\n\n    for root_dir in roots:\n        try:\n            for root, dirs, files in os.walk(root_dir):\n                dirs[:] = [d for d in dirs if d.lower() not in SKIP_DIRS]\n                base = os.path.basename(root).lower().replace(" ", "").replace("_", "").replace("-", "")\n                if query and (query in base or base in query):\n                    return {"name": os.path.basename(root), "path": root, "type": "Detected"}\n        except Exception:\n            continue\n    return None\n\n\ndef resolve_project_path(project_name):\n    project = find_project(project_name)\n    if not project:\n        return None, f"Project not found: {project_name}"\n    path = project.get("path")\n    if not path or not os.path.exists(path):\n        return None, f"Project path not found: {path}"\n    return path, None\n\n\ndef iter_project_files(project_path):\n    for root, dirs, files in os.walk(project_path):\n        dirs[:] = [d for d in dirs if d.lower() not in SKIP_DIRS]\n        for filename in files:\n            full = os.path.join(root, filename)\n            if os.path.splitext(filename)[1].lower() in TEXT_EXTENSIONS:\n                yield full\n\n\ndef find_file_in_project(project_name, file_query):\n    project_path, error = resolve_project_path(project_name)\n    if error:\n        return None, error\n\n    file_query_clean = normalize_lower(file_query).replace("\\\\", "/")\n    file_query_base = os.path.basename(file_query_clean)\n    candidates = []\n    for path in iter_project_files(project_path):\n        rel = os.path.relpath(path, project_path)\n        rel_lower = rel.lower().replace("\\\\", "/")\n        base_lower = os.path.basename(path).lower()\n        score = 0\n        if rel_lower == file_query_clean:\n            score = 100\n        elif base_lower == file_query_base:\n            score = 95\n        elif file_query_clean in rel_lower:\n            score = 85\n        elif file_query_base and file_query_base in base_lower:\n            score = 75\n        if score:\n            candidates.append((score, path, rel))\n    if not candidates:\n        return None, f"File not found in project {project_name}: {file_query}"\n    candidates.sort(key=lambda item: (-item[0], item[2].lower()))\n    return candidates[0][1], None\n\n# ==========================================================\n# STEP 1 - IDE MANAGER\n# ==========================================================\nIDE_ALIASES = {\n    "vscode": "vscode", "vs code": "vscode", "visual studio code": "vscode", "code": "vscode",\n    "visual studio": "visualstudio", "visual studio community": "visualstudio", "vs community": "visualstudio", "vscommunity": "visualstudio",\n    "intellij": "intellij", "intellij idea": "intellij", "idea": "intellij",\n    "pycharm": "pycharm", "android studio": "androidstudio", "androidstudio": "androidstudio", "eclipse": "eclipse",\n}\n\nIDE_CANDIDATES = {\n    "vscode": [os.path.expandvars(r"%LOCALAPPDATA%\\Programs\\Microsoft VS Code\\Code.exe"), r"C:\\Program Files\\Microsoft VS Code\\Code.exe", "code", "code.cmd"],\n    "visualstudio": [r"C:\\Program Files\\Microsoft Visual Studio\\2022\\Community\\Common7\\IDE\\devenv.exe", r"C:\\Program Files\\Microsoft Visual Studio\\2022\\Professional\\Common7\\IDE\\devenv.exe", r"C:\\Program Files\\Microsoft Visual Studio\\2022\\Enterprise\\Common7\\IDE\\devenv.exe", "devenv.exe"],\n    "intellij": [os.path.expandvars(r"%LOCALAPPDATA%\\JetBrains\\Toolbox\\scripts\\idea.cmd"), r"C:\\Program Files\\JetBrains\\IntelliJ IDEA Community Edition 2024.3\\bin\\idea64.exe", r"C:\\Program Files\\JetBrains\\IntelliJ IDEA 2024.3\\bin\\idea64.exe", "idea64.exe", "idea"],\n    "pycharm": [os.path.expandvars(r"%LOCALAPPDATA%\\JetBrains\\Toolbox\\scripts\\pycharm.cmd"), r"C:\\Program Files\\JetBrains\\PyCharm Community Edition 2024.3\\bin\\pycharm64.exe", r"C:\\Program Files\\JetBrains\\PyCharm 2024.3\\bin\\pycharm64.exe", "pycharm64.exe", "pycharm"],\n    "androidstudio": [r"C:\\Program Files\\Android\\Android Studio\\bin\\studio64.exe", "studio64.exe"],\n    "eclipse": ["eclipse.exe"],\n}\n\n\ndef resolve_ide(ide_name):\n    key = IDE_ALIASES.get(normalize_lower(ide_name), normalize_lower(ide_name))\n    for candidate in IDE_CANDIDATES.get(key, []):\n        expanded = os.path.expandvars(candidate)\n        if os.path.exists(expanded):\n            return key, expanded\n        found = shutil.which(expanded) or shutil.which(candidate)\n        if found:\n            return key, found\n    return key, None\n\n\ndef open_project_in_ide(project_name, ide_name):\n    project_path, error = resolve_project_path(project_name)\n    if error:\n        return error\n    ide_key, ide_path = resolve_ide(ide_name)\n    if not ide_path:\n        return f"Could not find IDE: {ide_name}\\nProject found at: {project_path}\\nInstall the IDE or add it to PATH, then try again."\n    try:\n        subprocess.Popen([ide_path, project_path], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, shell=False)\n        return f"Opening {project_name} in {ide_name}"\n    except Exception as e:\n        return f"Could not open project in {ide_name}: {e}"\n\n\ndef open_file_in_ide(project_name, file_query, ide_name):\n    file_path, error = find_file_in_project(project_name, file_query)\n    if error:\n        return error\n    ide_key, ide_path = resolve_ide(ide_name)\n    if not ide_path:\n        return f"Could not find IDE: {ide_name}"\n    try:\n        subprocess.Popen([ide_path, file_path], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, shell=False)\n        return f"Opening {os.path.basename(file_path)} in {ide_name}"\n    except Exception as e:\n        return f"Could not open file in {ide_name}: {e}"\n\n# ==========================================================\n# STEP 2 - CODE NAVIGATOR\n# ==========================================================\ndef read_file_lines(project_name, file_query, start_line=None, end_line=None):\n    file_path, error = find_file_in_project(project_name, file_query)\n    if error:\n        return error\n    text = safe_read_text(file_path)\n    if text.startswith("[READ ERROR]"):\n        return text\n    lines = text.splitlines()\n    start_line = 1 if start_line is None else max(1, int(start_line))\n    end_line = len(lines) if end_line is None else min(len(lines), int(end_line))\n    if start_line > end_line:\n        return "Invalid line range."\n    selected = "\\n".join(lines[start_line - 1:end_line])\n    return f"File: {file_path}\\nLines: {start_line}-{end_line}\\n\\n```text\\n{numbered_lines(selected, start_line)}\\n```"\n\n\ndef find_function_in_file(project_name, file_query, function_name):\n    file_path, error = find_file_in_project(project_name, file_query)\n    if error:\n        return error\n    lines = safe_read_text(file_path).splitlines()\n    pattern = re.compile(rf"^\\s*(def|function|const|let|var|class|public|private|protected|static)\\s+{re.escape(function_name)}\\b|^\\s*{re.escape(function_name)}\\s*[:=]", re.IGNORECASE)\n    matches = [(i, line) for i, line in enumerate(lines, start=1) if pattern.search(line)]\n    if not matches:\n        return f"Function/symbol not found: {function_name}"\n    output = [f"Matches for \'{function_name}\' in {file_path}:", ""]\n    output.extend(f"- line {line_no}: {line.strip()}" for line_no, line in matches[:20])\n    return "\\n".join(output)\n\n# ==========================================================\n# STEP 3 - CODE COPIER / SAFE LINE REPLACEMENT\n# ==========================================================\ndef copy_lines_between_files(source_project, source_file, source_start, source_end, target_project, target_file, target_start, target_end, mode="replace"):\n    source_path, error = find_file_in_project(source_project, source_file)\n    if error:\n        return error\n    target_path, error = find_file_in_project(target_project, target_file)\n    if error:\n        return error\n    source_lines = safe_read_text(source_path).splitlines()\n    target_lines = safe_read_text(target_path).splitlines()\n    source_start = max(1, int(source_start)); source_end = min(len(source_lines), int(source_end))\n    target_start = max(1, int(target_start)); target_end = min(len(target_lines), int(target_end))\n    if source_start > source_end or target_start > target_end:\n        return "Invalid source or target line range."\n    extracted = source_lines[source_start - 1:source_end]\n    backup = create_backup(target_path)\n    if mode == "insert":\n        new_lines = target_lines[:target_start - 1] + extracted + target_lines[target_start - 1:]\n    else:\n        new_lines = target_lines[:target_start - 1] + extracted + target_lines[target_end:]\n    safe_write_text(target_path, "\\n".join(new_lines) + "\\n")\n    return f"Code transfer completed safely.\\nSource: {source_path} lines {source_start}-{source_end}\\nTarget: {target_path} lines {target_start}-{target_end}\\nMode: {mode}\\nBackup: {backup}"\n\n\ndef replace_lines_with_text(project_name, file_query, start_line, end_line, new_code):\n    file_path, error = find_file_in_project(project_name, file_query)\n    if error:\n        return error\n    lines = safe_read_text(file_path).splitlines()\n    start_line = max(1, int(start_line)); end_line = min(len(lines), int(end_line))\n    if start_line > end_line:\n        return "Invalid line range."\n    backup = create_backup(file_path)\n    new_lines = lines[:start_line - 1] + str(new_code).splitlines() + lines[end_line:]\n    safe_write_text(file_path, "\\n".join(new_lines) + "\\n")\n    return f"Lines replaced successfully.\\nFile: {file_path}\\nLines: {start_line}-{end_line}\\nBackup: {backup}"\n\n# ==========================================================\n# STEP 4 - AI REVIEWER / IMPROVER\n# ==========================================================\ndef ask_ai(prompt):\n    try:\n        from llm_local import ask_llm\n        return ask_llm(prompt)\n    except Exception as e:\n        return "AI engine unavailable.\\nReason: " + str(e) + "\\n\\nFallback: check readability, naming, duplication, error handling, security and edge cases."\n\n\ndef review_code_text(code_text, context="code"):\n    prompt = f"""\nYou are J.A.R.V.I.S, a senior software engineering assistant.\nReview the following {context}.\nReturn:\n1. Short summary\n2. Bugs or risks\n3. Security issues\n4. Performance issues\n5. Maintainability improvements\n6. Improved version of the code if possible\n\nCODE:\n```text\n{code_text}\n```\n"""\n    return ask_ai(prompt)\n\n\ndef improve_code_text(code_text, context="code"):\n    prompt = f"""\nYou are J.A.R.V.I.S, a senior software engineer.\nImprove the following {context}.\nKeep behavior the same unless there is an obvious bug.\nReturn:\n1. What you changed\n2. Improved code only in a code block\n3. Notes\n\nCODE:\n```text\n{code_text}\n```\n"""\n    return ask_ai(prompt)\n\n\ndef review_file(project_name, file_query, start_line=None, end_line=None):\n    file_path, error = find_file_in_project(project_name, file_query)\n    if error:\n        return error\n    text = safe_read_text(file_path)\n    lines = text.splitlines()\n    if start_line is not None and end_line is not None:\n        start_line = max(1, int(start_line)); end_line = min(len(lines), int(end_line))\n        code = "\\n".join(lines[start_line - 1:end_line])\n        context = f"{file_path}, lines {start_line}-{end_line}"\n    else:\n        code = text\n        context = file_path\n    return review_code_text(code, context=context)\n\n\ndef improve_file(project_name, file_query, start_line=None, end_line=None):\n    file_path, error = find_file_in_project(project_name, file_query)\n    if error:\n        return error\n    text = safe_read_text(file_path)\n    lines = text.splitlines()\n    if start_line is not None and end_line is not None:\n        start_line = max(1, int(start_line)); end_line = min(len(lines), int(end_line))\n        code = "\\n".join(lines[start_line - 1:end_line])\n        context = f"{file_path}, lines {start_line}-{end_line}"\n    else:\n        code = text\n        context = file_path\n    return improve_code_text(code, context=context)\n\n# ==========================================================\n# STEP 5 - REPORT GENERATOR\n# ==========================================================\ndef create_markdown_report(title, content, open_after=True):\n    ensure_dir(REPORT_DIR)\n    path = Path(REPORT_DIR) / f"{title.replace(\' \', \'_\')}_{timestamp()}.md"\n    safe_write_text(path, "\\n".join([f"# {title}", "", f"Generated: {datetime.now().strftime(\'%Y-%m-%d %H:%M:%S\')}", "", str(content), ""]))\n    if open_after:\n        open_path(path)\n    return str(path)\n\n\ndef create_docx_report(title, content, open_after=True):\n    ensure_dir(REPORT_DIR)\n    try:\n        from docx import Document\n    except Exception:\n        md_path = create_markdown_report(title, content, open_after=open_after)\n        return f"python-docx is not installed. Markdown report created instead:\\n{md_path}"\n    path = Path(REPORT_DIR) / f"{title.replace(\' \', \'_\')}_{timestamp()}.docx"\n    doc = Document()\n    doc.add_heading(title, level=1)\n    doc.add_paragraph(f"Generated: {datetime.now().strftime(\'%Y-%m-%d %H:%M:%S\')}")\n    for block in str(content).split("\\n\\n"):\n        doc.add_paragraph(block)\n    doc.save(path)\n    if open_after:\n        open_path(path)\n    return str(path)\n\n\ndef create_pdf_report(title, content, open_after=True):\n    ensure_dir(REPORT_DIR)\n    path = Path(REPORT_DIR) / f"{title.replace(\' \', \'_\')}_{timestamp()}.pdf"\n    try:\n        from reportlab.lib.pagesizes import A4\n        from reportlab.pdfgen import canvas\n    except Exception:\n        md_path = create_markdown_report(title, content, open_after=open_after)\n        return f"reportlab is not installed. Markdown report created instead:\\n{md_path}"\n    c = canvas.Canvas(str(path), pagesize=A4)\n    width, height = A4\n    x, y = 42, height - 50\n    c.setFont("Helvetica-Bold", 16); c.drawString(x, y, title[:80]); y -= 28\n    c.setFont("Helvetica", 9); c.drawString(x, y, f"Generated: {datetime.now().strftime(\'%Y-%m-%d %H:%M:%S\')}"); y -= 28\n    for raw_line in str(content).splitlines():\n        if y < 50:\n            c.showPage(); c.setFont("Helvetica", 9); y = height - 50\n        c.drawString(x, y, raw_line[:115]); y -= 13\n    c.save()\n    if open_after:\n        open_path(path)\n    return str(path)\n\n\ndef create_pptx_report(title, content, open_after=True):\n    ensure_dir(REPORT_DIR)\n    try:\n        from pptx import Presentation\n    except Exception:\n        md_path = create_markdown_report(title, content, open_after=open_after)\n        return f"python-pptx is not installed. Markdown report created instead:\\n{md_path}"\n    path = Path(REPORT_DIR) / f"{title.replace(\' \', \'_\')}_{timestamp()}.pptx"\n    prs = Presentation()\n    slide = prs.slides.add_slide(prs.slide_layouts[0])\n    slide.shapes.title.text = title\n    slide.placeholders[1].text = f"Generated: {datetime.now().strftime(\'%Y-%m-%d %H:%M:%S\')}"\n    for section in str(content).split("\\n\\n")[:12]:\n        slide = prs.slides.add_slide(prs.slide_layouts[1])\n        lines = section.splitlines()\n        slide.shapes.title.text = short(lines[0] if lines else "Report", 60)\n        slide.placeholders[1].text = short("\\n".join(lines[1:]) if len(lines) > 1 else section, 900)\n    prs.save(path)\n    if open_after:\n        open_path(path)\n    return str(path)\n\n\ndef create_report(title, content, format_type="md", open_after=True):\n    fmt = normalize_lower(format_type)\n    if fmt in {"word", "docx", "doc"}:\n        path = create_docx_report(title, content, open_after=open_after)\n    elif fmt == "pdf":\n        path = create_pdf_report(title, content, open_after=open_after)\n    elif fmt in {"ppt", "pptx", "powerpoint", "presentation"}:\n        path = create_pptx_report(title, content, open_after=open_after)\n    else:\n        path = create_markdown_report(title, content, open_after=open_after)\n    return f"Report created:\\n{path}"\n\n\ndef review_project_report(project_name, format_type="md"):\n    try:\n        from project_review_assistant import review_project\n        content = review_project(project_name)\n    except Exception:\n        project_path, error = resolve_project_path(project_name)\n        if error:\n            return error\n        content = f"Project review fallback for {project_name}\\nPath: {project_path}\\n\\nReview engine unavailable."\n    return create_report(f"JARVIS Project Review - {project_name}", content, format_type=format_type, open_after=True)\n\n\ndef review_file_report(project_name, file_query, format_type="md"):\n    return create_report(f"JARVIS File Review - {file_query}", review_file(project_name, file_query), format_type=format_type, open_after=True)\n\n# ==========================================================\n# STEP 6 - REFACTORING HELPERS\n# ==========================================================\ndef remove_trailing_whitespace(project_name, file_query):\n    file_path, error = find_file_in_project(project_name, file_query)\n    if error:\n        return error\n    lines = safe_read_text(file_path).splitlines()\n    backup = create_backup(file_path)\n    safe_write_text(file_path, "\\n".join(line.rstrip() for line in lines) + "\\n")\n    return f"Trailing whitespace removed.\\nFile: {file_path}\\nBackup: {backup}"\n\n\ndef add_header_comment(project_name, file_query, comment):\n    file_path, error = find_file_in_project(project_name, file_query)\n    if error:\n        return error\n    text = safe_read_text(file_path)\n    backup = create_backup(file_path)\n    ext = Path(file_path).suffix.lower()\n    if ext in {".py", ".sh", ".ps1"}:\n        header = "\\n".join(f"# {line}" for line in str(comment).splitlines())\n    elif ext in {".js", ".jsx", ".ts", ".tsx", ".java", ".c", ".cpp", ".cs", ".go", ".rs", ".php"}:\n        header = "/*\\n" + "\\n".join(str(comment).splitlines()) + "\\n*/"\n    else:\n        header = str(comment)\n    safe_write_text(file_path, header + "\\n\\n" + text)\n    return f"Header comment added.\\nFile: {file_path}\\nBackup: {backup}"\n\n\ndef format_light(project_name, file_query):\n    return remove_trailing_whitespace(project_name, file_query)\n\n# ==========================================================\n# COMMAND ROUTER\n# ==========================================================\ndef handle_developer_command(command):\n    original = str(command).strip()\n    lower = normalize_lower(original)\n\n    match = re.match(r"^open project (.+?) in (vs code|vscode|visual studio code|visual studio community|vs community|visual studio|intellij|intellij idea|idea|pycharm|android studio|eclipse)$", lower)\n    if match:\n        project_name = original[13: original.lower().rfind(" in ")].strip()\n        ide_name = original[original.lower().rfind(" in ") + 4:].strip()\n        return open_project_in_ide(project_name, ide_name)\n\n    match = re.match(r"^open file (.+?) from (.+?) in (.+)$", original, flags=re.IGNORECASE)\n    if match:\n        return open_file_in_ide(match.group(2).strip(), match.group(1).strip(), match.group(3).strip())\n\n    match = re.match(r"^(show|read|get|give me) lines? (\\d+) (to|-|until) (\\d+) from file (.+?) (in|from) project (.+)$", original, flags=re.IGNORECASE)\n    if match:\n        return read_file_lines(match.group(7).strip(), match.group(5).strip(), int(match.group(2)), int(match.group(4)))\n\n    match = re.match(r"^(give me|show|get|read) (the )?(code )?(from )?line (\\d+) (to|-|until) (line )?(\\d+) from (.+?) (in|from) project (.+)$", original, flags=re.IGNORECASE)\n    if match:\n        return read_file_lines(match.group(11).strip(), match.group(9).strip(), int(match.group(5)), int(match.group(8)))\n\n    match = re.match(r"^find (function|class|symbol) (.+?) in file (.+?) (from|in) project (.+)$", original, flags=re.IGNORECASE)\n    if match:\n        return find_function_in_file(match.group(5).strip(), match.group(3).strip(), match.group(2).strip())\n\n    match = re.match(r"^(copy|replace|move) lines? (\\d+) (to|-|until) (\\d+) from file (.+?) (in|from) project (.+?) to file (.+?) (in|from) project (.+?) lines? (\\d+) (to|-|until) (\\d+)$", original, flags=re.IGNORECASE)\n    if match:\n        return copy_lines_between_files(match.group(7).strip(), match.group(5).strip(), int(match.group(2)), int(match.group(4)), match.group(10).strip(), match.group(8).strip(), int(match.group(11)), int(match.group(13)), mode="replace")\n\n    match = re.match(r"^review lines? (\\d+) (to|-|until) (\\d+) from file (.+?) (in|from) project (.+)$", original, flags=re.IGNORECASE)\n    if match:\n        return review_file(match.group(6).strip(), match.group(4).strip(), int(match.group(1)), int(match.group(3)))\n\n    match = re.match(r"^review file (.+?) (in|from) project (.+)$", original, flags=re.IGNORECASE)\n    if match:\n        return review_file(match.group(3).strip(), match.group(1).strip())\n\n    match = re.match(r"^(improve|optimize|refactor) lines? (\\d+) (to|-|until) (\\d+) from file (.+?) (in|from) project (.+)$", original, flags=re.IGNORECASE)\n    if match:\n        return improve_file(match.group(7).strip(), match.group(5).strip(), int(match.group(2)), int(match.group(4)))\n\n    match = re.match(r"^(improve|optimize|refactor) file (.+?) (in|from) project (.+)$", original, flags=re.IGNORECASE)\n    if match:\n        return improve_file(match.group(4).strip(), match.group(2).strip())\n\n    match = re.match(r"^generate (word|docx|pdf|ppt|pptx|powerpoint|markdown|md) (report|review) for project (.+)$", original, flags=re.IGNORECASE)\n    if match:\n        return review_project_report(match.group(3).strip(), match.group(1).strip())\n\n    match = re.match(r"^generate (word|docx|pdf|ppt|pptx|powerpoint|markdown|md) (report|review) for file (.+?) (in|from) project (.+)$", original, flags=re.IGNORECASE)\n    if match:\n        return review_file_report(match.group(5).strip(), match.group(3).strip(), match.group(1).strip())\n\n    match = re.match(r"^(format|clean) file (.+?) (in|from) project (.+)$", original, flags=re.IGNORECASE)\n    if match:\n        return format_light(match.group(4).strip(), match.group(2).strip())\n\n    return None\n\n\nif __name__ == "__main__":\n    print("J.A.R.V.I.S Developer Assistant")\n    print("Type a developer command or \'exit\'.")\n    while True:\n        cmd = input("\\nDeveloper command: ").strip()\n        if cmd.lower() in {"exit", "quit"}:\n            break\n        result = handle_developer_command(cmd)\n        print(result if result is not None else "Command not handled by developer_assistant.py")\n'

_EMBEDDED_DEVELOPER_ASSISTANT_NAMESPACE = {
    "__name__": "_embedded_developer_assistant",
    "__file__": "embedded_developer_assistant.py",
}

exec(
    _EMBEDDED_DEVELOPER_ASSISTANT_CODE,
    _EMBEDDED_DEVELOPER_ASSISTANT_NAMESPACE
)

handle_developer_command = _EMBEDDED_DEVELOPER_ASSISTANT_NAMESPACE["handle_developer_command"]


# ==========================================================
# NATURAL DEVELOPER COMMAND ROUTER
# Extra layer for commands like:
# - read lines 1-100 from hud.py in project J.A.R.V.I.S
# - find function handle_command in jarvis_agent.py from project J.A.R.V.I.S
# - copy lines 20-80 from logger.py in CyberShield AI to app.py in ManagerApp at line 150
# - generate Excel report for project CyberShield AI
# ==========================================================
def _jarvis_norm(text):
    return str(text).strip()


def _jarvis_parse_project_file(project_name, file_query):
    """
    Uses the embedded developer assistant resolver first.
    Falls back to project_file_assistant if available.
    """
    try:
        ns = globals().get("_EMBEDDED_DEVELOPER_ASSISTANT_NAMESPACE", {})
        finder = ns.get("find_file_in_project")

        if finder:
            path, error = finder(project_name, file_query)

            if not error and path:
                return path, None
    except Exception as e:
        last_error = str(e)
    else:
        last_error = None

    try:
        from project_file_assistant import find_project_file
        item, error = find_project_file(project_name, file_query)

        if not error and item:
            return item.get("full_path"), None

        return None, error
    except Exception as e:
        return None, last_error or str(e)


def _jarvis_read_lines_direct(project_name, file_query, start_line, end_line):
    try:
        ns = globals().get("_EMBEDDED_DEVELOPER_ASSISTANT_NAMESPACE", {})
        reader = ns.get("read_file_lines")

        if reader:
            return reader(project_name, file_query, start_line, end_line)
    except Exception:
        pass

    try:
        from project_file_assistant import read_project_file_lines
        return read_project_file_lines(project_name, file_query, start_line, end_line)
    except Exception as e:
        return f"Could not read lines: {e}"


def _jarvis_find_symbol_direct(project_name, file_query, symbol_name, symbol_type="symbol"):
    try:
        ns = globals().get("_EMBEDDED_DEVELOPER_ASSISTANT_NAMESPACE", {})
        finder = ns.get("find_function_in_file")

        if finder:
            return finder(project_name, file_query, symbol_name)
    except Exception:
        pass

    file_path, error = _jarvis_parse_project_file(project_name, file_query)

    if error:
        return error

    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            lines = f.read().splitlines()
    except Exception as e:
        return f"Could not read file: {e}"

    if symbol_type == "class":
        patterns = [
            rf"^\s*class\s+{re.escape(symbol_name)}\b",
            rf"^\s*export\s+class\s+{re.escape(symbol_name)}\b",
        ]
    elif symbol_type == "function":
        patterns = [
            rf"^\s*def\s+{re.escape(symbol_name)}\b",
            rf"^\s*async\s+def\s+{re.escape(symbol_name)}\b",
            rf"^\s*function\s+{re.escape(symbol_name)}\b",
            rf"^\s*const\s+{re.escape(symbol_name)}\s*=",
            rf"^\s*export\s+function\s+{re.escape(symbol_name)}\b",
        ]
    else:
        patterns = [
            rf"\b{re.escape(symbol_name)}\b",
        ]

    output = [
        f"Matches for {symbol_type} '{symbol_name}' in {file_path}:",
        ""
    ]

    found = 0

    for index, line in enumerate(lines, start=1):
        if any(re.search(pattern, line, flags=re.IGNORECASE) for pattern in patterns):
            output.append(f"- line {index}: {line.strip()}")
            found += 1

            if found >= 30:
                break

    if found == 0:
        return f"{symbol_type.title()} not found: {symbol_name}"

    return "\n".join(output)


def _jarvis_copy_lines_direct(source_project, source_file, source_start, source_end, target_project, target_file, target_start, target_end=None):
    try:
        from project_file_assistant import copy_project_file_lines_between_projects
        return copy_project_file_lines_between_projects(
            source_project,
            source_file,
            source_start,
            source_end,
            target_project,
            target_file,
            target_start,
            target_end,
            mode="insert" if target_end is None else "replace"
        )
    except Exception:
        pass

    try:
        ns = globals().get("_EMBEDDED_DEVELOPER_ASSISTANT_NAMESPACE", {})
        copier = ns.get("copy_lines_between_files")

        if copier and target_end is not None:
            return copier(
                source_project,
                source_file,
                source_start,
                source_end,
                target_project,
                target_file,
                target_start,
                target_end,
                mode="replace"
            )
    except Exception as e:
        return f"Could not copy lines: {e}"

    return "Copy command requires project_file_assistant.py for insert-at-line mode."


def _jarvis_build_report_content(project_name, report_kind="project"):
    kind = str(report_kind).lower().strip()

    try:
        if kind in {"security", "security review", "audit"}:
            return full_security_audit(project_name)
    except Exception:
        pass

    try:
        if kind in {"architecture", "architecture review"}:
            return strict_architecture_analyzer_project(project_name)
    except Exception:
        pass

    try:
        if kind in {"score", "score report"}:
            return score_project(project_name)
    except Exception:
        pass

    try:
        # Prefer enterprise report if your upgraded project_review_assistant.py has it.
        from project_review_assistant import build_enterprise_project_review
        return build_enterprise_project_review(project_name)
    except Exception:
        pass

    try:
        return autonomous_review_project(project_name)
    except Exception:
        pass

    try:
        return review_project(project_name)
    except Exception as e:
        return f"Could not build report content: {e}"


def _jarvis_export_excel_report(project_name, content, report_kind="project_review", open_after=True):
    reports_dir = "developer_reports"
    os.makedirs(reports_dir, exist_ok=True)

    safe_project = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(project_name)).strip("_") or "project"
    safe_kind = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(report_kind)).strip("_") or "report"

    timestamp = __import__("datetime").datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    path = os.path.join(reports_dir, f"{safe_project}_{safe_kind}_{timestamp}.xlsx")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        ws = wb.active
        ws.title = "Summary"

        ws["A1"] = "J.A.R.V.I.S Report"
        ws["A1"].font = Font(bold=True, size=16)

        ws["A3"] = "Project"
        ws["B3"] = project_name
        ws["A4"] = "Report type"
        ws["B4"] = report_kind
        ws["A5"] = "Generated"
        ws["B5"] = timestamp

        ws["A7"] = "Full report preview"
        ws["A7"].font = Font(bold=True)

        lines = str(content).splitlines()

        for row_index, line in enumerate(lines[:400], start=8):
            ws.cell(row=row_index, column=1).value = line

        ws.column_dimensions["A"].width = 120
        ws.column_dimensions["B"].width = 35

        ws2 = wb.create_sheet("Report Lines")
        ws2.append(["Line", "Text"])

        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

        for index, line in enumerate(lines, start=1):
            ws2.append([index, line])

        ws2.column_dimensions["A"].width = 12
        ws2.column_dimensions["B"].width = 140

        ws3 = wb.create_sheet("Risk Matrix")
        ws3.append(["Severity", "Finding"])

        for cell in ws3[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FCE4D6")

        risk_keywords = ["critical", "high", "medium", "low", "risk", "vulnerability", "security"]
        for line in lines:
            lower = line.lower()

            if any(keyword in lower for keyword in risk_keywords):
                if "critical" in lower:
                    severity = "CRITICAL"
                elif "high" in lower:
                    severity = "HIGH"
                elif "medium" in lower:
                    severity = "MEDIUM"
                elif "low" in lower:
                    severity = "LOW"
                else:
                    severity = "INFO"

                ws3.append([severity, line])

        ws3.column_dimensions["A"].width = 16
        ws3.column_dimensions["B"].width = 140

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        wb.save(path)

    except Exception as e:
        # Fallback CSV if openpyxl is missing.
        csv_path = path.replace(".xlsx", ".csv")

        with open(csv_path, "w", encoding="utf-8", errors="ignore", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["project", "report_type", "line", "text"])

            for index, line in enumerate(str(content).splitlines(), start=1):
                writer.writerow([project_name, report_kind, index, line])

        path = csv_path

    if open_after:
        try:
            os.startfile(os.path.abspath(path))
        except Exception:
            pass

    return f"Report created:\n{path}"


def _jarvis_export_report_any(project_name, format_type, report_kind="project"):
    fmt = str(format_type).lower().strip()
    content = _jarvis_build_report_content(project_name, report_kind=report_kind)

    if fmt in {"excel", "xls", "xlsx", "spreadsheet"}:
        return _jarvis_export_excel_report(project_name, content, report_kind=report_kind, open_after=True)

    # Use upgraded project_review_assistant.py if available.
    try:
        from project_review_assistant import export_report_content
        return export_report_content(
            project_name,
            content,
            format_type=fmt,
            report_kind=report_kind,
            open_after=True
        )
    except Exception:
        pass

    # Fallback to embedded developer assistant report generator.
    try:
        ns = globals().get("_EMBEDDED_DEVELOPER_ASSISTANT_NAMESPACE", {})
        creator = ns.get("create_report")

        if creator:
            return creator(
                f"JARVIS {report_kind.title()} - {project_name}",
                content,
                format_type=fmt,
                open_after=True
            )
    except Exception as e:
        return f"Could not export report: {e}"

    return "Could not export report. No report generator available."


def handle_natural_developer_command(command):
    original = _jarvis_norm(command)

    if not original:
        return None

    # Normalize en dash/range hyphen forms.
    original = original.replace("–", "-").replace("—", "-")
    lower = original.lower().strip()

    # read/show lines 1-100 from hud.py in project J.A.R.V.I.S
    match = re.match(
        r"^(read|show|get|give me)\s+(?:the\s+)?(?:code\s+)?lines?\s+(\d+)\s*(?:-|to|until)\s*(\d+)\s+from\s+(?:file\s+)?(.+?)\s+(?:in|from)\s+project\s+(.+)$",
        original,
        flags=re.IGNORECASE
    )

    if match:
        start = int(match.group(2))
        end = int(match.group(3))
        file_query = match.group(4).strip()
        project_name = match.group(5).strip()

        return _jarvis_read_lines_direct(project_name, file_query, start, end)

    # review/improve lines 120-200 from logger.py in project CyberShield AI
    match = re.match(
        r"^(review|analyze|improve|optimize|refactor|explain)\s+lines?\s+(\d+)\s*(?:-|to|until)\s*(\d+)\s+from\s+(?:file\s+)?(.+?)\s+(?:in|from)\s+project\s+(.+)$",
        original,
        flags=re.IGNORECASE
    )

    if match:
        action = match.group(1).lower()
        start = int(match.group(2))
        end = int(match.group(3))
        file_query = match.group(4).strip()
        project_name = match.group(5).strip()

        try:
            if action in {"improve", "optimize", "refactor"}:
                from project_file_assistant import improve_project_file_lines
                return improve_project_file_lines(project_name, file_query, start, end)

            if action == "explain":
                from project_file_assistant import explain_project_file_lines
                return explain_project_file_lines(project_name, file_query, start, end)

            from project_file_assistant import review_project_file_lines
            return review_project_file_lines(project_name, file_query, start, end)
        except Exception:
            try:
                ns = globals().get("_EMBEDDED_DEVELOPER_ASSISTANT_NAMESPACE", {})
                reviewer = ns.get("review_file")
                improver = ns.get("improve_file")

                if action in {"improve", "optimize", "refactor"} and improver:
                    return improver(project_name, file_query, start, end)

                if reviewer:
                    return reviewer(project_name, file_query, start, end)
            except Exception as e:
                return f"Could not {action} lines: {e}"

    # find function handle_command in jarvis_agent.py from project J.A.R.V.I.S
    match = re.match(
        r"^find\s+(function|class|symbol)\s+(.+?)\s+in\s+(?:file\s+)?(.+?)\s+(?:from|in)\s+project\s+(.+)$",
        original,
        flags=re.IGNORECASE
    )

    if match:
        symbol_type = match.group(1).lower()
        symbol_name = match.group(2).strip()
        file_query = match.group(3).strip()
        project_name = match.group(4).strip()

        return _jarvis_find_symbol_direct(project_name, file_query, symbol_name, symbol_type)

    # copy lines 20-80 from logger.py in CyberShield AI to app.py in ManagerApp at line 150
    match = re.match(
        r"^(copy|move)\s+lines?\s+(\d+)\s*(?:-|to|until)\s*(\d+)\s+from\s+(?:file\s+)?(.+?)\s+in\s+(.+?)\s+to\s+(?:file\s+)?(.+?)\s+in\s+(.+?)\s+at\s+line\s+(\d+)$",
        original,
        flags=re.IGNORECASE
    )

    if match:
        source_start = int(match.group(2))
        source_end = int(match.group(3))
        source_file = match.group(4).strip()
        source_project = match.group(5).strip()
        target_file = match.group(6).strip()
        target_project = match.group(7).strip()
        target_start = int(match.group(8))

        return _jarvis_copy_lines_direct(
            source_project,
            source_file,
            source_start,
            source_end,
            target_project,
            target_file,
            target_start,
            None
        )

    # copy/replace lines 20-80 from file a.py in project X to file b.py in project Y lines 100-120
    match = re.match(
        r"^(copy|replace|move)\s+lines?\s+(\d+)\s*(?:-|to|until)\s*(\d+)\s+from\s+(?:file\s+)?(.+?)\s+(?:in|from)\s+project\s+(.+?)\s+to\s+(?:file\s+)?(.+?)\s+(?:in|from)\s+project\s+(.+?)\s+lines?\s+(\d+)\s*(?:-|to|until)\s*(\d+)$",
        original,
        flags=re.IGNORECASE
    )

    if match:
        return _jarvis_copy_lines_direct(
            match.group(5).strip(),
            match.group(4).strip(),
            int(match.group(2)),
            int(match.group(3)),
            match.group(7).strip(),
            match.group(6).strip(),
            int(match.group(8)),
            int(match.group(9))
        )

    # generate Excel/PDF/Word/PPT report for project X
    match = re.match(
        r"^generate\s+(word|doc|docx|pdf|ppt|pptx|powerpoint|presentation|excel|xls|xlsx|spreadsheet|markdown|md|html|json|csv)\s+(?:report|review)\s+for\s+project\s+(.+)$",
        original,
        flags=re.IGNORECASE
    )

    if match:
        fmt = match.group(1).strip()
        project_name = match.group(2).strip()

        return _jarvis_export_report_any(project_name, fmt, report_kind="project_review")

    # generate security/architecture/score Excel report for project X
    match = re.match(
        r"^generate\s+(security|architecture|score)\s+(word|doc|docx|pdf|ppt|pptx|powerpoint|excel|xls|xlsx|markdown|md|html|json|csv)\s+(?:report|review)\s+for\s+project\s+(.+)$",
        original,
        flags=re.IGNORECASE
    )

    if match:
        kind = match.group(1).strip()
        fmt = match.group(2).strip()
        project_name = match.group(3).strip()

        return _jarvis_export_report_any(project_name, fmt, report_kind=kind)

    return None



from screen_reader import (
    read_screen,
    analyze_screen,
    read_screen_center,
    analyze_screen_center,
    read_terminal,
    analyze_terminal,
    read_browser,
    analyze_browser,
    take_screenshot,
    save_screenshot,
    describe_screenshot,
    remember_current_screenshot,
    what_was_on_my_screen,
    search_screenshot_memory,
    read_code_on_screen,
    read_error_on_screen,
    review_code_on_screen,
    find_bugs_on_screen,
    explain_error_on_screen,
    what_error_is_on_screen
)

from screen_memory import load_memory

from memory_manager import (
    remember,
    recent_memories,
    search_memory,
    memory_stats,
    last_project,
    last_file,
    last_audit,
    last_security_report,
    last_task,
    continue_last_task,
    memory_summary
)

from project_memory import (
    remember_project,
    get_project_memory,
    search_project_memory,
    show_remembered_projects,
    project_memory_stats,
    last_project_name,
    last_project as last_project_memory,
    last_project_file,
    last_project_audit,
    last_project_security_report,
    continue_last_project_task,
    project_conversation_summary,
    what_was_i_working_on,
    what_file_did_we_review_last,
    continue_last_audit,
    show_last_security_report
)

from deep_project_memory import (
    remember_deep_project,
    list_deep_projects,
    get_deep_project,
    search_deep_project_code,
    show_project_files,
    show_project_tech_stack,
    deep_project_stats,
    last_deep_project,
    continue_last_deep_project,
    last_project_summary,
    compare_remembered_projects,
    what_projects_do_you_remember,
    what_was_i_working_on_last,
    resume_last_project,
    remember_project_event,
    project_timeline,
    audit_history,
    vulnerability_history,
    remembered_fixes,
    project_evolution,
    session_summary,
    continue_previous_session,
    last_20_audits,
    last_vulnerabilities,
    last_improvements
)

from project_file_assistant import (
    open_project_file,
    read_project_file,
    search_project_files,
    analyze_project_file,
    improve_project_file,
    optimize_project_file,
    preview_project_file,
    create_project_file_backup,
    suggest_safe_project_file_patch,
    apply_safe_project_file_replacement,
    restore_latest_project_file_backup,
    list_project_file_backups,
    apply_ai_safe_project_file_patch
)

from file_memory_assistant import (
    open_memory_file,
    read_memory_file,
    analyze_memory_file,
    review_memory_file,
    improve_memory_file,
    optimize_memory_file,
    security_review_memory_file,
    create_file_backup,
    suggest_safe_patch_for_file,
    apply_safe_full_replacement,
    restore_latest_backup,
    list_file_backups
)

from universal_file_resolver import (
    find_universal_file,
    rank_universal_file,
    open_universal_file,
    read_universal_file,
    analyze_universal_file,
    review_universal_file,
    improve_universal_file,
    optimize_universal_file,
    security_review_universal_file,
    open_best_universal_file,
    read_best_universal_file,
    analyze_best_universal_file,
    review_best_universal_file,
    security_review_best_universal_file,
    open_numbered_file,
    read_numbered_file,
    analyze_numbered_file,
    review_numbered_file,
    security_review_numbered_file,
    improve_numbered_file,
    optimize_numbered_file,
    open_universal_folder,
    rank_universal_folder,
    open_project_folder_universal,
    safe_preview_file
)

from project_review_assistant import (
    review_project,
    find_security_issues,
    find_dead_code,
    find_duplicate_code,
    generate_architecture_report,
    generate_improvement_roadmap,
    optimize_project,
    analyze_project_structure,
    generate_project_documentation,
    generate_documentation,
    project_evidence_report,
    evidence_project,
    grounded_review_project,
    grounded_security_review_project,
    grounded_architecture_review_project,
    grounded_documentation_project,
    grounded_review,
    grounded_security_review,
    grounded_architecture_review,
    grounded_docs,
    strict_grounded_analyzer_project,
    strict_security_analyzer_project,
    strict_architecture_analyzer_project,
    strict_grounded_analyzer,
    strict_security_analyzer,
    strict_architecture_analyzer,
    score_project,
    compare_projects,
    compare_security,
    compare_architecture,
    rank_projects_by_security,
    rank_projects_by_architecture,
    rank_projects_by_maintainability,
    best_project_in_memory,
    export_project_report,
    export_report,
    export_project_markdown_report,
    daily_project_check,
    smart_daily_check,
    daily_check,
    suggest_fixes_for_project,
    suggest_project_fixes,
    project_fixes,
    suggest_fixes,
    find_api_keys,
    find_passwords,
    find_hardcoded_secrets,
    find_sql_injection,
    find_xss_risks,
    find_dangerous_imports,
    full_security_audit,
    generate_security_roadmap,
    autonomous_review_project,
    autonomous_improve_project,
    autonomous_secure_project,
    autonomous_optimize_project,
    autonomous_fix_project,
    review_everything,
    fix_project,
    secure_project,
    improve_project,
    optimize_project_autonomous,
    memory_aware_review,
    memory_aware_security_review,
    project_evolution_report,
    engineering_session_summary,
    generate_project_roadmap,
    next_best_improvements,
    what_should_i_fix_next,
    highest_risk_vulnerabilities,
    estimate_project_maturity,
    estimate_production_readiness,
    generate_sprint_plan,
    generate_release_checklist,
    generate_deployment_checklist,
    become_project_architect,
    generate_roadmap,
    project_roadmap,
    next_improvements,
    fix_next,
    high_risk_vulnerabilities,
    production_readiness,
    project_maturity,
    sprint_plan,
    release_checklist,
    deployment_checklist,
    project_architect
)

from tools import (
    open_website,
    search_files,
    open_installed_app,
    refresh_app_index,
    list_apps,
    open_project,
    list_projects,
    list_projects_detailed,
    list_projects_by_drive,
    search_projects,
    refresh_project_index,
    analyze_project_by_name,
    open_project_in_vscode,
    open_anything,
    open_project_in_app,
    open_file_in_app,
    refresh_all_indexes,
    ensure_indexes_ready as smart_ensure_indexes_ready,
    smart_refresh_all_indexes,
    index_status
)


# ==========================
# HUD CONTEXT HELPERS
# ==========================
HUD_STATUS_FILE = "hud_status.txt"
HUD_COMMAND_FILE = "hud_command.txt"
HUD_RESULT_FILE = "hud_result.txt"
HUD_PROJECT_FILE = "hud_project.txt"
HUD_CURRENT_FILE = "hud_current_file.txt"
HUD_ACTION_FILE = "hud_action.txt"
HUD_AI_STATUS_FILE = "hud_ai_status.txt"

HUD_SECURITY_SCORE_FILE = "hud_security_score.txt"
HUD_PROJECT_SCORE_FILE = "hud_project_score.txt"
HUD_MEMORY_STATUS_FILE = "hud_memory_status.txt"
HUD_VISION_STATUS_FILE = "hud_vision_status.txt"
HUD_OLLAMA_STATUS_FILE = "hud_ollama_status.txt"

# Step 15 - Enterprise Workflow HUD files
HUD_NEXT_TASK_FILE = "hud_next_task.txt"
HUD_ROADMAP_STATUS_FILE = "hud_roadmap_status.txt"
HUD_SPRINT_STATUS_FILE = "hud_sprint_status.txt"
HUD_RELEASE_STATUS_FILE = "hud_release_status.txt"
HUD_DEPLOYMENT_STATUS_FILE = "hud_deployment_status.txt"
HUD_PRODUCTION_STATUS_FILE = "hud_production_status.txt"
HUD_COMMANDER_MODE_FILE = "hud_commander_mode.txt"


def write_hud_file(path, value):
    try:
        with open(path, "w", encoding="utf-8", errors="ignore") as f:
            f.write(str(value))
    except Exception:
        pass


def update_hud_context(
    status=None,
    command=None,
    result=None,
    project=None,
    file_path=None,
    action=None,
    ai_status=None,
    security_score=None,
    project_score=None,
    memory_status=None,
    vision_status=None,
    ollama_status=None,
    next_task=None,
    roadmap_status=None,
    sprint_status=None,
    release_status=None,
    deployment_status=None,
    production_status=None,
    commander_mode=None
):
    if status is not None:
        write_hud_file(HUD_STATUS_FILE, status)

    if command is not None:
        write_hud_file(HUD_COMMAND_FILE, command)

    if result is not None:
        text = str(result).replace("\n", " ").strip()
        write_hud_file(HUD_RESULT_FILE, text[:220])

    if project is not None:
        write_hud_file(HUD_PROJECT_FILE, project)

    if file_path is not None:
        write_hud_file(HUD_CURRENT_FILE, file_path)

    if action is not None:
        write_hud_file(HUD_ACTION_FILE, action)

    if ai_status is not None:
        write_hud_file(HUD_AI_STATUS_FILE, ai_status)

    if security_score is not None:
        write_hud_file(HUD_SECURITY_SCORE_FILE, security_score)

    if project_score is not None:
        write_hud_file(HUD_PROJECT_SCORE_FILE, project_score)

    if memory_status is not None:
        write_hud_file(HUD_MEMORY_STATUS_FILE, memory_status)

    if vision_status is not None:
        write_hud_file(HUD_VISION_STATUS_FILE, vision_status)

    if ollama_status is not None:
        write_hud_file(HUD_OLLAMA_STATUS_FILE, ollama_status)

    if next_task is not None:
        write_hud_file(HUD_NEXT_TASK_FILE, next_task)

    if roadmap_status is not None:
        write_hud_file(HUD_ROADMAP_STATUS_FILE, roadmap_status)

    if sprint_status is not None:
        write_hud_file(HUD_SPRINT_STATUS_FILE, sprint_status)

    if release_status is not None:
        write_hud_file(HUD_RELEASE_STATUS_FILE, release_status)

    if deployment_status is not None:
        write_hud_file(HUD_DEPLOYMENT_STATUS_FILE, deployment_status)

    if production_status is not None:
        write_hud_file(HUD_PRODUCTION_STATUS_FILE, production_status)

    if commander_mode is not None:
        write_hud_file(HUD_COMMANDER_MODE_FILE, commander_mode)



# ==========================
# HUD FUSION HELPERS
# ==========================
def extract_hud_project_score(text):
    text = str(text)
    lower = text.lower()

    patterns = [
        r"overall[^0-9]{0,25}([0-9]+(?:\.[0-9]+)?\s*/\s*10)",
        r"score[^0-9]{0,25}([0-9]+(?:\.[0-9]+)?\s*/\s*10)",
        r"overall[^0-9]{0,25}([0-9]+(?:\.[0-9]+)?\s*%)",
        r"score[^0-9]{0,25}([0-9]+(?:\.[0-9]+)?\s*%)",
    ]

    for pattern in patterns:
        match = re.search(pattern, lower)

        if match:
            return match.group(1).replace(" ", "")

    return "DONE"


def extract_hud_security_score(text):
    text = str(text)
    lower = text.lower()

    patterns = [
        r"security[^0-9]{0,25}([0-9]+(?:\.[0-9]+)?\s*/\s*10)",
        r"security risk level[^a-zA-Z]{0,20}(low|medium|high)",
        r"risk level[^a-zA-Z]{0,20}(low|medium|high)",
    ]

    for pattern in patterns:
        match = re.search(pattern, lower)

        if match:
            return match.group(1).upper().replace(" ", "")

    if "high" in lower and "risk" in lower:
        return "HIGH"

    if "medium" in lower and "risk" in lower:
        return "MEDIUM"

    if "low" in lower and "risk" in lower:
        return "LOW"

    return "DONE"


def hud_start_action(command, action, project=None, file_path=None, thinking=False, vision=False):
    update_hud_context(
        status="PROCESSING",
        command=command,
        project=project,
        file_path=file_path,
        action=action,
        ai_status="THINKING" if thinking else "READY",
        memory_status="SYNC",
        vision_status="ACTIVE" if vision else None,
        ollama_status="THINKING" if thinking else "LOCAL",
        commander_mode="ACTIVE" if project else None
    )


def hud_finish_action(command, result, action, project=None, file_path=None, project_score=None, security_score=None, vision=False):
    update_hud_context(
        status="SUCCESS",
        command=command,
        result=result,
        project=project,
        file_path=file_path,
        action=action,
        ai_status="READY",
        project_score=project_score,
        security_score=security_score,
        memory_status="SYNC",
        vision_status="ACTIVE" if vision else None,
        ollama_status="READY",
        commander_mode="READY" if project else None
    )






# ==========================
# OLLAMA STARTUP AUTO-REPAIR
# ==========================
def startup_ollama_auto_repair():
    """
    Enterprise startup guard.
    Checks Ollama before J.A.R.V.I.S starts accepting commands.
    If Ollama is stuck or the model crashes, llm_local.py repairs it automatically.
    """
    update_hud_context(
        status="PROCESSING",
        command="Startup Ollama check",
        result="Checking Ollama and local model...",
        action="Ollama auto-repair",
        ai_status="CHECKING",
        ollama_status="CHECKING",
        commander_mode="STARTUP",
        next_task="Verify local AI engine"
    )

    if ensure_ollama_ready is None:
        update_hud_context(
            status="ERROR",
            result="ensure_ollama_ready not available in llm_local.py",
            action="Ollama auto-repair unavailable",
            ai_status="ERROR",
            ollama_status="ERROR"
        )
        return False, "ensure_ollama_ready not available in llm_local.py"

    try:
        ready, message = ensure_ollama_ready(auto_repair=True)

        update_hud_context(
            status="SUCCESS" if ready else "ERROR",
            result=message,
            action="Ollama ready" if ready else "Ollama repair failed",
            ai_status="READY" if ready else "ERROR",
            ollama_status="READY" if ready else "ERROR",
            commander_mode="READY" if ready else "ERROR",
            next_task="Local AI ready" if ready else "Check Ollama manually"
        )

        return ready, message

    except Exception as e:
        update_hud_context(
            status="ERROR",
            result=f"Ollama startup check failed: {e}",
            action="Ollama startup exception",
            ai_status="ERROR",
            ollama_status="ERROR",
            commander_mode="ERROR"
        )
        return False, str(e)





# ==========================================================
# PROJECT / FILE COMMAND ROUTER
# Fixes commands like:
# - find project CyberShield AI
# - where is project CyberShield AI
# - list projects
# - show projects
# - find file app.py in project CyberShield AI
# - where is file app.py in project CyberShield AI
# ==========================================================
def _jarvis_project_to_text(project):
    if not project:
        return "Project not found."

    if isinstance(project, dict):
        name = project.get("name", "Unknown")
        path = project.get("path", "Unknown path")
        project_type = project.get("type", project.get("project_type", "Unknown"))

        return (
            f"Project found:\n"
            f"Name: {name}\n"
            f"Type: {project_type}\n"
            f"Path: {path}"
        )

    return str(project)


def _jarvis_find_project_command(project_name):
    try:
        from tools import find_project as tools_find_project
        project = tools_find_project(project_name)

        if project:
            result = _jarvis_project_to_text(project)

            try:
                update_hud_context(
                    status="SUCCESS",
                    command=f"find project {project_name}",
                    result=result,
                    project=project.get("name", project_name) if isinstance(project, dict) else project_name,
                    action="Find project",
                    ai_status="READY",
                    ollama_status="READY",
                    commander_mode="READY"
                )
            except Exception:
                pass

            return result

    except Exception as e:
        return f"Could not search project: {e}"

    return f"Project not found: {project_name}"


def _jarvis_list_projects_command(detailed=False):
    try:
        if detailed:
            result = list_projects_detailed()
        else:
            result = list_projects()

        if isinstance(result, list):
            if not result:
                return "No indexed projects found."

            return "Indexed projects:\n" + "\n".join(f"- {item}" for item in result)

        return str(result)

    except Exception as e:
        return f"Could not list projects: {e}"


def _jarvis_search_projects_command(query):
    try:
        result = search_projects(query)

        if isinstance(result, list):
            if not result:
                return f"No projects found for: {query}"

            lines = [f"Projects matching '{query}':"]

            for item in result:
                if isinstance(item, dict):
                    lines.append(f"- {item.get('name', 'Unknown')} -> {item.get('path', 'Unknown path')}")
                else:
                    lines.append(f"- {item}")

            return "\n".join(lines)

        return str(result)

    except Exception as e:
        return f"Could not search projects: {e}"


def _jarvis_find_file_command(file_query, project_name):
    try:
        from project_file_assistant import find_project_file
        item, error = find_project_file(project_name, file_query)

        if error:
            return error

        if item:
            path = item.get("full_path", "")
            rel = item.get("relative_path", file_query)

            result = (
                f"File found:\n"
                f"Project: {project_name}\n"
                f"File: {rel}\n"
                f"Path: {path}"
            )

            try:
                update_hud_context(
                    status="SUCCESS",
                    command=f"find file {file_query} in project {project_name}",
                    result=result,
                    project=project_name,
                    file_path=rel,
                    action="Find file",
                    ai_status="READY",
                    ollama_status="READY",
                    commander_mode="READY"
                )
            except Exception:
                pass

            return result

    except Exception:
        pass

    try:
        path, error = _jarvis_parse_project_file(project_name, file_query)

        if error:
            return error

        return (
            f"File found:\n"
            f"Project: {project_name}\n"
            f"File: {file_query}\n"
            f"Path: {path}"
        )

    except Exception as e:
        return f"Could not find file: {e}"


def handle_project_lookup_command(command):
    original = str(command).strip()
    lower = original.lower().strip()

    if lower in {"list projects", "show projects", "projects"}:
        return _jarvis_list_projects_command(detailed=False)

    if lower in {"list projects detailed", "show projects detailed", "detailed projects"}:
        return _jarvis_list_projects_command(detailed=True)

    if lower in {"refresh projects", "refresh project index", "reindex projects"}:
        try:
            return refresh_project_index()
        except Exception as e:
            return f"Could not refresh project index: {e}"

    if lower in {"index status", "indexes status", "project index status"}:
        try:
            return index_status()
        except Exception as e:
            return f"Could not read index status: {e}"

    match = re.match(
        r"^(find|where is|locate|search)\s+project\s+(.+)$",
        original,
        flags=re.IGNORECASE
    )

    if match:
        project_name = match.group(2).strip()

        if match.group(1).lower() == "search":
            return _jarvis_search_projects_command(project_name)

        return _jarvis_find_project_command(project_name)

    match = re.match(
        r"^project\s+(.+?)\s+(location|path|where)$",
        original,
        flags=re.IGNORECASE
    )

    if match:
        return _jarvis_find_project_command(match.group(1).strip())

    match = re.match(
        r"^(where is|find|locate|search)\s+file\s+(.+?)\s+(?:in|from)\s+project\s+(.+)$",
        original,
        flags=re.IGNORECASE
    )

    if match:
        file_query = match.group(2).strip()
        project_name = match.group(3).strip()
        return _jarvis_find_file_command(file_query, project_name)

    match = re.match(
        r"^list\s+files\s+(?:in|from)\s+project\s+(.+)$",
        original,
        flags=re.IGNORECASE
    )

    if match:
        project_name = match.group(1).strip()

        try:
            return show_project_files_local(project_name)
        except Exception as e:
            return f"Could not list project files: {e}"

    match = re.match(
        r"^show\s+structure\s+(?:for|of)\s+project\s+(.+)$",
        original,
        flags=re.IGNORECASE
    )

    if match:
        project_name = match.group(1).strip()

        try:
            return show_project_structure_local(project_name)
        except Exception as e:
            return f"Could not show project structure: {e}"

    match = re.match(
        r"^project\s+statistics\s+(.+)$",
        original,
        flags=re.IGNORECASE
    )

    if match:
        project_name = match.group(1).strip()

        try:
            return show_project_statistics_local(project_name)
        except Exception as e:
            return f"Could not show project statistics: {e}"

    return None




# ==========================================================
# ENTERPRISE NATURAL COMMAND LAYER
# Makes J.A.R.V.I.S understand natural commands like:
# - could you open firefox
# - open cyber shield in vs code
# - create a pdf report for cyber
# - review it and export word
# - open the last report
# ==========================================================
JARVIS_CONTEXT_FILE = "jarvis_context.json"


def _safe_json_load(path, default=None):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as file:
            data = json.load(file)

        return data

    except Exception:
        return default if default is not None else {}


def _safe_json_save(path, data):
    try:
        with open(path, "w", encoding="utf-8") as file:
            json.dump(data, file, indent=2)

        return True

    except Exception:
        return False


def _ctx_get(key, default=None):
    data = _safe_json_load(JARVIS_CONTEXT_FILE, {})
    return data.get(key, default)


def _ctx_set(**kwargs):
    data = _safe_json_load(JARVIS_CONTEXT_FILE, {})

    for key, value in kwargs.items():
        if value is not None:
            data[key] = value

    return _safe_json_save(JARVIS_CONTEXT_FILE, data)


def _clean_natural_command(text):
    text = str(text or "").strip()
    text = text.replace("’", "'").replace("`", "'")
    text = re.sub(r"\s+", " ", text)

    lower = text.lower().strip()

    polite_prefixes = [
        "jarvis ",
        "hey jarvis ",
        "please ",
        "can you ",
        "could you ",
        "would you ",
        "i need you to ",
        "i want you to ",
        "please can you ",
        "please could you ",
        "for me ",
    ]

    changed = True

    while changed:
        changed = False

        for prefix in polite_prefixes:
            if lower.startswith(prefix):
                text = text[len(prefix):].strip()
                lower = text.lower().strip()
                changed = True

    filler = [
        "please",
        "for me",
        "right now",
        "now",
        "the",
        "my",
    ]

    # Remove harmless filler only at the beginning.
    for word in filler:
        if lower.startswith(word + " "):
            text = text[len(word):].strip()
            lower = text.lower().strip()

    return text.strip()


def _normalize_project_spoken_name(text):
    lower = str(text or "").lower().strip()

    aliases = {
        "cyber": "CyberShield AI",
        "cyber shield": "CyberShield AI",
        "cyber shield ai": "CyberShield AI",
        "cybershield": "CyberShield AI",
        "cybershield ai": "CyberShield AI",
        "cybers in the": "CyberShield AI",
        "jarvis": "J.A.R.V.I.S",
        "jervis": "J.A.R.V.I.S",
        "j a r v i s": "J.A.R.V.I.S",
        "j.a.r.v.i.s": "J.A.R.V.I.S",
        "manager app": "ManagerApp",
        "managerapp": "ManagerApp",
    }

    if lower in aliases:
        return aliases[lower]

    # Keep original title if no alias matched.
    return str(text).strip()


def _normalize_app_spoken_name(text):
    lower = str(text or "").lower().strip()

    aliases = {
        "fire": "firefox",
        "fox": "firefox",
        "mozilla": "firefox",
        "mozilla firefox": "firefox",
        "browser": "chrome",
        "google browser": "chrome",
        "google chrome": "chrome",
        "vs": "vscode",
        "vs code": "vscode",
        "visual code": "vscode",
        "visual studio code": "vscode",
        "visual studio": "visual studio community",
        "vs community": "visual studio community",
        "idea": "intellij",
        "intellij idea": "intellij",
        "files": "file explorer",
        "file manager": "file explorer",
    }

    return aliases.get(lower, lower)


def _extract_report_format(text):
    lower = str(text).lower()

    if any(word in lower for word in ["word", "docx", "doc"]):
        return "word"

    if "pdf" in lower:
        return "pdf"

    if any(word in lower for word in ["powerpoint", "ppt", "pptx", "presentation"]):
        return "powerpoint"

    if any(word in lower for word in ["excel", "xlsx", "xls", "spreadsheet"]):
        return "excel"

    if "html" in lower:
        return "html"

    if "markdown" in lower or re.search(r"\bmd\b", lower):
        return "markdown"

    return "pdf"


def _extract_project_from_text(text, default_last=True):
    original = str(text or "").strip()
    lower = original.lower()

    # Direct alias detection anywhere in the command.
    for alias in [
        "cyber shield ai",
        "cybershield ai",
        "cyber shield",
        "cyber",
        "j.a.r.v.i.s",
        "j a r v i s",
        "jarvis",
        "manager app",
        "managerapp",
    ]:
        if re.search(rf"\b{re.escape(alias)}\b", lower):
            return _normalize_project_spoken_name(alias)

    # Common "for project X", "from project X"
    match = re.search(
        r"\b(?:for|from|in|of)\s+(?:project\s+)?(.+?)(?:\s+as\s+|\s+and\s+|\s+then\s+|$)",
        original,
        flags=re.IGNORECASE
    )

    if match:
        candidate = match.group(1).strip(" .,:;")
        candidate = re.sub(r"\b(word|pdf|powerpoint|ppt|excel|report|review|security|architecture|score)\b", "", candidate, flags=re.IGNORECASE).strip()

        if candidate:
            return _normalize_project_spoken_name(candidate)

    if default_last:
        return _ctx_get("last_project", "CyberShield AI")

    return None


def _friendly_result(result):
    text = str(result)

    # Keep generated report/file paths because user may need them,
    # but opening commands should stay clean.
    if text.lower().startswith(("opening ", "could not", "project not found", "file not found", "website target")):
        return text

    if "Opening project" in text and ":" in text:
        text = re.sub(r":\s*[A-Z]:\\.*$", "", text)

    if "Opening file in" in text and ":" in text:
        text = re.sub(r":\s*[A-Z]:\\.*$", "", text)

    return text


def _open_natural_target(target):
    target = str(target or "").strip()

    if not target:
        return "Missing target."

    target = _normalize_app_spoken_name(target)

    try:
        result = open_anything(target)
        return _friendly_result(result)
    except Exception:
        pass

    try:
        return open_installed_app(target)
    except Exception as e:
        return f"Could not open {_normalize_app_spoken_name(target)}: {e}"


def _handle_natural_open_command(original):
    text = _clean_natural_command(original)
    lower = text.lower()

    # Open project X in IDE.
    match = re.match(
        r"^(?:open|launch|start)\s+(?:project\s+)?(.+?)\s+(?:in|with)\s+(vs code|vscode|visual studio code|visual studio|visual studio community|vs community|intellij|intellij idea|idea|pycharm|android studio|eclipse)$",
        text,
        flags=re.IGNORECASE
    )

    if match:
        project_name = _normalize_project_spoken_name(match.group(1).strip())
        app_name = _normalize_app_spoken_name(match.group(2).strip())

        _ctx_set(last_project=project_name, last_ide=app_name)

        try:
            result = open_project_in_app(project_name, app_name)
        except Exception:
            result = open_project(project_name)

        return _friendly_result(result)

    # Open a project.
    match = re.match(
        r"^(?:open|launch|start)\s+(?:the\s+)?(?:project\s+)?(.+?)$",
        text,
        flags=re.IGNORECASE
    )

    if match:
        target = match.group(1).strip()

        if target.lower() in {"it", "this", "that", "project"}:
            target = _ctx_get("last_project", "CyberShield AI")

        project_name = _normalize_project_spoken_name(target)

        # If it looks like project alias, use project opener.
        if project_name in {"CyberShield AI", "J.A.R.V.I.S", "ManagerApp"} or "project" in lower:
            _ctx_set(last_project=project_name)

            try:
                return _friendly_result(open_project(project_name))
            except Exception:
                pass

        return _open_natural_target(target)

    return None


def _handle_natural_report_command(original):
    text = _clean_natural_command(original)
    lower = text.lower()

    if "report" not in lower and "presentation" not in lower and "review" not in lower:
        return None

    if not any(word in lower for word in [
        "create", "generate", "make", "export", "build", "review", "security", "architecture", "score", "pdf", "word", "powerpoint", "excel"
    ]):
        return None

    fmt = _extract_report_format(lower)
    project_name = _extract_project_from_text(text, default_last=True)

    if not project_name:
        project_name = "CyberShield AI"

    report_kind = "project_review"

    if "security" in lower or "audit" in lower:
        report_kind = "security"

    if "architecture" in lower or "architect" in lower:
        report_kind = "architecture"

    if "score" in lower:
        report_kind = "score"

    _ctx_set(last_project=project_name, last_report_format=fmt, last_report_kind=report_kind)

    try:
        return _jarvis_export_report_any(project_name, fmt, report_kind=report_kind)
    except Exception as e:
        return f"Could not generate {fmt} report for {project_name}: {e}"


def _handle_contextual_followup(original):
    text = _clean_natural_command(original)
    lower = text.lower()

    project_name = _ctx_get("last_project", "CyberShield AI")

    if lower in {"review it", "review this", "review the project", "analyze it", "analyze this"}:
        return review_project(project_name)

    if lower in {"secure it", "audit it", "security review it", "scan it"}:
        return full_security_audit(project_name)

    if lower in {"score it", "score this", "score the project"}:
        return score_project(project_name)

    if lower in {"make pdf", "create pdf", "export pdf", "make a pdf", "create a pdf"}:
        return _jarvis_export_report_any(project_name, "pdf", report_kind=_ctx_get("last_report_kind", "project_review"))

    if lower in {"make word", "create word", "export word", "make a word report", "create a word report"}:
        return _jarvis_export_report_any(project_name, "word", report_kind=_ctx_get("last_report_kind", "project_review"))

    if lower in {"make powerpoint", "create powerpoint", "export powerpoint", "make presentation", "create presentation"}:
        return _jarvis_export_report_any(project_name, "powerpoint", report_kind=_ctx_get("last_report_kind", "project_review"))

    if lower in {"make excel", "create excel", "export excel", "make spreadsheet", "create spreadsheet"}:
        return _jarvis_export_report_any(project_name, "excel", report_kind=_ctx_get("last_report_kind", "project_review"))

    return None


def handle_enterprise_natural_command(command):
    original = str(command or "").strip()

    if not original:
        return None

    original = original.replace("–", "-").replace("—", "-")

    # Contextual follow-ups should run first.
    result = _handle_contextual_followup(original)

    if result is not None:
        return result

    # Reports / document generation.
    result = _handle_natural_report_command(original)

    if result is not None:
        return result

    # Open / launch / start.
    lower = _clean_natural_command(original).lower()

    if lower.startswith(("open ", "launch ", "start ")):
        result = _handle_natural_open_command(original)

        if result is not None:
            return result

    # "I need VS Code", "I need Chrome"
    match = re.match(r"^(?:i need|give me|bring up|show me)\s+(.+)$", lower, flags=re.IGNORECASE)

    if match:
        return _open_natural_target(match.group(1).strip())

    return None


def parse_project_file_args(text):
    """
    Parses commands like:
    <project name> <file name>

    It supports project names with spaces by matching the longest
    remembered/indexed project name prefix.
    """

    text = text.strip()

    if not text:
        return None, None, "Missing project and file."

    projects = list_projects()

    best_match = None

    for project in projects:
        if text.lower().startswith(project.lower() + " "):
            if best_match is None or len(project) > len(best_match):
                best_match = project

    if best_match:
        file_query = text[len(best_match):].strip()

        if not file_query:
            return None, None, "Missing file name."

        return best_match, file_query, None

    parts = text.split(" ", 1)

    if len(parts) < 2:
        return None, None, "Usage: <project name> <file name>"

    return parts[0], parts[1], None


def parse_file_from_project_text(text):
    """
    Supports commands like:
    app.py from CyberShield AI
    auth.py from cyber
    CyberShield AI app.py
    """

    text = str(text).strip()

    if not text:
        return None, None, "Missing file and project."

    lower = text.lower()

    if " from " in lower:
        index = lower.rfind(" from ")
        file_query = text[:index].strip()
        project_name = text[index + len(" from "):].strip()

        if not file_query:
            return None, None, "Missing file name."

        if not project_name:
            return None, None, "Missing project name."

        return project_name, file_query, None

    return parse_project_file_args(text)




# ==========================
# PROJECT ASSISTANT INSPECTION HELPERS
# ==========================
PROJECT_ASSISTANT_SKIP_DIRS = {
    "node_modules",
    "venv",
    ".venv",
    "jarvis-env",
    "__pycache__",
    ".git",
    ".idea",
    ".vscode",
    "dist",
    "build",
    ".next",
    ".cache",
    "site-packages",
    "file_backups",
}

PROJECT_ASSISTANT_TEXT_EXTS = {
    ".py", ".js", ".jsx", ".ts", ".tsx",
    ".html", ".css", ".scss", ".sass",
    ".json", ".md", ".txt", ".yml", ".yaml",
    ".env", ".ini", ".cfg", ".toml",
    ".xml", ".java", ".c", ".cpp", ".h", ".hpp",
    ".cs", ".php", ".rb", ".go", ".rs",
    ".sql", ".bat", ".ps1", ".sh",
}


def find_project_for_assistant(project_name):
    try:
        from tools import find_project
        project = find_project(project_name)
    except Exception:
        project = None

    if not project:
        return None, f"Project not found: {project_name}"

    path = project.get("path", "")

    if not path or not os.path.exists(path):
        return None, f"Project path not found on disk: {path}"

    return project, None


def iter_project_files_for_assistant(project_path):
    for root, dirs, files in os.walk(project_path):
        dirs[:] = [
            d for d in dirs
            if d.lower() not in PROJECT_ASSISTANT_SKIP_DIRS
        ]

        for filename in files:
            full_path = os.path.join(root, filename)

            try:
                size = os.path.getsize(full_path)
            except Exception:
                size = 0

            yield {
                "name": filename,
                "path": full_path,
                "relative_path": os.path.relpath(full_path, project_path),
                "size": size,
                "extension": os.path.splitext(filename)[1].lower(),
            }


def show_project_structure_local(project_name, max_depth=3, max_items=160):
    project, error = find_project_for_assistant(project_name)

    if error:
        return error

    project_path = project["path"]
    base_depth = len(Path(project_path).parts)

    output = [
        f"Project structure: {project.get('name', project_name)}",
        f"Path: {project_path}",
        ""
    ]

    shown = 0

    for root, dirs, files in os.walk(project_path):
        dirs[:] = [
            d for d in dirs
            if d.lower() not in PROJECT_ASSISTANT_SKIP_DIRS
        ]

        depth = len(Path(root).parts) - base_depth

        if depth > max_depth:
            dirs[:] = []
            continue

        indent = "  " * depth
        folder_name = os.path.basename(root) or root

        output.append(f"{indent}{folder_name}/")
        shown += 1

        file_indent = "  " * (depth + 1)

        for filename in sorted(files)[:12]:
            if shown >= max_items:
                output.append("... output truncated")
                return "\n".join(output)

            output.append(f"{file_indent}{filename}")
            shown += 1

    return "\n".join(output)


def show_project_files_local(project_name, limit=140):
    project, error = find_project_for_assistant(project_name)

    if error:
        return error

    files = list(iter_project_files_for_assistant(project["path"]))
    files.sort(key=lambda item: item["relative_path"].lower())

    output = [
        f"Project files: {project.get('name', project_name)}",
        f"Total files shown: {min(len(files), limit)} / {len(files)}",
        ""
    ]

    for item in files[:limit]:
        output.append(f"- {item['relative_path']}")

    if len(files) > limit:
        output.append(f"... and {len(files) - limit} more files")

    return "\n".join(output)


def show_project_statistics_local(project_name):
    project, error = find_project_for_assistant(project_name)

    if error:
        return error

    files = list(iter_project_files_for_assistant(project["path"]))

    total_size = sum(item["size"] for item in files)
    text_files = 0
    by_ext = {}

    for item in files:
        ext = item["extension"] or "[no extension]"
        by_ext[ext] = by_ext.get(ext, 0) + 1

        if item["extension"] in PROJECT_ASSISTANT_TEXT_EXTS:
            text_files += 1

    top_ext = sorted(
        by_ext.items(),
        key=lambda item: item[1],
        reverse=True
    )[:15]

    output = [
        f"Project statistics: {project.get('name', project_name)}",
        f"Path: {project['path']}",
        f"Total files: {len(files)}",
        f"Text/code files: {text_files}",
        f"Approx total size: {round(total_size / 1024 / 1024, 2)} MB",
        "",
        "Top file types:"
    ]

    for ext, count in top_ext:
        output.append(f"- {ext}: {count}")

    return "\n".join(output)


def show_largest_project_files_local(project_name, limit=25):
    project, error = find_project_for_assistant(project_name)

    if error:
        return error

    files = list(iter_project_files_for_assistant(project["path"]))
    files.sort(key=lambda item: item["size"], reverse=True)

    output = [
        f"Largest files: {project.get('name', project_name)}",
        ""
    ]

    for item in files[:limit]:
        mb = round(item["size"] / 1024 / 1024, 2)
        output.append(f"- {item['relative_path']} -> {mb} MB")

    return "\n".join(output)


def show_all_backups():
    output = []

    try:
        output.append("General file backups:")
        output.append(list_file_backups())
    except Exception as e:
        output.append(f"General backup error: {e}")

    output.append("")

    try:
        output.append("Project file backups:")
        output.append(list_project_file_backups())
    except Exception as e:
        output.append(f"Project backup error: {e}")

    return "\n".join(output)




# ==========================
# STEP 15 - AUTONOMOUS WORKFLOW ENGINE
# Safe orchestration. Generates consolidated reports only.
# ==========================
def workflow_section(title, content, max_chars=9000):
    text = str(content).strip()

    if len(text) > max_chars:
        text = text[:max_chars] + "\n... section truncated"

    return (
        "\n\n"
        + "=" * 76
        + f"\n{title}\n"
        + "=" * 76
        + "\n"
        + text
    )


def set_workflow_hud_stage(
    command,
    project_name,
    stage,
    next_task=None,
    roadmap_status=None,
    sprint_status=None,
    release_status=None,
    deployment_status=None,
    production_status=None
):
    update_hud_context(
        status="PROCESSING",
        command=command,
        project=project_name,
        action=stage,
        ai_status="THINKING",
        memory_status="ACTIVE",
        ollama_status="THINKING",
        commander_mode="WORKFLOW",
        next_task=next_task or stage,
        roadmap_status=roadmap_status,
        sprint_status=sprint_status,
        release_status=release_status,
        deployment_status=deployment_status,
        production_status=production_status
    )


def run_workflow_step(title, func, project_name):
    try:
        return workflow_section(
            title,
            func(project_name)
        )
    except Exception as e:
        return workflow_section(
            title,
            f"Step failed safely: {e}"
        )


def autonomous_project_workflow(project_name, workflow_name="full"):
    command = f"{workflow_name} workflow {project_name}"

    set_workflow_hud_stage(
        command,
        project_name,
        "Workflow started",
        next_task="Score project",
        roadmap_status="PROCESSING",
        sprint_status="PENDING",
        release_status="PENDING",
        deployment_status="PENDING",
        production_status="PENDING"
    )

    report = [
        "JARVIS AUTONOMOUS WORKFLOW REPORT",
        "Mode: safe orchestration / no automatic code changes",
        f"Workflow: {workflow_name}",
        f"Project: {project_name}",
        "",
        "Executed stages:",
        "1. Project score",
        "2. Security audit",
        "3. Architecture review",
        "4. Memory-aware review",
        "5. Project roadmap",
        "6. Sprint plan",
        "7. Production readiness",
        "8. Release checklist",
        "9. Deployment checklist",
        "10. Next fixes",
    ]

    set_workflow_hud_stage(command, project_name, "Scoring project", next_task="Security audit")
    report.append(
        run_workflow_step(
            "1. PROJECT SCORE",
            score_project,
            project_name
        )
    )

    set_workflow_hud_stage(command, project_name, "Running security audit", next_task="Architecture review")
    report.append(
        run_workflow_step(
            "2. FULL SECURITY AUDIT",
            full_security_audit,
            project_name
        )
    )

    set_workflow_hud_stage(command, project_name, "Reviewing architecture", next_task="Memory-aware review")
    report.append(
        run_workflow_step(
            "3. ARCHITECTURE REVIEW",
            strict_architecture_analyzer_project,
            project_name
        )
    )

    set_workflow_hud_stage(command, project_name, "Running memory-aware review", next_task="Project roadmap")
    report.append(
        run_workflow_step(
            "4. MEMORY-AWARE REVIEW",
            memory_aware_review,
            project_name
        )
    )

    set_workflow_hud_stage(
        command,
        project_name,
        "Generating roadmap",
        next_task="Sprint plan",
        roadmap_status="ACTIVE"
    )
    report.append(
        run_workflow_step(
            "5. PROJECT ROADMAP",
            generate_project_roadmap,
            project_name
        )
    )

    set_workflow_hud_stage(
        command,
        project_name,
        "Generating sprint plan",
        next_task="Production readiness",
        roadmap_status="DONE",
        sprint_status="ACTIVE"
    )
    report.append(
        run_workflow_step(
            "6. SPRINT PLAN",
            generate_sprint_plan,
            project_name
        )
    )

    set_workflow_hud_stage(
        command,
        project_name,
        "Estimating production readiness",
        next_task="Release checklist",
        sprint_status="DONE",
        production_status="ACTIVE"
    )
    report.append(
        run_workflow_step(
            "7. PRODUCTION READINESS",
            estimate_production_readiness,
            project_name
        )
    )

    set_workflow_hud_stage(
        command,
        project_name,
        "Generating release checklist",
        next_task="Deployment checklist",
        production_status="DONE",
        release_status="ACTIVE"
    )
    report.append(
        run_workflow_step(
            "8. RELEASE CHECKLIST",
            generate_release_checklist,
            project_name
        )
    )

    set_workflow_hud_stage(
        command,
        project_name,
        "Generating deployment checklist",
        next_task="Next best fixes",
        release_status="DONE",
        deployment_status="ACTIVE"
    )
    report.append(
        run_workflow_step(
            "9. DEPLOYMENT CHECKLIST",
            generate_deployment_checklist,
            project_name
        )
    )

    set_workflow_hud_stage(
        command,
        project_name,
        "Finding next fixes",
        next_task="Workflow complete",
        deployment_status="DONE"
    )
    report.append(
        run_workflow_step(
            "10. NEXT BEST FIXES",
            what_should_i_fix_next,
            project_name
        )
    )

    final_report = "\n".join(report)

    try:
        remember_project_event(
            project_name,
            "workflow",
            f"Autonomous workflow: {workflow_name}",
            final_report[:16000],
            tags=["workflow", "autonomous", "orchestration", workflow_name],
            metadata={"source": "jarvis_agent_step15"}
        )
    except Exception:
        pass

    update_hud_context(
        status="SUCCESS",
        command=command,
        result=final_report,
        project=project_name,
        action="Workflow completed",
        ai_status="READY",
        memory_status="SYNC",
        ollama_status="READY",
        commander_mode="READY",
        next_task="Review workflow report",
        roadmap_status="DONE",
        sprint_status="DONE",
        release_status="DONE",
        deployment_status="DONE",
        production_status="DONE",
        project_score=extract_hud_project_score(final_report),
        security_score=extract_hud_security_score(final_report)
    )

    return final_report


def review_project_workflow(project_name):
    return autonomous_project_workflow(project_name, "review")


def secure_project_workflow(project_name):
    command = f"secure workflow {project_name}"

    set_workflow_hud_stage(
        command,
        project_name,
        "Security workflow started",
        next_task="Full security audit",
        roadmap_status="PENDING",
        production_status="PENDING"
    )

    sections = [
        "JARVIS SECURITY WORKFLOW REPORT",
        "Mode: safe security orchestration / no automatic code changes",
        f"Project: {project_name}",
        run_workflow_step("1. FULL SECURITY AUDIT", full_security_audit, project_name),
        run_workflow_step("2. HIGHEST RISK VULNERABILITIES", highest_risk_vulnerabilities, project_name),
        run_workflow_step("3. SECURITY ROADMAP", generate_security_roadmap, project_name),
        run_workflow_step("4. PRODUCTION READINESS", estimate_production_readiness, project_name),
        run_workflow_step("5. NEXT SECURITY FIXES", what_should_i_fix_next, project_name),
    ]

    result = "\n".join(sections)

    try:
        remember_project_event(
            project_name,
            "workflow",
            "Security workflow",
            result[:16000],
            tags=["workflow", "security", "audit"],
            metadata={"source": "jarvis_agent_step15"}
        )
    except Exception:
        pass

    update_hud_context(
        status="SUCCESS",
        command=command,
        result=result,
        project=project_name,
        action="Security workflow completed",
        ai_status="READY",
        memory_status="SYNC",
        ollama_status="READY",
        commander_mode="READY",
        next_task="Fix highest-risk security issues",
        roadmap_status="DONE",
        production_status="DONE",
        security_score=extract_hud_security_score(result)
    )

    return result


def prepare_project_for_production_workflow(project_name):
    command = f"prepare production workflow {project_name}"

    set_workflow_hud_stage(
        command,
        project_name,
        "Production workflow started",
        next_task="Production readiness",
        release_status="PENDING",
        deployment_status="PENDING",
        production_status="ACTIVE"
    )

    sections = [
        "JARVIS PRODUCTION PREPARATION WORKFLOW",
        "Mode: safe production planning / no automatic deployment",
        f"Project: {project_name}",
        run_workflow_step("1. PRODUCTION READINESS", estimate_production_readiness, project_name),
        run_workflow_step("2. FULL SECURITY AUDIT", full_security_audit, project_name),
        run_workflow_step("3. RELEASE CHECKLIST", generate_release_checklist, project_name),
        run_workflow_step("4. DEPLOYMENT CHECKLIST", generate_deployment_checklist, project_name),
        run_workflow_step("5. SPRINT PLAN", generate_sprint_plan, project_name),
        run_workflow_step("6. NEXT FIXES BEFORE RELEASE", what_should_i_fix_next, project_name),
    ]

    result = "\n".join(sections)

    try:
        remember_project_event(
            project_name,
            "workflow",
            "Production preparation workflow",
            result[:16000],
            tags=["workflow", "production", "release", "deployment"],
            metadata={"source": "jarvis_agent_step15"}
        )
    except Exception:
        pass

    update_hud_context(
        status="SUCCESS",
        command=command,
        result=result,
        project=project_name,
        action="Production workflow completed",
        ai_status="READY",
        memory_status="SYNC",
        ollama_status="READY",
        commander_mode="READY",
        next_task="Complete release blockers",
        release_status="DONE",
        deployment_status="DONE",
        production_status="DONE",
        project_score=extract_hud_project_score(result),
        security_score=extract_hud_security_score(result)
    )

    return result


def release_project_workflow(project_name):
    command = f"release workflow {project_name}"

    set_workflow_hud_stage(
        command,
        project_name,
        "Release workflow started",
        next_task="Release checklist",
        release_status="ACTIVE",
        deployment_status="PENDING"
    )

    sections = [
        "JARVIS RELEASE WORKFLOW",
        "Mode: release planning / no automatic deployment",
        f"Project: {project_name}",
        run_workflow_step("1. RELEASE CHECKLIST", generate_release_checklist, project_name),
        run_workflow_step("2. DEPLOYMENT CHECKLIST", generate_deployment_checklist, project_name),
        run_workflow_step("3. PRODUCTION READINESS", estimate_production_readiness, project_name),
        run_workflow_step("4. SECURITY AUDIT", full_security_audit, project_name),
        run_workflow_step("5. FINAL NEXT FIXES", what_should_i_fix_next, project_name),
    ]

    result = "\n".join(sections)

    try:
        remember_project_event(
            project_name,
            "workflow",
            "Release workflow",
            result[:16000],
            tags=["workflow", "release", "deployment"],
            metadata={"source": "jarvis_agent_step15"}
        )
    except Exception:
        pass

    update_hud_context(
        status="SUCCESS",
        command=command,
        result=result,
        project=project_name,
        action="Release workflow completed",
        ai_status="READY",
        memory_status="SYNC",
        ollama_status="READY",
        commander_mode="READY",
        next_task="Review release checklist",
        release_status="DONE",
        deployment_status="DONE",
        production_status="DONE",
        security_score=extract_hud_security_score(result)
    )

    return result



# ==========================
# STRICT COMMAND ROUTER HELPERS
# ==========================
SAFE_COMMAND_PREFIXES = (
    "generate excel report ",
    "generate xls report ",
    "generate xlsx report ",
    "generate html report ",
    "generate json report ",
    "generate csv report ",
    "generate security excel report ",
    "generate architecture excel report ",
    "generate score excel report ",
    # Developer Assistant embedded commands
    "open project ",
    "open file ",
    "show lines ",
    "read lines ",
    "get lines ",
    "give me lines ",
    "give me code ",
    "find function ",
    "find class ",
    "find symbol ",
    "copy lines ",
    "replace lines ",
    "move lines ",
    "review lines ",
    "review file ",
    "improve lines ",
    "optimize lines ",
    "refactor lines ",
    "improve file ",
    "optimize file ",
    "refactor file ",
    "generate word report ",
    "generate docx report ",
    "generate pdf report ",
    "generate ppt report ",
    "generate pptx report ",
    "generate powerpoint report ",
    "generate markdown report ",
    "format file ",
    "clean file ",
    # Developer Assistant commands
    "open project ",
    "open file ",
    "show lines ",
    "read lines ",
    "get lines ",
    "give me lines ",
    "give me code ",
    "find function ",
    "find class ",
    "find symbol ",
    "copy lines ",
    "replace lines ",
    "move lines ",
    "review lines ",
    "review file ",
    "improve lines ",
    "optimize lines ",
    "refactor lines ",
    "improve file ",
    "optimize file ",
    "refactor file ",
    "generate word report ",
    "generate docx report ",
    "generate pdf report ",
    "generate ppt report ",
    "generate pptx report ",
    "generate powerpoint report ",
    "generate markdown report ",
    "format file ",
    "clean file ",
    "open ",
    "open app ",
    "open application ",
    "open program ",
    "open website ",
    "open site ",
    "go to ",
    "visit ",
    "open folder ",
    "open directory ",
    "open file ",
    "open document ",
    "open project ",
    "open code ",
    "edit project ",
    "find file ",
    "rank file ",
    "read file ",
    "show file ",
    "analyze file ",
    "review file ",
    "improve file ",
    "optimize file ",
    "security review ",
    "preview file ",
    "safe preview file ",
    "backup file ",
    "create backup file ",
    "create backup for file ",
    "restore file ",
    "restore backup file ",
    "restore latest backup file ",
    "suggest safe patch file ",
    "suggest patch file ",
    "safe patch file ",
    "list backups",
    "list file backups",
    "preview project file ",
    "backup project file ",
    "create backup project file ",
    "restore project file ",
    "restore backup project file ",
    "suggest safe patch project file ",
    "suggest patch project file ",
    "apply safe patch project file ",
    "apply ai patch project file ",
    "auto improve project file ",
    "score project ",
    "review project ",
    "analyze project ",
    "security review project ",
    "strict security analyzer project ",
    "strict grounded analyzer project ",
    "strict architecture analyzer project ",
    "suggest fixes ",
    "suggest fixes for project ",
    "export report ",
    "export project report ",
    "daily check",
    "daily project check",
    "smart daily check",
    "show projects",
    "show all projects",
    "list projects",
    "show project structure ",
    "project structure ",
    "show project files ",
    "project files ",
    "show project statistics ",
    "show project stats ",
    "project statistics ",
    "project stats ",
    "show largest files ",
    "largest files ",
    "show security report ",
    "security report ",
    "find api keys ",
    "scan api keys ",
    "find passwords ",
    "scan passwords ",
    "find hardcoded secrets ",
    "scan hardcoded secrets ",
    "find secrets ",
    "scan secrets ",
    "find sql injection ",
    "scan sql injection ",
    "find xss ",
    "scan xss ",
    "find dangerous imports ",
    "scan dangerous imports ",
    "find vulnerable files ",
    "scan vulnerable files ",
    "full security audit ",
    "enterprise audit ",
    "scan entire project ",
    "generate security roadmap ",
    "security roadmap ",

    # Step 7 - Iron Man Developer Mode natural commands
    "secure cyber shield ai",
    "secure cybershield ai",
    "secure cyber",
    "secure project",
    "secure my project",
    "scan cyber shield ai",
    "scan cybershield ai",
    "scan cyber",
    "scan project",
    "scan my project",
    "audit cyber shield ai",
    "audit cybershield ai",
    "audit cyber",
    "audit project",
    "audit my project",
    "find vulnerabilities",
    "find security issues",
    "find security problems",
    "check vulnerabilities",
    "check security",

    # Step 8 - Autonomous Coding Assistant
    "autonomous review ",
    "autonomous improve ",
    "autonomous secure ",
    "autonomous optimize ",
    "autonomous fix ",
    "review everything",
    "review everything ",
    "improve cyber shield ai",
    "improve cybershield ai",
    "improve cyber",
    "fix cyber shield ai",
    "fix cybershield ai",
    "fix cyber",
    "optimize cyber shield ai",
    "optimize cybershield ai",
    "optimize cyber",
    "review all cyber shield ai",
    "review all cybershield ai",
    "review all cyber",
    "improve project",
    "fix project",
    "optimize project",
    "review all project",
    "show backups",
    "show file backups",
    "show project backups",
    "restore backup ",
    "list applications",
    "refresh projects",
    "refresh applications",
    "refresh all",
    "refresh indexes",
    "index status",
    "show index status",
    "smart refresh indexes",
    "force refresh indexes",
    "read screen",
    "analyze screen",
    "read terminal",
    "analyze terminal",
    "read browser",
    "analyze browser",
    "take screenshot",
    "save screenshot",
    "describe screenshot",
    "read code on screen",
    "review code on screen",
    "find bugs on screen",
    "explain error on screen",
    "what error is on screen",
    "read error on screen",
    "fix file ",
    "fix project file ",
    "secure file ",
    "secure project file ",
    "optimize file automatically ",
    "improve file automatically ",
    "apply fix file ",

    # Step 9 - Conversational Memory
    "what was i working on",
    "what was i working on last",
    "what project was i working on",
    "what project was i working on last",
    "what file did we review last",
    "what file did you review last",
    "what vulnerabilities did you find",
    "show last security report",
    "show last audit",
    "continue last audit",
    "continue last project",
    "resume last project",
    "resume last task",
    "continue last task",
    "what projects do you remember",
    "compare remembered projects",
    "memory summary",
    "show memory summary",
    "project conversation summary",
    "last project summary",

    # Step 10 - Deep Memory Integration
    "memory aware review ",
    "memory aware security review ",
    "project timeline ",
    "show project timeline ",
    "audit history",
    "audit history ",
    "vulnerability history",
    "vulnerability history ",
    "remembered fixes",
    "remembered fixes ",
    "project evolution ",
    "show project evolution ",
    "engineering session summary",
    "session summary",
    "continue previous session",
    "last 20 audits",
    "last vulnerabilities",
    "last vulnerabilities ",
    "last improvements",
    "last improvements ",

    # Step 12 - Autonomous Project Commander / generic project commands
    "plan project ",
    "project plan ",
    "architect project ",
    "project architect ",
    "become project architect ",
    "prepare release ",
    "release checklist ",
    "prepare deployment ",
    "deployment checklist ",
    "project maturity ",
    "estimate project maturity ",
    "production readiness ",
    "estimate production readiness ",
    "what should i fix next ",
    "what should i fix next project ",
    "next best improvements ",
    "highest risk vulnerabilities ",
    "high risk vulnerabilities ",
    "generate sprint plan ",
    "sprint plan ",
    "generate roadmap ",
    "generate project roadmap ",
    "project roadmap ",

    # Step 15 - Autonomous Workflow Engine
    "workflow project ",
    "run workflow ",
    "run project workflow ",
    "full workflow ",
    "review workflow ",
    "secure workflow ",
    "production workflow ",
    "release workflow ",
    "review project workflow ",
    "secure project workflow ",
    "prepare project ",
    "prepare project for production ",
    "prepare project for release ",
    "release project ",
    "prepare production ",
    "prepare for production ",
)

DANGEROUS_COMMAND_WORDS = (
    "delete ",
    "remove ",
    "format ",
    "wipe ",
    "erase ",
    "overwrite ",
    "replace code",
    "paste code",
    "write code",
    "modify file",
    "edit file content",
)

BAD_RECOGNITION_MARKERS = (
    "/no_think",
    "call project",
    "fibres",
    "fiber",
    "fibre",
)


def normalize_agent_command(command):
    text = str(command).strip()
    lower = text.lower().strip()

    replacements = {
        "open cyber in vs code": "open project CyberShield AI in VS Code",
        "open cyber shield in vs code": "open project CyberShield AI in VS Code",
        "open cyber shield ai in vs code": "open project CyberShield AI in VS Code",
        "open jarvis in vs code": "open project J.A.R.V.I.S in VS Code",
        "review app.py from cyber": "review file app.py from project CyberShield AI",
        "generate pdf report cyber": "generate pdf report for project CyberShield AI",
        "generate word report cyber": "generate word report for project CyberShield AI",
        "generate ppt report cyber": "generate ppt report for project CyberShield AI",
        "open google chrome": "open chrome",
        "open fire": "open firefox",
        "open download": "open folder downloads",
        "open downloads": "open folder downloads",
        "open document": "open folder documents",
        "open documents": "open folder documents",
        "project cyber": "open project CyberShield AI",
        "project cyber shield": "open project CyberShield AI",
        "project cyber shield ai": "open project CyberShield AI",
        "open project cyber": "open project CyberShield AI",
        "open project cyber shield": "open project CyberShield AI",
        "open project cyber shield ai": "open project CyberShield AI",
        "score project cyber": "score project CyberShield AI",
        "score project cyber shield": "score project CyberShield AI",
        "score project cyber shield ai": "score project CyberShield AI",
        "show project structure cyber": "show project structure CyberShield AI",
        "show project structure cyber shield ai": "show project structure CyberShield AI",
        "show project files cyber": "show project files CyberShield AI",
        "show project files cyber shield ai": "show project files CyberShield AI",
        "show project statistics cyber": "show project statistics CyberShield AI",
        "show project statistics cyber shield ai": "show project statistics CyberShield AI",
        "show project stats cyber": "show project stats CyberShield AI",
        "show project stats cyber shield ai": "show project stats CyberShield AI",
        "show largest files cyber": "show largest files CyberShield AI",
        "show largest files cyber shield ai": "show largest files CyberShield AI",
        "show security report cyber": "show security report CyberShield AI",
        "show security report cyber shield ai": "show security report CyberShield AI",
        "security report cyber": "security report CyberShield AI",
        "security report cyber shield ai": "security report CyberShield AI",
        "find api keys cyber": "find api keys CyberShield AI",
        "find api keys cyber shield ai": "find api keys CyberShield AI",
        "scan api keys cyber": "scan api keys CyberShield AI",
        "find passwords cyber": "find passwords CyberShield AI",
        "find passwords cyber shield ai": "find passwords CyberShield AI",
        "find hardcoded secrets cyber": "find hardcoded secrets CyberShield AI",
        "find hardcoded secrets cyber shield ai": "find hardcoded secrets CyberShield AI",
        "find secrets cyber": "find hardcoded secrets CyberShield AI",
        "find sql injection cyber": "find sql injection CyberShield AI",
        "find sql injection cyber shield ai": "find sql injection CyberShield AI",
        "find xss cyber": "find xss CyberShield AI",
        "find xss cyber shield ai": "find xss CyberShield AI",
        "find dangerous imports cyber": "find dangerous imports CyberShield AI",
        "find dangerous imports cyber shield ai": "find dangerous imports CyberShield AI",
        "find vulnerable files cyber": "full security audit CyberShield AI",
        "find vulnerable files cyber shield ai": "full security audit CyberShield AI",
        "scan entire project cyber": "full security audit CyberShield AI",
        "scan entire project cyber shield ai": "full security audit CyberShield AI",
        "full security audit cyber": "full security audit CyberShield AI",
        "full security audit cyber shield ai": "full security audit CyberShield AI",
        "enterprise audit cyber": "enterprise audit CyberShield AI",
        "enterprise audit cyber shield ai": "enterprise audit CyberShield AI",
        "generate security roadmap cyber": "generate security roadmap CyberShield AI",
        "generate security roadmap cyber shield ai": "generate security roadmap CyberShield AI",
        "security roadmap cyber": "security roadmap CyberShield AI",
        "security roadmap cyber shield ai": "security roadmap CyberShield AI",
        "show backups": "show backups",
        "show file backups": "show backups",
        "show project backups": "show backups",
        "review code": "review code on screen",
        "review this code": "review code on screen",
        "find bugs": "find bugs on screen",
        "find bugs on the screen": "find bugs on screen",
        "explain error": "explain error on screen",
        "explain this error": "explain error on screen",
        "what error": "what error is on screen",
        "fix app.py from cyber": "apply safe patch project file CyberShield AI app.py",

        # Step 9 - Conversational Memory
        "what was i working on": "what was i working on last",
        "what was i working on last": "what was i working on last",
        "what project was i working on": "what was i working on last",
        "what project was i working on last": "what was i working on last",
        "what file did we review last": "what file did we review last",
        "what file did you review last": "what file did we review last",
        "what vulnerabilities did you find": "what vulnerabilities did you find",
        "show last security report": "show last security report",
        "show last audit": "show last audit",
        "continue last audit": "continue last audit",
        "continue last project": "continue last project",
        "resume last project": "resume last project",
        "resume last task": "resume last task",
        "continue last task": "continue last task",
        "what projects do you remember": "what projects do you remember",
        "compare remembered projects": "compare remembered projects",
        "memory summary": "memory summary",
        "show memory summary": "memory summary",
        "project conversation summary": "project conversation summary",
        "last project summary": "last project summary",

        # Step 10 - Deep Memory Integration
        "memory aware review cyber": "memory aware review CyberShield AI",
        "memory aware review cyber shield ai": "memory aware review CyberShield AI",
        "memory aware security review cyber": "memory aware security review CyberShield AI",
        "memory aware security review cyber shield ai": "memory aware security review CyberShield AI",
        "project timeline cyber": "project timeline CyberShield AI",
        "project timeline cyber shield ai": "project timeline CyberShield AI",
        "show project timeline cyber": "project timeline CyberShield AI",
        "show project timeline cyber shield ai": "project timeline CyberShield AI",
        "audit history": "audit history",
        "audit history cyber": "audit history CyberShield AI",
        "audit history cyber shield ai": "audit history CyberShield AI",
        "vulnerability history": "vulnerability history",
        "vulnerability history cyber": "vulnerability history CyberShield AI",
        "vulnerability history cyber shield ai": "vulnerability history CyberShield AI",
        "remembered fixes": "remembered fixes",
        "remembered fixes cyber": "remembered fixes CyberShield AI",
        "remembered fixes cyber shield ai": "remembered fixes CyberShield AI",
        "project evolution cyber": "project evolution CyberShield AI",
        "project evolution cyber shield ai": "project evolution CyberShield AI",
        "show project evolution cyber": "project evolution CyberShield AI",
        "show project evolution cyber shield ai": "project evolution CyberShield AI",
        "engineering session summary": "engineering session summary",
        "session summary": "engineering session summary",
        "continue previous session": "continue previous session",
        "last 20 audits": "last 20 audits",
        "last vulnerabilities": "last vulnerabilities",
        "last vulnerabilities cyber": "last vulnerabilities CyberShield AI",
        "last vulnerabilities cyber shield ai": "last vulnerabilities CyberShield AI",
        "last improvements": "last improvements",
        "last improvements cyber": "last improvements CyberShield AI",
        "last improvements cyber shield ai": "last improvements CyberShield AI",

        # Step 12 - Generic Project Commander examples
        "plan project cyber": "plan project CyberShield AI",
        "plan project cyber shield ai": "plan project CyberShield AI",
        "architect project cyber": "architect project CyberShield AI",
        "architect project cyber shield ai": "architect project CyberShield AI",
        "prepare release cyber": "prepare release CyberShield AI",
        "prepare release cyber shield ai": "prepare release CyberShield AI",
        "prepare deployment cyber": "prepare deployment CyberShield AI",
        "prepare deployment cyber shield ai": "prepare deployment CyberShield AI",
        "project maturity cyber": "project maturity CyberShield AI",
        "project maturity cyber shield ai": "project maturity CyberShield AI",
        "production readiness cyber": "production readiness CyberShield AI",
        "production readiness cyber shield ai": "production readiness CyberShield AI",
        "what should i fix next cyber": "what should i fix next CyberShield AI",
        "what should i fix next cyber shield ai": "what should i fix next CyberShield AI",
        "next best improvements cyber": "next best improvements CyberShield AI",
        "next best improvements cyber shield ai": "next best improvements CyberShield AI",
        "highest risk vulnerabilities cyber": "highest risk vulnerabilities CyberShield AI",
        "highest risk vulnerabilities cyber shield ai": "highest risk vulnerabilities CyberShield AI",
        "generate sprint plan cyber": "generate sprint plan CyberShield AI",
        "generate sprint plan cyber shield ai": "generate sprint plan CyberShield AI",
        "generate roadmap cyber": "generate roadmap CyberShield AI",
        "generate roadmap cyber shield ai": "generate roadmap CyberShield AI",

        # Step 15 - Autonomous Workflow Engine
        "workflow cyber": "workflow project CyberShield AI",
        "workflow cyber shield ai": "workflow project CyberShield AI",
        "run workflow cyber": "workflow project CyberShield AI",
        "run workflow cyber shield ai": "workflow project CyberShield AI",
        "review workflow cyber": "review workflow CyberShield AI",
        "review workflow cyber shield ai": "review workflow CyberShield AI",
        "secure workflow cyber": "secure workflow CyberShield AI",
        "secure workflow cyber shield ai": "secure workflow CyberShield AI",
        "production workflow cyber": "production workflow CyberShield AI",
        "production workflow cyber shield ai": "production workflow CyberShield AI",
        "release workflow cyber": "release workflow CyberShield AI",
        "release workflow cyber shield ai": "release workflow CyberShield AI",
        "prepare project cyber for production": "production workflow CyberShield AI",
        "prepare project cyber shield ai for production": "production workflow CyberShield AI",
        "release project cyber": "release workflow CyberShield AI",
        "release project cyber shield ai": "release workflow CyberShield AI",
    }

    if lower in replacements:
        return replacements[lower]

    # Step 9 - Natural Conversational Memory.
    if (
        "working on" in lower
        or "last project" in lower
        or "current project" in lower
    ):
        if "file" not in lower:
            return "what was i working on last"

    if "last file" in lower or "file did" in lower:
        return "what file did we review last"

    if "last audit" in lower:
        if "continue" in lower or "resume" in lower:
            return "continue last audit"
        return "show last audit"

    if "security report" in lower and "last" in lower:
        return "show last security report"

    if "vulnerabilities" in lower and ("find" in lower or "found" in lower or "did you" in lower):
        return "what vulnerabilities did you find"

    if "continue" in lower and "project" in lower:
        return "continue last project"

    if "resume" in lower and "project" in lower:
        return "resume last project"

    if ("continue" in lower or "resume" in lower) and "task" in lower:
        return "continue last task"

    if "projects do you remember" in lower or "remembered projects" in lower:
        if "compare" in lower:
            return "compare remembered projects"
        return "what projects do you remember"

    if "memory summary" in lower or "summarize memory" in lower:
        return "memory summary"

    # Step 15 - Autonomous Workflow Engine.
    workflow_prefixes = [
        "workflow project ",
        "run workflow ",
        "run project workflow ",
        "full workflow ",
        "review workflow ",
        "secure workflow ",
        "production workflow ",
        "release workflow ",
        "review project workflow ",
        "secure project workflow ",
        "prepare project for production ",
        "prepare project for release ",
        "release project ",
        "prepare production ",
        "prepare for production ",
    ]

    for prefix in workflow_prefixes:
        if lower.startswith(prefix):
            target = text[len(prefix):].strip()

            if target:
                if prefix in ["run workflow ", "run project workflow ", "full workflow "]:
                    return "workflow project " + target
                if prefix in ["review project workflow "]:
                    return "review workflow " + target
                if prefix in ["secure project workflow "]:
                    return "secure workflow " + target
                if prefix in ["prepare project for production ", "prepare production ", "prepare for production "]:
                    return "production workflow " + target
                if prefix in ["prepare project for release ", "release project "]:
                    return "release workflow " + target

                return prefix + target

    if lower.startswith("prepare project ") and " for production" in lower:
        target = text[len("prepare project "):].strip()
        target = re.sub(r"\s+for\s+production\s*$", "", target, flags=re.IGNORECASE).strip()
        if target:
            return "production workflow " + target

    if lower.startswith("prepare project ") and " for release" in lower:
        target = text[len("prepare project "):].strip()
        target = re.sub(r"\s+for\s+release\s*$", "", target, flags=re.IGNORECASE).strip()
        if target:
            return "release workflow " + target

    # Step 12 - Generic Autonomous Project Commander.
    # These commands keep the project name dynamic instead of forcing CyberShield AI.
    generic_project_prefixes = [
        "plan project ",
        "project plan ",
        "architect project ",
        "project architect ",
        "become project architect ",
        "prepare release ",
        "release checklist ",
        "prepare deployment ",
        "deployment checklist ",
        "project maturity ",
        "estimate project maturity ",
        "production readiness ",
        "estimate production readiness ",
        "what should i fix next ",
        "what should i fix next project ",
        "next best improvements ",
        "highest risk vulnerabilities ",
        "high risk vulnerabilities ",
        "generate sprint plan ",
        "sprint plan ",
        "generate roadmap ",
        "generate project roadmap ",
        "project roadmap ",
    ]

    for prefix in generic_project_prefixes:
        if lower.startswith(prefix):
            target = text[len(prefix):].strip()

            if target:
                if prefix in ["project plan ", "generate project roadmap ", "project roadmap "]:
                    return "plan project " + target
                if prefix in ["project architect ", "become project architect "]:
                    return "architect project " + target
                if prefix == "release checklist ":
                    return "prepare release " + target
                if prefix == "deployment checklist ":
                    return "prepare deployment " + target
                if prefix == "estimate project maturity ":
                    return "project maturity " + target
                if prefix == "estimate production readiness ":
                    return "production readiness " + target
                if prefix == "what should i fix next project ":
                    return "what should i fix next " + target
                if prefix == "high risk vulnerabilities ":
                    return "highest risk vulnerabilities " + target
                if prefix == "sprint plan ":
                    return "generate sprint plan " + target
                if prefix == "generate roadmap ":
                    return "plan project " + target

                return prefix + target

    if lower.startswith("prepare ") and " for release" in lower:
        target = text[len("prepare "):].strip()
        target = re.sub(r"\s+for\s+release\s*$", "", target, flags=re.IGNORECASE).strip()
        if target:
            return "prepare release " + target

    if lower.startswith("prepare ") and " for deployment" in lower:
        target = text[len("prepare "):].strip()
        target = re.sub(r"\s+for\s+deployment\s*$", "", target, flags=re.IGNORECASE).strip()
        if target:
            return "prepare deployment " + target

    # Step 10 - Natural Deep Memory Integration.
    if "memory aware" in lower and "security" in lower:
        if "cyber" in lower or "project" in lower:
            return "memory aware security review CyberShield AI"

    if "memory aware" in lower and "review" in lower:
        if "cyber" in lower or "project" in lower:
            return "memory aware review CyberShield AI"

    if "project timeline" in lower or "timeline project" in lower:
        if "cyber" in lower or "project" in lower:
            return "project timeline CyberShield AI"

    if "audit history" in lower:
        if "cyber" in lower:
            return "audit history CyberShield AI"
        return "audit history"

    if "vulnerability history" in lower or "vulnerabilities history" in lower:
        if "cyber" in lower:
            return "vulnerability history CyberShield AI"
        return "vulnerability history"

    if "remembered fixes" in lower or "previous fixes" in lower or "last fixes" in lower:
        if "cyber" in lower:
            return "remembered fixes CyberShield AI"
        return "remembered fixes"

    if "project evolution" in lower or "evolution project" in lower:
        if "cyber" in lower or "project" in lower:
            return "project evolution CyberShield AI"

    if "engineering session summary" in lower or "session summary" in lower:
        return "engineering session summary"

    if "continue previous session" in lower or "resume previous session" in lower:
        return "continue previous session"

    if "last 20 audits" in lower:
        return "last 20 audits"

    if "last vulnerabilities" in lower or "recent vulnerabilities" in lower:
        if "cyber" in lower:
            return "last vulnerabilities CyberShield AI"
        return "last vulnerabilities"

    if "last improvements" in lower or "recent improvements" in lower:
        if "cyber" in lower:
            return "last improvements CyberShield AI"
        return "last improvements"

    # Step 7 - Natural Iron Man Developer Mode.
    if lower in {
        "secure project",
        "secure my project",
        "scan project",
        "scan my project",
        "audit project",
        "audit my project",
        "find vulnerabilities",
        "find security issues",
        "find security problems",
        "check vulnerabilities",
        "check security",
    }:
        return "full security audit CyberShield AI"

    if (
        any(word in lower for word in ["secure", "scan", "audit"])
        and any(word in lower for word in ["cyber", "cybershield", "cyber shield"])
    ):
        if "roadmap" in lower:
            return "generate security roadmap CyberShield AI"
        if "audit" in lower:
            return "enterprise audit CyberShield AI"
        return "full security audit CyberShield AI"

    if (
        "vulnerab" in lower
        or "security issues" in lower
        or "security problems" in lower
        or "weaknesses" in lower
    ):
        if "cyber" in lower or "project" in lower or lower == "find vulnerabilities":
            return "full security audit CyberShield AI"

    if "api key" in lower or "api keys" in lower:
        if "cyber" in lower or "project" in lower:
            return "find api keys CyberShield AI"

    if "password" in lower or "passwords" in lower:
        if "cyber" in lower or "project" in lower:
            return "find passwords CyberShield AI"

    if "secret" in lower or "secrets" in lower:
        if "cyber" in lower or "project" in lower:
            return "find hardcoded secrets CyberShield AI"

    if "sql injection" in lower:
        if "cyber" in lower or "project" in lower:
            return "find sql injection CyberShield AI"

    if "xss" in lower:
        if "cyber" in lower or "project" in lower:
            return "find xss CyberShield AI"

    if "dangerous import" in lower or "dangerous imports" in lower:
        if "cyber" in lower or "project" in lower:
            return "find dangerous imports CyberShield AI"

    if "security roadmap" in lower or "roadmap security" in lower:
        return "generate security roadmap CyberShield AI"

    # Step 8 - Natural Autonomous Coding Assistant.
    if lower in {
        "review everything",
        "review all",
        "review all project",
        "review whole project",
    }:
        return "review everything CyberShield AI"

    if (
        any(word in lower for word in ["improve", "fix", "optimize", "review"])
        and any(word in lower for word in ["cyber", "cybershield", "cyber shield"])
    ):
        if "fix" in lower:
            return "autonomous fix CyberShield AI"
        if "optimize" in lower:
            return "autonomous optimize CyberShield AI"
        if "review" in lower:
            return "autonomous review CyberShield AI"
        if "improve" in lower:
            return "autonomous improve CyberShield AI"

    if lower in {
        "improve project",
        "improve my project",
        "fix project",
        "fix my project",
        "optimize project",
        "optimize my project",
        "review project",
        "review my project",
    }:
        if lower.startswith("fix"):
            return "autonomous fix CyberShield AI"
        if lower.startswith("optimize"):
            return "autonomous optimize CyberShield AI"
        if lower.startswith("review"):
            return "autonomous review CyberShield AI"
        return "autonomous improve CyberShield AI"

    if lower.startswith((
        "open project ",
        "score project ",
        "review project ",
        "analyze project ",
        "suggest fixes ",
        "strict security ",
        "export report ",
        "show project structure ",
        "show project files ",
        "show project statistics ",
        "show project stats ",
        "show largest files ",
        "show security report ",
        "security report ",
        "find api keys ",
        "scan api keys ",
        "find passwords ",
        "scan passwords ",
        "find hardcoded secrets ",
        "scan hardcoded secrets ",
        "find secrets ",
        "scan secrets ",
        "find sql injection ",
        "scan sql injection ",
        "find xss ",
        "scan xss ",
        "find dangerous imports ",
        "scan dangerous imports ",
        "find vulnerable files ",
        "scan vulnerable files ",
        "full security audit ",
        "enterprise audit ",
        "scan entire project ",
        "generate security roadmap ",
        "security roadmap ",
        "autonomous review ",
        "autonomous improve ",
        "autonomous secure ",
        "autonomous optimize ",
        "autonomous fix ",
        "review everything ",
        "memory aware review ",
        "memory aware security review ",
        "project timeline ",
        "show project timeline ",
        "audit history ",
        "vulnerability history ",
        "remembered fixes ",
        "project evolution ",
        "show project evolution ",
        "last vulnerabilities ",
        "last improvements ",
        "plan project ",
        "project plan ",
        "architect project ",
        "project architect ",
        "prepare release ",
        "release checklist ",
        "prepare deployment ",
        "deployment checklist ",
        "project maturity ",
        "production readiness ",
        "what should i fix next ",
        "next best improvements ",
        "highest risk vulnerabilities ",
        "high risk vulnerabilities ",
        "generate sprint plan ",
        "sprint plan ",
        "generate roadmap ",
        "generate project roadmap ",
        "project roadmap ",
        "workflow project ",
        "run workflow ",
        "run project workflow ",
        "full workflow ",
        "review workflow ",
        "secure workflow ",
        "production workflow ",
        "release workflow ",
        "review project workflow ",
        "secure project workflow ",
        "prepare project ",
        "release project ",
    )) and "cyber" in lower:
        if lower.startswith("open project "):
            return "open project CyberShield AI"
        if lower.startswith("score project "):
            return "score project CyberShield AI"
        if lower.startswith("review project "):
            return "review project CyberShield AI"
        if lower.startswith("analyze project "):
            return "analyze project CyberShield AI"
        if lower.startswith("suggest fixes "):
            return "suggest fixes for project CyberShield AI"
        if lower.startswith("strict security "):
            return "strict security analyzer project CyberShield AI"
        if lower.startswith("export report "):
            return "export report CyberShield AI"
        if lower.startswith("show project structure "):
            return "show project structure CyberShield AI"
        if lower.startswith("show project files "):
            return "show project files CyberShield AI"
        if lower.startswith("show project statistics "):
            return "show project statistics CyberShield AI"
        if lower.startswith("show project stats "):
            return "show project stats CyberShield AI"
        if lower.startswith("show largest files "):
            return "show largest files CyberShield AI"
        if lower.startswith("show security report "):
            return "show security report CyberShield AI"
        if lower.startswith("security report "):
            return "security report CyberShield AI"
        if lower.startswith("find api keys ") or lower.startswith("scan api keys "):
            return "find api keys CyberShield AI"
        if lower.startswith("find passwords ") or lower.startswith("scan passwords "):
            return "find passwords CyberShield AI"
        if lower.startswith("find hardcoded secrets ") or lower.startswith("scan hardcoded secrets ") or lower.startswith("find secrets ") or lower.startswith("scan secrets "):
            return "find hardcoded secrets CyberShield AI"
        if lower.startswith("find sql injection ") or lower.startswith("scan sql injection "):
            return "find sql injection CyberShield AI"
        if lower.startswith("find xss ") or lower.startswith("scan xss "):
            return "find xss CyberShield AI"
        if lower.startswith("find dangerous imports ") or lower.startswith("scan dangerous imports "):
            return "find dangerous imports CyberShield AI"
        if lower.startswith("find vulnerable files ") or lower.startswith("scan vulnerable files "):
            return "full security audit CyberShield AI"
        if lower.startswith("full security audit ") or lower.startswith("enterprise audit ") or lower.startswith("scan entire project "):
            return "full security audit CyberShield AI"
        if lower.startswith("generate security roadmap ") or lower.startswith("security roadmap "):
            return "generate security roadmap CyberShield AI"
        if lower.startswith("autonomous review ") or lower.startswith("review everything "):
            return "autonomous review CyberShield AI"
        if lower.startswith("autonomous improve "):
            return "autonomous improve CyberShield AI"
        if lower.startswith("autonomous secure "):
            return "autonomous secure CyberShield AI"
        if lower.startswith("autonomous optimize "):
            return "autonomous optimize CyberShield AI"
        if lower.startswith("autonomous fix "):
            return "autonomous fix CyberShield AI"
        if lower.startswith("memory aware security review "):
            return "memory aware security review CyberShield AI"
        if lower.startswith("memory aware review "):
            return "memory aware review CyberShield AI"
        if lower.startswith("project timeline ") or lower.startswith("show project timeline "):
            return "project timeline CyberShield AI"
        if lower.startswith("audit history "):
            return "audit history CyberShield AI"
        if lower.startswith("vulnerability history "):
            return "vulnerability history CyberShield AI"
        if lower.startswith("remembered fixes "):
            return "remembered fixes CyberShield AI"
        if lower.startswith("project evolution ") or lower.startswith("show project evolution "):
            return "project evolution CyberShield AI"
        if lower.startswith("last vulnerabilities "):
            return "last vulnerabilities CyberShield AI"
        if lower.startswith("last improvements "):
            return "last improvements CyberShield AI"
        if lower.startswith("plan project ") or lower.startswith("project plan ") or lower.startswith("generate roadmap ") or lower.startswith("generate project roadmap ") or lower.startswith("project roadmap "):
            return "plan project CyberShield AI"
        if lower.startswith("architect project ") or lower.startswith("project architect "):
            return "architect project CyberShield AI"
        if lower.startswith("prepare release ") or lower.startswith("release checklist "):
            return "prepare release CyberShield AI"
        if lower.startswith("prepare deployment ") or lower.startswith("deployment checklist "):
            return "prepare deployment CyberShield AI"
        if lower.startswith("project maturity ") or lower.startswith("estimate project maturity "):
            return "project maturity CyberShield AI"
        if lower.startswith("production readiness ") or lower.startswith("estimate production readiness "):
            return "production readiness CyberShield AI"
        if lower.startswith("what should i fix next ") or lower.startswith("next best improvements "):
            return "what should i fix next CyberShield AI"
        if lower.startswith("highest risk vulnerabilities ") or lower.startswith("high risk vulnerabilities "):
            return "highest risk vulnerabilities CyberShield AI"
        if lower.startswith("generate sprint plan ") or lower.startswith("sprint plan "):
            return "generate sprint plan CyberShield AI"
        if lower.startswith(("workflow project ", "run workflow ", "run project workflow ", "full workflow ")):
            return "workflow project CyberShield AI"
        if lower.startswith(("review workflow ", "review project workflow ")):
            return "review workflow CyberShield AI"
        if lower.startswith(("secure workflow ", "secure project workflow ")):
            return "secure workflow CyberShield AI"
        if lower.startswith(("production workflow ", "prepare project ")):
            return "production workflow CyberShield AI"
        if lower.startswith(("release workflow ", "release project ")):
            return "release workflow CyberShield AI"

    return text


def is_bad_or_unclear_command(command):
    lower = str(command).lower().strip()

    if not lower:
        return True

    if any(marker in lower for marker in BAD_RECOGNITION_MARKERS):
        return True

    return False


def is_dangerous_command(command):
    lower = str(command).lower().strip()

    safe_edit_prefixes = (
        "backup file ",
        "create backup file ",
        "create backup for file ",
        "backup project file ",
        "create backup project file ",
        "restore file ",
        "restore backup file ",
        "restore latest backup file ",
        "restore project file ",
        "restore backup project file ",
        "suggest safe patch file ",
        "suggest patch file ",
        "safe patch file ",
        "suggest safe patch project file ",
        "suggest patch project file ",
        "list backups",
        "list file backups",
        "apply safe patch project file ",
        "apply ai patch project file ",
        "auto improve project file ",
        "fix file ",
        "fix project file ",
        "secure file ",
        "secure project file ",
        "optimize file automatically ",
        "improve file automatically ",
        "apply fix file ",
    )

    if lower.startswith(safe_edit_prefixes):
        return False

    return any(word in lower for word in DANGEROUS_COMMAND_WORDS)


def has_safe_command_intent(command):
    lower = str(command).lower().strip()

    if lower == "exit":
        return True

    return any(lower.startswith(prefix) for prefix in SAFE_COMMAND_PREFIXES)


def route_universal_open_command(command):
    lower = command.lower().strip()

    if lower.startswith("open app "):
        target = command[len("open app "):].strip()
        return open_installed_app(target)

    if lower.startswith("open application "):
        target = command[len("open application "):].strip()
        return open_installed_app(target)

    if lower.startswith("open program "):
        target = command[len("open program "):].strip()
        return open_installed_app(target)

    if lower.startswith("open folder "):
        target = command[len("open folder "):].strip()
        return open_universal_folder(target)

    if lower.startswith("open directory "):
        target = command[len("open directory "):].strip()
        return open_universal_folder(target)

    if lower.startswith("open project "):
        target = command[len("open project "):].strip()
        return open_project_folder_universal(target)

    if lower.startswith("open code "):
        target = command[len("open code "):].strip()
        return open_project_in_vscode(target)

    if lower.startswith("open website "):
        target = command[len("open website "):].strip()
        return open_website(target)

    if lower.startswith("open site "):
        target = command[len("open site "):].strip()
        return open_website(target)

    if lower.startswith("go to "):
        target = command[len("go to "):].strip()
        return open_website(target)

    if lower.startswith("visit "):
        target = command[len("visit "):].strip()
        return open_website(target)

    return None



# ==========================
# AUTO INDEX INITIALIZATION
# ==========================
def ensure_indexes_ready(force=False):
    """
    Smart startup index initialization.

    This uses the Smart Auto Indexer from tools.py.
    It does not rebuild everything at every startup.
    It rebuilds only when indexes are missing, empty, old or the project roots changed.
    """
    print("[JARVIS] Smart index check started...")

    update_hud_context(
        status="PROCESSING",
        command="startup",
        result="Checking smart indexes...",
        action="Smart Auto Indexer",
        ai_status="READY",
        memory_status="SYNC",
        ollama_status="LOCAL"
    )

    try:
        result = smart_ensure_indexes_ready(force=force)

        print("[JARVIS] Smart index check completed.")
        print(result)

        update_hud_context(
            status="SUCCESS",
            command="startup",
            result=result,
            action="Indexes ready",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    except Exception as e:
        error_message = f"Smart index check failed: {e}"
        print(f"[JARVIS] {error_message}")

        # Safe fallback: try the older refresh mechanism.
        try:
            fallback_result = refresh_all_indexes()

            update_hud_context(
                status="SUCCESS",
                command="startup",
                result=fallback_result,
                action="Fallback index refresh",
                ai_status="READY",
                memory_status="SYNC",
                ollama_status="READY"
            )

            return fallback_result

        except Exception as fallback_error:
            final_error = f"{error_message}. Fallback also failed: {fallback_error}"

            update_hud_context(
                status="ERROR",
                command="startup",
                result=final_error,
                action="Index check failed",
                ai_status="READY",
                memory_status="SYNC",
                ollama_status="LOCAL"
            )

            return final_error


def handle_command(command):
    enterprise_natural_result = handle_enterprise_natural_command(command)

    if enterprise_natural_result is not None:
        hud_finish_action(
            command,
            enterprise_natural_result,
            action="Enterprise natural command",
            project=_ctx_get("last_project", None)
        )
        return enterprise_natural_result

    project_lookup_result = handle_project_lookup_command(command)

    if project_lookup_result is not None:
        update_hud_context(
            status="SUCCESS",
            command=command,
            result=project_lookup_result,
            action="Project/file lookup",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY"
        )
        return project_lookup_result

    # ==========================
    # DEVELOPER ASSISTANT ROUTING - EARLY
    # Runs before old project/file routers.
    # ==========================
    developer_result = handle_natural_developer_command(command)

    if developer_result is None:
        developer_result = handle_developer_command(command)

    if developer_result is not None:
        update_hud_context(
            status="SUCCESS",
            command=command,
            result=developer_result,
            action="Developer assistant",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY"
        )
        return developer_result

    command = normalize_agent_command(command).strip()

    update_hud_context(
        status="PROCESSING",
        command=command,
        action="Agent routing",
        ai_status="READY",
        memory_status="SYNC",
        ollama_status="LOCAL"
    )

    if not command:
        return "No command received."

    if is_bad_or_unclear_command(command):
        return (
            "Command was not clear enough, so I did not execute it. "
            "Please repeat using a clear command like: open calculator, open folder downloads, "
            "open project CyberShield AI, or score project CyberShield AI."
        )

    if is_dangerous_command(command):
        return (
            "This command can modify or delete content. For safety, I will not execute it here. "
            "Use a dedicated safe edit command with backup support."
        )

    if not has_safe_command_intent(command):
        return (
            "I did not detect a clear JARVIS command, so I did not send it to Ollama. "
            "Use: open, read, analyze, review, score project, suggest fixes, export report, or daily check."
        )

    command_lower = command.lower()

    # ==========================
    # EXIT
    # ==========================
    if command_lower == "exit":
        return "exit"

    # ==========================
    # DEVELOPER ASSISTANT ROUTING
    # ==========================
    developer_result = handle_developer_command(command)

    if developer_result is not None:
        update_hud_context(
            status="SUCCESS",
            command=command,
            result=developer_result,
            action="Developer assistant",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY"
        )
        return developer_result

    # ==========================
    # SMART INDEX COMMANDS
    # ==========================
    if command_lower in {
        "index status",
        "show index status",
    }:
        return index_status()

    if command_lower in {
        "smart refresh indexes",
        "refresh indexes",
        "refresh all",
    }:
        return smart_refresh_all_indexes(force=False)

    if command_lower in {
        "force refresh indexes",
        "force refresh all indexes",
    }:
        return smart_refresh_all_indexes(force=True)

    # ==========================
    # UNIVERSAL OPEN ROUTING
    # Must stay early so app/site/folder/project commands do not fall through.
    # ==========================
    universal_open_result = route_universal_open_command(command)

    if universal_open_result is not None:
        remember(
            "universal_open",
            f"Universal open: {command}",
            str(universal_open_result)[:3000],
            ["open", "universal", command]
        )

        return universal_open_result

    if command_lower in ["refresh all", "refresh indexes", "refresh all indexes"]:
        result = refresh_all_indexes()

        remember(
            "system",
            "Refreshed all indexes",
            result[:3000],
            ["refresh", "indexes"]
        )

        return result


    # ==========================
    # DEBUG
    # ==========================
    if command_lower == "jarvis debug project review":
        return (
            "Project review routing is active.\n"
            "Available commands:\n"
            "- review project <project>\n"
            "- security review project <project>\n"
            "- project evidence <project>\n"
            "- grounded review project <project>\n"
            "- grounded security review project <project>\n"
            "- grounded architecture review project <project>\n"
            "- grounded docs project <project>\n"
            "- strict grounded analyzer project <project>\n"
            "- strict security analyzer project <project>\n"
            "- strict architecture analyzer project <project>\n"
            "- score project <project>\n"
            "- export report <project>\n"
            "- export project report <project>\n"
            "- daily project check\n"
            "- daily check\n"
            "- jarvis daily check\n"
            "- smart daily check\n"
            "- suggest fixes for project <project>\n"
            "- suggest fixes <project>\n"
            "- project fixes <project>\n"
            "- compare projects <project A> and <project B>\n"
            "- compare security <project A> and <project B>\n"
            "- compare architecture <project A> and <project B>\n"
            "- rank projects by security\n"
            "- rank projects by architecture\n"
            "- rank projects by maintainability\n"
            "- best project in memory\n"
            "- find security issues in project <project>\n"
            "- find dead code in project <project>\n"
            "- find duplicates in project <project>\n"
            "- generate architecture report for <project>\n"
            "- architecture review project <project>\n"
            "- generate improvement roadmap for <project>\n"
            "- generate documentation for project <project>\n"
            "- documentation project <project>\n"
            "- project docs <project>\n"
            "- analyze project structure <project>\n"
            "- optimize project <project>\n"
            "- refresh deep project <project>\n"
            "- show projects\n- show project structure <project>\n- show project files <project>\n- show project statistics <project>\n- show largest files <project>\n- show security report <project>\n- find api keys <project>\n- find passwords <project>\n- find hardcoded secrets <project>\n- find sql injection <project>\n- find xss <project>\n- find dangerous imports <project>\n- full security audit <project>\n- enterprise audit <project>\n- secure cyber shield ai\n- scan cyber shield ai\n- audit cyber shield ai\n- find vulnerabilities\n- autonomous review <project>\n- autonomous improve <project>\n- autonomous secure <project>\n- autonomous optimize <project>\n- autonomous fix <project>\n- review everything\n- what was i working on last\n- what file did we review last\n- continue last audit\n- show last security report\n- what vulnerabilities did you find\n- what projects do you remember\n- memory summary\n- memory aware review <project>\n- memory aware security review <project>\n- project timeline <project>\n- audit history <project>\n- vulnerability history <project>\n- remembered fixes <project>\n- project evolution <project>\n- engineering session summary\n- continue previous session\n- plan project <project>\n- architect project <project>\n- prepare release <project>\n- prepare deployment <project>\n- project maturity <project>\n- production readiness <project>\n- what should i fix next <project>\n- highest risk vulnerabilities <project>\n- generate sprint plan <project>\n- workflow project <project>\n- review workflow <project>\n- secure workflow <project>\n- production workflow <project>\n- release workflow <project>\n- generate security roadmap <project>\n- show backups\n"
            "- show projects from <drive>\n"
            "- search projects <keyword>\n"
            "- find file <file>\n"
            "- rank file <file>\n"
            "- open file #<number>\n"
            "- read file #<number>\n"
            "- review file #<number>\n"
            "- security review file #<number>\n"
            "- open file <file> from <project>\n"
            "- read file <file> from <project>\n"
            "- open best <file> from <project>\n"
            "- read best <file> from <project>\n"
            "- review best <file> from <project>\n"
            "- security review best <file> from <project>\n- preview file <file> from <project>\n- backup file <file> from <project>\n- suggest safe patch file <file> from <project>\n- restore backup file <file> from <project>\n- list backups"
        )


    # ==========================
    # GENERAL MEMORY COMMANDS
    # ==========================
    # ==========================
    # STEP 15 - AUTONOMOUS WORKFLOW ENGINE
    # Consolidated workflows for any indexed/deep project.
    # ==========================
    if command_lower.startswith("workflow project "):
        project_name = command[
            len("workflow project "):
        ].strip()

        return autonomous_project_workflow(
            project_name,
            "full"
        )

    if command_lower.startswith("review workflow "):
        project_name = command[
            len("review workflow "):
        ].strip()

        return review_project_workflow(project_name)

    if command_lower.startswith("secure workflow "):
        project_name = command[
            len("secure workflow "):
        ].strip()

        return secure_project_workflow(project_name)

    if command_lower.startswith("production workflow "):
        project_name = command[
            len("production workflow "):
        ].strip()

        return prepare_project_for_production_workflow(project_name)

    if command_lower.startswith("release workflow "):
        project_name = command[
            len("release workflow "):
        ].strip()

        return release_project_workflow(project_name)

    # ==========================
    # STEP 12 - AUTONOMOUS PROJECT COMMANDER
    # Generic planner commands for any indexed/deep project.
    # ==========================
    if command_lower.startswith("plan project "):
        project_name = command[
            len("plan project "):
        ].strip()

        hud_start_action(
            command,
            "Generating project roadmap",
            project=project_name,
            thinking=True
        )

        result = generate_project_roadmap(project_name)

        remember_project_event(
            project_name,
            "roadmap",
            "Project commander roadmap",
            result[:12000],
            tags=["commander", "roadmap", "planner"],
            metadata={"source": "jarvis_agent_step12"}
        )

        hud_finish_action(
            command,
            result,
            "Project roadmap completed",
            project=project_name,
            project_score=extract_hud_project_score(result),
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("architect project "):
        project_name = command[
            len("architect project "):
        ].strip()

        hud_start_action(
            command,
            "Project architect mode",
            project=project_name,
            thinking=True
        )

        result = become_project_architect(project_name)

        remember_project_event(
            project_name,
            "architecture",
            "Project architect mode",
            result[:12000],
            tags=["commander", "architect", "architecture"],
            metadata={"source": "jarvis_agent_step12"}
        )

        hud_finish_action(
            command,
            result,
            "Project architect report completed",
            project=project_name,
            project_score=extract_hud_project_score(result),
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("prepare release "):
        project_name = command[
            len("prepare release "):
        ].strip()

        hud_start_action(
            command,
            "Preparing release checklist",
            project=project_name,
            thinking=False
        )

        result = generate_release_checklist(project_name)

        remember_project_event(
            project_name,
            "release",
            "Release checklist",
            result[:8000],
            tags=["commander", "release", "checklist"],
            metadata={"source": "jarvis_agent_step12"}
        )

        hud_finish_action(
            command,
            result,
            "Release checklist completed",
            project=project_name
        )

        return result

    if command_lower.startswith("prepare deployment "):
        project_name = command[
            len("prepare deployment "):
        ].strip()

        hud_start_action(
            command,
            "Preparing deployment checklist",
            project=project_name,
            thinking=False
        )

        result = generate_deployment_checklist(project_name)

        remember_project_event(
            project_name,
            "deployment",
            "Deployment checklist",
            result[:8000],
            tags=["commander", "deployment", "checklist"],
            metadata={"source": "jarvis_agent_step12"}
        )

        hud_finish_action(
            command,
            result,
            "Deployment checklist completed",
            project=project_name
        )

        return result

    if command_lower.startswith("project maturity "):
        project_name = command[
            len("project maturity "):
        ].strip()

        result = estimate_project_maturity(project_name)

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            project=project_name,
            action="Project maturity estimated",
            ai_status="READY",
            project_score=extract_hud_project_score(result),
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower.startswith("production readiness "):
        project_name = command[
            len("production readiness "):
        ].strip()

        result = estimate_production_readiness(project_name)

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            project=project_name,
            action="Production readiness estimated",
            ai_status="READY",
            security_score=extract_hud_security_score(result),
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower.startswith("what should i fix next "):
        project_name = command[
            len("what should i fix next "):
        ].strip()

        hud_start_action(
            command,
            "Finding next best fixes",
            project=project_name,
            thinking=True
        )

        result = what_should_i_fix_next(project_name)

        remember_project_event(
            project_name,
            "fix_plan",
            "What should I fix next",
            result[:10000],
            tags=["commander", "fix_next", "improvements"],
            metadata={"source": "jarvis_agent_step12"}
        )

        hud_finish_action(
            command,
            result,
            "Next fixes ready",
            project=project_name,
            project_score=extract_hud_project_score(result),
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("next best improvements "):
        project_name = command[
            len("next best improvements "):
        ].strip()

        return what_should_i_fix_next(project_name)

    if command_lower.startswith("highest risk vulnerabilities "):
        project_name = command[
            len("highest risk vulnerabilities "):
        ].strip()

        hud_start_action(
            command,
            "Finding highest risk vulnerabilities",
            project=project_name,
            thinking=False
        )

        result = highest_risk_vulnerabilities(project_name)

        remember_project_event(
            project_name,
            "audit",
            "Highest risk vulnerabilities",
            result[:10000],
            tags=["commander", "security", "vulnerabilities", "risk"],
            metadata={"source": "jarvis_agent_step12"}
        )

        hud_finish_action(
            command,
            result,
            "Highest risk vulnerabilities ready",
            project=project_name,
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("generate sprint plan "):
        project_name = command[
            len("generate sprint plan "):
        ].strip()

        hud_start_action(
            command,
            "Generating sprint plan",
            project=project_name,
            thinking=True
        )

        result = generate_sprint_plan(project_name)

        remember_project_event(
            project_name,
            "sprint",
            "Sprint plan",
            result[:10000],
            tags=["commander", "sprint", "planner"],
            metadata={"source": "jarvis_agent_step12"}
        )

        hud_finish_action(
            command,
            result,
            "Sprint plan completed",
            project=project_name,
            project_score=extract_hud_project_score(result),
            security_score=extract_hud_security_score(result)
        )

        return result

    # ==========================
    # STEP 10 - DEEP MEMORY INTEGRATION
    # ==========================
    if command_lower.startswith("memory aware security review "):
        project_name = command[
            len("memory aware security review "):
        ].strip()

        hud_start_action(
            command,
            "Memory-aware security review",
            project=project_name,
            thinking=True
        )

        result = memory_aware_security_review(project_name)

        remember_project_event(
            project_name,
            "audit",
            "Memory-aware security review",
            result[:12000],
            tags=["security", "memory", "review", "audit"],
            metadata={"source": "jarvis_agent_step10"}
        )

        hud_finish_action(
            command,
            result,
            "Memory-aware security review completed",
            project=project_name,
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("memory aware review "):
        project_name = command[
            len("memory aware review "):
        ].strip()

        hud_start_action(
            command,
            "Memory-aware project review",
            project=project_name,
            thinking=True
        )

        result = memory_aware_review(project_name)

        remember_project_event(
            project_name,
            "review",
            "Memory-aware project review",
            result[:12000],
            tags=["memory", "review", "project"],
            metadata={"source": "jarvis_agent_step10"}
        )

        hud_finish_action(
            command,
            result,
            "Memory-aware review completed",
            project=project_name,
            project_score=extract_hud_project_score(result),
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("project timeline "):
        project_name = command[
            len("project timeline "):
        ].strip()

        result = project_timeline(project_name)

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            project=project_name,
            action="Project timeline loaded",
            ai_status="READY",
            memory_status="ACTIVE",
            ollama_status="READY"
        )

        return result

    if command_lower.startswith("audit history "):
        project_name = command[
            len("audit history "):
        ].strip()

        return audit_history(project_name)

    if command_lower == "audit history":
        return audit_history()

    if command_lower.startswith("vulnerability history "):
        project_name = command[
            len("vulnerability history "):
        ].strip()

        return vulnerability_history(project_name)

    if command_lower == "vulnerability history":
        return vulnerability_history()

    if command_lower.startswith("remembered fixes "):
        project_name = command[
            len("remembered fixes "):
        ].strip()

        return remembered_fixes(project_name)

    if command_lower == "remembered fixes":
        return remembered_fixes()

    if command_lower.startswith("project evolution "):
        project_name = command[
            len("project evolution "):
        ].strip()

        hud_start_action(
            command,
            "Project evolution report",
            project=project_name,
            thinking=True
        )

        try:
            result = project_evolution_report(project_name)
        except Exception:
            result = project_evolution(project_name)

        hud_finish_action(
            command,
            result,
            "Project evolution completed",
            project=project_name,
            project_score=extract_hud_project_score(result),
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower in [
        "engineering session summary",
        "session summary",
    ]:
        try:
            result = engineering_session_summary()
        except Exception:
            result = session_summary()

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            action="Engineering session summary",
            ai_status="READY",
            memory_status="ACTIVE",
            ollama_status="READY"
        )

        return result

    if command_lower == "continue previous session":
        result = continue_previous_session()

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            action="Previous session loaded",
            ai_status="READY",
            memory_status="ACTIVE",
            ollama_status="READY"
        )

        return result

    if command_lower == "last 20 audits":
        return last_20_audits()

    if command_lower.startswith("last vulnerabilities "):
        project_name = command[
            len("last vulnerabilities "):
        ].strip()

        return last_vulnerabilities(project_name)

    if command_lower == "last vulnerabilities":
        return last_vulnerabilities()

    if command_lower.startswith("last improvements "):
        project_name = command[
            len("last improvements "):
        ].strip()

        return last_improvements(project_name)

    if command_lower == "last improvements":
        return last_improvements()

    # ==========================
    # STEP 9 - CONVERSATIONAL MEMORY
    # ==========================
    if command_lower in [
        "what was i working on",
        "what was i working on last",
        "what project was i working on",
        "what project was i working on last",
    ]:
        update_hud_context(
            status="PROCESSING",
            command=command,
            action="Reading last project memory",
            ai_status="READY",
            memory_status="ACTIVE",
            ollama_status="LOCAL"
        )

        result = what_was_i_working_on_last()

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            action="Last project memory found",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower in [
        "what file did we review last",
        "what file did you review last",
    ]:
        update_hud_context(
            status="PROCESSING",
            command=command,
            action="Reading last file memory",
            ai_status="READY",
            memory_status="ACTIVE",
            ollama_status="LOCAL"
        )

        result = what_file_did_we_review_last()

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            action="Last file memory found",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower in [
        "show last audit",
        "continue last audit",
    ]:
        update_hud_context(
            status="PROCESSING",
            command=command,
            action="Reading last audit",
            ai_status="READY",
            memory_status="ACTIVE",
            ollama_status="LOCAL"
        )

        if command_lower == "continue last audit":
            result = continue_last_audit()
        else:
            result = last_project_audit()

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            action="Last audit loaded",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY",
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower in [
        "show last security report",
        "what vulnerabilities did you find",
    ]:
        update_hud_context(
            status="PROCESSING",
            command=command,
            action="Reading last security report",
            ai_status="READY",
            memory_status="ACTIVE",
            ollama_status="LOCAL"
        )

        result = show_last_security_report()

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            action="Last security report loaded",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY",
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower in [
        "continue last project",
        "resume last project",
    ]:
        update_hud_context(
            status="PROCESSING",
            command=command,
            action="Resuming last project",
            ai_status="READY",
            memory_status="ACTIVE",
            ollama_status="LOCAL"
        )

        result = resume_last_project()

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            action="Last project resumed",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower in [
        "resume last task",
        "continue last task",
    ]:
        update_hud_context(
            status="PROCESSING",
            command=command,
            action="Continuing last task",
            ai_status="READY",
            memory_status="ACTIVE",
            ollama_status="LOCAL"
        )

        result = continue_last_task()

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            action="Last task loaded",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower == "what projects do you remember":
        return what_projects_do_you_remember()

    if command_lower == "compare remembered projects":
        return compare_remembered_projects()

    if command_lower == "last project summary":
        return last_project_summary()

    if command_lower == "project conversation summary":
        return project_conversation_summary()

    if command_lower == "memory summary":
        return memory_summary()

    if command_lower.startswith("remember this "):
        content = command[len("remember this "):].strip()

        return remember(
            "note",
            "Manual note",
            content,
            ["note", "manual"]
        )

    if command_lower == "show recent memory":
        return recent_memories()

    if command_lower.startswith("search memory "):
        keyword = command[len("search memory "):].strip()
        return search_memory(keyword)

    if command_lower == "memory stats":
        return memory_stats()

    # ==========================
    # PROJECT MEMORY
    # IMPORTANT:
    # This block must stay before OPEN PROJECT / OPEN APPLICATION.
    # ==========================
    if command_lower == "show remembered projects":
        return show_remembered_projects()

    if command_lower == "project memory stats":
        return project_memory_stats()

    if command_lower.startswith("remember project "):
        project_name = command[
            len("remember project "):
        ].strip()

        result = analyze_project_by_name(project_name)

        if (
            result.startswith("Project not found")
            or result.startswith("Project index not found")
        ):
            description = f"User project: {project_name}"
        else:
            description = result[:5000]

        memory_result = remember_project(
            project_name,
            description
        )

        remember(
            "project",
            f"Remembered project: {project_name}",
            description[:3000],
            ["project", "memory", project_name]
        )

        return memory_result

    if command_lower.startswith("what do you remember about "):
        project_name = command[
            len("what do you remember about "):
        ].strip()

        return get_project_memory(
            project_name
        )

    if command_lower.startswith("search project memory "):
        keyword = command[
            len("search project memory "):
        ].strip()

        return search_project_memory(
            keyword
        )


    # ==========================
    # DEEP PROJECT MEMORY
    # IMPORTANT:
    # This block must stay before OPEN PROJECT / OPEN APPLICATION.
    # ==========================
    if (
        command_lower.startswith("remember deep project ")
        or command_lower.startswith("refresh deep project ")
    ):
        if command_lower.startswith("remember deep project "):
            project_name = command[
                len("remember deep project "):
            ].strip()
            action_name = "Deep remembered project"
        else:
            project_name = command[
                len("refresh deep project "):
            ].strip()
            action_name = "Deep refreshed project"

        from tools import find_project

        project = find_project(project_name)

        if not project:
            return f"Project not found: {project_name}"

        result = remember_deep_project(
            project["name"],
            project["path"]
        )

        remember(
            "deep_project",
            f"{action_name}: {project['name']}",
            result,
            ["project", "deep", "memory", "refresh", project["name"]]
        )

        return result

    if command_lower == "show deep projects":
        return list_deep_projects()

    if command_lower == "deep project stats":
        return deep_project_stats()

    if command_lower.startswith("what do you remember deeply about "):
        project_name = command[
            len("what do you remember deeply about "):
        ].strip()

        return get_deep_project(project_name)

    if command_lower.startswith("what files are in project "):
        project_name = command[
            len("what files are in project "):
        ].strip()

        return show_project_files(project_name)

    if command_lower.startswith("what is the tech stack of "):
        project_name = command[
            len("what is the tech stack of "):
        ].strip()

        return show_project_tech_stack(project_name)

    if command_lower.startswith("search project code "):
        query = command[
            len("search project code "):
        ].strip()

        return search_deep_project_code(query)

    # ==========================
    # REFRESH APPLICATIONS
    # ==========================
    if command_lower == "refresh applications":
        result = refresh_app_index()

        remember(
            "application",
            "Application index refreshed",
            result,
            ["application", "refresh"]
        )

        return result

    # ==========================
    # REFRESH PROJECTS
    # ==========================
    if command_lower == "refresh projects":
        result = refresh_project_index()

        remember(
            "project",
            "Project index refreshed",
            result,
            ["project", "refresh"]
        )

        return result

    # ==========================
    # PROJECT EXPLORER
    # IMPORTANT:
    # These commands must stay before LLM fallback,
    # so JARVIS does not invent project names.
    # ==========================
    if command_lower in [
        "show projects",
        "show all projects",
        "list all projects",
        "show indexed projects"
    ]:
        result = list_projects_detailed()

        remember(
            "project",
            "Showed indexed projects",
            result[:5000],
            ["project", "list", "detailed"]
        )

        return result

    if command_lower.startswith("show projects from "):
        drive = command[
            len("show projects from "):
        ].strip()

        result = list_projects_by_drive(drive)

        remember(
            "project",
            f"Showed projects from drive: {drive}",
            result[:5000],
            ["project", "list", "drive", drive]
        )

        return result

    if command_lower.startswith("list projects from "):
        drive = command[
            len("list projects from "):
        ].strip()

        result = list_projects_by_drive(drive)

        remember(
            "project",
            f"Listed projects from drive: {drive}",
            result[:5000],
            ["project", "list", "drive", drive]
        )

        return result

    if command_lower.startswith("search projects "):
        keyword = command[
            len("search projects "):
        ].strip()

        result = search_projects(keyword)

        remember(
            "project",
            f"Searched projects: {keyword}",
            result[:5000],
            ["project", "search", keyword]
        )

        return result

    if command_lower.startswith("find project "):
        keyword = command[
            len("find project "):
        ].strip()

        result = search_projects(keyword)

        remember(
            "project",
            f"Found project: {keyword}",
            result[:5000],
            ["project", "find", keyword]
        )

        return result

    # ==========================
    # LIST APPLICATIONS
    # ==========================
    if command_lower == "list applications":
        apps = list_apps()

        result = "\n".join(apps) if apps else "No applications found."

        remember(
            "application",
            "Listed applications",
            result,
            ["application", "list"]
        )

        return result

    # ==========================
    # LIST PROJECTS
    # ==========================
    if command_lower == "list projects":
        projects = list_projects()

        result = "\n".join(projects) if projects else "No projects found."

        remember(
            "project",
            "Listed projects",
            result,
            ["project", "list"]
        )

        return result

    # ==========================
    # READ PDF
    # ==========================
    if command_lower.startswith("read pdf "):
        pdf_name = command[len("read pdf "):].strip()
        result = read_pdf(pdf_name)

        remember(
            "pdf",
            f"Read PDF: {pdf_name}",
            result[:3000],
            ["pdf", pdf_name]
        )

        return result

    # ==========================
    # ANALYZE PDF
    # ==========================
    if command_lower.startswith("analyze pdf "):
        pdf_name = command[len("analyze pdf "):].strip()
        result = analyze_pdf(pdf_name)

        remember(
            "pdf",
            f"Analyzed PDF: {pdf_name}",
            result,
            ["pdf", "analysis", pdf_name]
        )

        return result

    # ==========================
    # OPEN PDF
    # ==========================
    if command_lower.startswith("open pdf "):
        pdf_name = command[len("open pdf "):].strip()
        result = open_pdf(pdf_name)

        remember(
            "pdf",
            f"Opened PDF: {pdf_name}",
            result,
            ["pdf", "open", pdf_name]
        )

        return result

    # ==========================
    # READ SCREEN
    # ==========================
    if command_lower in [
        "read screen",
        "what is on my screen"
    ]:
        result = read_screen()

        remember(
            "screen",
            "Read screen",
            result[:3000],
            ["screen", "ocr"]
        )

        return result

    if command_lower == "read screen center":
        result = read_screen_center()

        remember(
            "screen",
            "Read screen center",
            result[:3000],
            ["screen", "center", "ocr"]
        )

        return result

    if command_lower == "read terminal":
        result = read_terminal()

        remember(
            "screen",
            "Read terminal",
            result[:3000],
            ["screen", "terminal", "ocr"]
        )

        return result

    if command_lower == "read browser":
        result = read_browser()

        remember(
            "screen",
            "Read browser",
            result[:3000],
            ["screen", "browser", "ocr"]
        )

        return result

    # ==========================
    # ANALYZE SCREEN
    # ==========================
    if command_lower in [
        "analyze screen",
        "analyze my screen"
    ]:
        result = analyze_screen()

        remember(
            "screen",
            "Analyzed screen",
            result,
            ["screen", "analysis"]
        )

        return result

    if command_lower == "analyze screen center":
        result = analyze_screen_center()

        remember(
            "screen",
            "Analyzed screen center",
            result,
            ["screen", "center", "analysis"]
        )

        return result

    if command_lower == "analyze terminal":
        result = analyze_terminal()

        remember(
            "screen",
            "Analyzed terminal",
            result,
            ["screen", "terminal", "analysis"]
        )

        return result

    if command_lower == "analyze browser":
        result = analyze_browser()

        remember(
            "screen",
            "Analyzed browser",
            result,
            ["screen", "browser", "analysis"]
        )

        return result


    # ==========================
    # SCREEN VISION UPGRADE
    # ==========================
    if command_lower == "read code on screen":
        update_hud_context(
            status="PROCESSING",
            command=command,
            action="Reading code on screen",
            ai_status="READY",
            vision_status="ACTIVE",
            memory_status="SYNC",
            ollama_status="LOCAL"
        )

        result = read_code_on_screen()

        remember(
            "screen",
            "Read code on screen",
            result[:3000],
            ["screen", "code", "ocr"]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            action="Code OCR completed",
            ai_status="READY",
            vision_status="ACTIVE",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower == "read error on screen":
        update_hud_context(
            status="PROCESSING",
            command=command,
            action="Reading error on screen",
            ai_status="READY",
            vision_status="ACTIVE",
            memory_status="SYNC",
            ollama_status="LOCAL"
        )

        result = read_error_on_screen()

        remember(
            "screen",
            "Read error on screen",
            result[:3000],
            ["screen", "error", "ocr"]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            action="Error OCR completed",
            ai_status="READY",
            vision_status="ACTIVE",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower in [
        "review code on screen",
        "review this code",
        "analyze code on screen"
    ]:
        update_hud_context(
            status="PROCESSING",
            command=command,
            action="Reviewing code on screen",
            ai_status="THINKING",
            vision_status="ACTIVE",
            memory_status="SYNC",
            ollama_status="THINKING"
        )

        result = review_code_on_screen()

        remember(
            "screen",
            "Reviewed code on screen",
            result[:5000],
            ["screen", "code", "review"]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            action="Code screen review completed",
            ai_status="READY",
            vision_status="ACTIVE",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower in [
        "find bugs on screen",
        "find bugs on the screen",
        "find bugs in code on screen"
    ]:
        update_hud_context(
            status="PROCESSING",
            command=command,
            action="Finding bugs on screen",
            ai_status="THINKING",
            vision_status="ACTIVE",
            memory_status="SYNC",
            ollama_status="THINKING"
        )

        result = find_bugs_on_screen()

        remember(
            "screen",
            "Found bugs on screen",
            result[:5000],
            ["screen", "bugs", "code"]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            action="Bug scan completed",
            ai_status="READY",
            vision_status="ACTIVE",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower in [
        "explain error on screen",
        "explain this error",
        "what error is on screen"
    ]:
        update_hud_context(
            status="PROCESSING",
            command=command,
            action="Explaining screen error",
            ai_status="THINKING",
            vision_status="ACTIVE",
            memory_status="SYNC",
            ollama_status="THINKING"
        )

        result = explain_error_on_screen()

        remember(
            "screen",
            "Explained error on screen",
            result[:5000],
            ["screen", "error", "debug"]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            action="Error explanation completed",
            ai_status="READY",
            vision_status="ACTIVE",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    # ==========================
    # SCREENSHOTS
    # ==========================
    if command_lower == "take screenshot":
        result = take_screenshot()

        remember(
            "screenshot",
            "Took screenshot",
            result,
            ["screenshot"]
        )

        return result

    if command_lower == "save screenshot":
        result = save_screenshot()

        remember(
            "screenshot",
            "Saved screenshot",
            result,
            ["screenshot", "save"]
        )

        return result

    if command_lower == "describe screenshot":
        result = describe_screenshot()

        remember(
            "screenshot",
            "Described screenshot",
            result,
            ["screenshot", "description"]
        )

        return result

    # ==========================
    # SCREEN MEMORY
    # ==========================
    if command_lower == "remember screenshot":
        result = remember_current_screenshot()

        remember(
            "screenshot",
            "Remembered screenshot",
            result,
            ["screenshot", "memory"]
        )

        return result

    if command_lower == "what was on my screen":
        return what_was_on_my_screen()

    if command_lower.startswith("search screenshots "):
        keyword = command[
            len("search screenshots "):
        ].strip()

        return search_screenshot_memory(keyword)

    # ==========================
    # MEMORY STATS
    # ==========================
    if command_lower == "how many screenshots do you remember":
        data = load_memory()

        if not data:
            return "JARVIS remembers 0 screenshots."

        first = data[0].get("timestamp", "Unknown")
        last = data[-1].get("timestamp", "Unknown")

        return (
            f"JARVIS remembers {len(data)} screenshots.\n"
            f"First memory: {first}\n"
            f"Last memory: {last}"
        )

    # ==========================
    # RECENT SCREENSHOTS
    # ==========================
    if command_lower == "show recent screenshots":
        data = load_memory()

        if not data:
            return "No screenshots remembered."

        output = []

        for item in data[-10:]:
            output.append(
                f"{item.get('timestamp', 'Unknown')} -> "
                f"{item.get('summary', 'No summary')}"
            )

        return "\n".join(output)

    # ==========================
    # SMART SCREENSHOT SEARCH
    # ==========================
    if command_lower.startswith("find screenshot "):
        keyword = command[
            len("find screenshot "):
        ].strip()

        return search_screenshot_memory(keyword)


    # ==========================
    # PROJECT ASSISTANT INSPECTION COMMANDS
    # ==========================
    if command_lower.startswith("show project structure "):
        project_name = command[
            len("show project structure "):
        ].strip()

        update_hud_context(
            status="PROCESSING",
            command=command,
            project=project_name,
            action="Showing project structure",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="LOCAL"
        )

        result = show_project_structure_local(project_name)

        remember(
            "project_assistant",
            f"Showed project structure: {project_name}",
            result[:5000],
            ["project", "structure", project_name]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            project=project_name,
            action="Project structure ready",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower.startswith("project structure "):
        project_name = command[
            len("project structure "):
        ].strip()

        return show_project_structure_local(project_name)

    if command_lower.startswith("show project files "):
        project_name = command[
            len("show project files "):
        ].strip()

        update_hud_context(
            status="PROCESSING",
            command=command,
            project=project_name,
            action="Listing project files",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="LOCAL"
        )

        result = show_project_files_local(project_name)

        remember(
            "project_assistant",
            f"Showed project files: {project_name}",
            result[:5000],
            ["project", "files", project_name]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            project=project_name,
            action="Project files listed",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower.startswith("project files "):
        project_name = command[
            len("project files "):
        ].strip()

        return show_project_files_local(project_name)

    if command_lower.startswith("show project statistics "):
        project_name = command[
            len("show project statistics "):
        ].strip()

        result = show_project_statistics_local(project_name)

        remember(
            "project_assistant",
            f"Showed project statistics: {project_name}",
            result[:5000],
            ["project", "statistics", project_name]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            project=project_name,
            action="Project statistics ready",
            ai_status="READY",
            project_score="STATS",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower.startswith("show project stats "):
        project_name = command[
            len("show project stats "):
        ].strip()

        return show_project_statistics_local(project_name)

    if command_lower.startswith("project statistics "):
        project_name = command[
            len("project statistics "):
        ].strip()

        return show_project_statistics_local(project_name)

    if command_lower.startswith("project stats "):
        project_name = command[
            len("project stats "):
        ].strip()

        return show_project_statistics_local(project_name)

    if command_lower.startswith("show largest files "):
        project_name = command[
            len("show largest files "):
        ].strip()

        result = show_largest_project_files_local(project_name)

        remember(
            "project_assistant",
            f"Showed largest files: {project_name}",
            result[:5000],
            ["project", "largest", "files", project_name]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            project=project_name,
            action="Largest files ready",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower.startswith("largest files "):
        project_name = command[
            len("largest files "):
        ].strip()

        return show_largest_project_files_local(project_name)

    if command_lower.startswith("show security report "):
        project_name = command[
            len("show security report "):
        ].strip()

        update_hud_context(
            status="PROCESSING",
            command=command,
            project=project_name,
            action="Generating security report",
            ai_status="THINKING",
            memory_status="SYNC",
            ollama_status="THINKING"
        )

        result = find_security_issues(project_name)

        remember(
            "project_assistant",
            f"Security report: {project_name}",
            result[:5000],
            ["project", "security", "report", project_name]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            project=project_name,
            action="Security report ready",
            ai_status="READY",
            security_score=extract_hud_security_score(result),
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower.startswith("security report "):
        project_name = command[
            len("security report "):
        ].strip()

        return find_security_issues(project_name)

    if command_lower in [
        "show backups",
        "show file backups",
        "show project backups"
    ]:
        result = show_all_backups()

        remember(
            "safe_edit",
            "Showed all backups",
            result[:5000],
            ["backup", "show"]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            action="Backups listed",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower.startswith("restore backup "):
        file_query = command[
            len("restore backup "):
        ].strip()

        result = restore_latest_backup(file_query)

        remember(
            "safe_edit",
            f"Restored backup: {file_query}",
            result[:3000],
            ["backup", "restore", file_query]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            file_path=file_query,
            action="Backup restore attempted",
            ai_status="READY"
        )

        return result


    # ==========================
    # STEP 8 - AUTONOMOUS CODING ASSISTANT
    # Safe autonomous analysis. No automatic code changes.
    # ==========================
    if command_lower.startswith("autonomous review "):
        project_name = command[
            len("autonomous review "):
        ].strip()

        hud_start_action(
            command,
            "Autonomous project review",
            project=project_name,
            thinking=True
        )

        result = autonomous_review_project(project_name)

        remember(
            "autonomous_coding",
            f"Autonomous review: {project_name}",
            result[:10000],
            ["autonomous", "review", project_name]
        )

        hud_finish_action(
            command,
            result,
            "Autonomous review completed",
            project=project_name,
            project_score=extract_hud_project_score(result),
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("review everything "):
        project_name = command[
            len("review everything "):
        ].strip()

        hud_start_action(
            command,
            "Reviewing everything",
            project=project_name,
            thinking=True
        )

        result = review_everything(project_name)

        remember(
            "autonomous_coding",
            f"Review everything: {project_name}",
            result[:10000],
            ["autonomous", "review_everything", project_name]
        )

        hud_finish_action(
            command,
            result,
            "Review everything completed",
            project=project_name,
            project_score=extract_hud_project_score(result),
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower == "review everything":
        project_name = "CyberShield AI"

        hud_start_action(
            command,
            "Reviewing everything",
            project=project_name,
            thinking=True
        )

        result = review_everything(project_name)

        remember(
            "autonomous_coding",
            f"Review everything: {project_name}",
            result[:10000],
            ["autonomous", "review_everything", project_name]
        )

        hud_finish_action(
            command,
            result,
            "Review everything completed",
            project=project_name,
            project_score=extract_hud_project_score(result),
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("autonomous improve "):
        project_name = command[
            len("autonomous improve "):
        ].strip()

        hud_start_action(
            command,
            "Autonomous improvement plan",
            project=project_name,
            thinking=True
        )

        result = autonomous_improve_project(project_name)

        remember(
            "autonomous_coding",
            f"Autonomous improve: {project_name}",
            result[:12000],
            ["autonomous", "improve", project_name]
        )

        remember_project_event(
            project_name,
            "improvement",
            "Autonomous improvement plan",
            result[:12000],
            tags=["autonomous", "improve", "roadmap"],
            metadata={"source": "jarvis_agent_step10"}
        )

        hud_finish_action(
            command,
            result,
            "Autonomous improvement completed",
            project=project_name,
            project_score=extract_hud_project_score(result),
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("autonomous secure "):
        project_name = command[
            len("autonomous secure "):
        ].strip()

        hud_start_action(
            command,
            "Autonomous security hardening",
            project=project_name,
            thinking=True
        )

        result = autonomous_secure_project(project_name)

        remember(
            "autonomous_coding",
            f"Autonomous secure: {project_name}",
            result[:12000],
            ["autonomous", "secure", project_name]
        )

        hud_finish_action(
            command,
            result,
            "Autonomous security completed",
            project=project_name,
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("autonomous optimize "):
        project_name = command[
            len("autonomous optimize "):
        ].strip()

        hud_start_action(
            command,
            "Autonomous optimization plan",
            project=project_name,
            thinking=True
        )

        result = autonomous_optimize_project(project_name)

        remember(
            "autonomous_coding",
            f"Autonomous optimize: {project_name}",
            result[:12000],
            ["autonomous", "optimize", project_name]
        )

        hud_finish_action(
            command,
            result,
            "Autonomous optimization completed",
            project=project_name,
            project_score=extract_hud_project_score(result)
        )

        return result

    if command_lower.startswith("autonomous fix "):
        project_name = command[
            len("autonomous fix "):
        ].strip()

        hud_start_action(
            command,
            "Autonomous fix plan",
            project=project_name,
            thinking=True
        )

        result = autonomous_fix_project(project_name)

        remember(
            "autonomous_coding",
            f"Autonomous fix: {project_name}",
            result[:12000],
            ["autonomous", "fix", project_name]
        )

        remember_project_event(
            project_name,
            "fix",
            "Autonomous fix plan",
            result[:12000],
            tags=["autonomous", "fix", "patch_plan"],
            metadata={"source": "jarvis_agent_step10"}
        )

        hud_finish_action(
            command,
            result,
            "Autonomous fix plan completed",
            project=project_name,
            project_score=extract_hud_project_score(result),
            security_score=extract_hud_security_score(result)
        )

        return result

    # ==========================
    # STEP 7 - IRON MAN DEVELOPER MODE
    # Natural security commands routed to Step 6 auditor.
    # ==========================
    if command_lower in [
        "secure cyber shield ai",
        "secure cybershield ai",
        "secure cyber",
        "secure project",
        "secure my project",
        "scan cyber shield ai",
        "scan cybershield ai",
        "scan cyber",
        "scan project",
        "scan my project",
        "find vulnerabilities",
        "find security issues",
        "find security problems",
        "check vulnerabilities",
        "check security",
    ]:
        project_name = "CyberShield AI"

        hud_start_action(
            command,
            "Iron Man security scan",
            project=project_name,
            thinking=False
        )

        result = full_security_audit(project_name)

        remember(
            "security_audit",
            f"Iron Man security scan: {project_name}",
            result[:8000],
            ["security", "ironman", "audit", project_name]
        )

        hud_finish_action(
            command,
            result,
            "Iron Man security scan completed",
            project=project_name,
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower in [
        "audit cyber shield ai",
        "audit cybershield ai",
        "audit cyber",
        "audit project",
        "audit my project",
    ]:
        project_name = "CyberShield AI"

        hud_start_action(
            command,
            "Enterprise audit",
            project=project_name,
            thinking=False
        )

        result = full_security_audit(project_name)

        remember(
            "security_audit",
            f"Enterprise audit: {project_name}",
            result[:8000],
            ["security", "enterprise", "audit", project_name]
        )

        hud_finish_action(
            command,
            result,
            "Enterprise audit completed",
            project=project_name,
            security_score=extract_hud_security_score(result)
        )

        return result

    # ==========================
    # STEP 6 - AI SECURITY AUDITOR
    # Rule-based project security scans. No automatic code changes.
    # ==========================
    if command_lower.startswith("find api keys "):
        project_name = command[
            len("find api keys "):
        ].strip()

        hud_start_action(
            command,
            "Scanning API keys",
            project=project_name,
            thinking=False
        )

        result = find_api_keys(project_name)

        remember(
            "security_audit",
            f"API key scan: {project_name}",
            result[:5000],
            ["security", "api_keys", project_name]
        )

        hud_finish_action(
            command,
            result,
            "API key scan completed",
            project=project_name,
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("scan api keys "):
        project_name = command[
            len("scan api keys "):
        ].strip()

        return find_api_keys(project_name)

    if command_lower.startswith("find passwords "):
        project_name = command[
            len("find passwords "):
        ].strip()

        hud_start_action(
            command,
            "Scanning passwords",
            project=project_name,
            thinking=False
        )

        result = find_passwords(project_name)

        remember(
            "security_audit",
            f"Password scan: {project_name}",
            result[:5000],
            ["security", "passwords", project_name]
        )

        hud_finish_action(
            command,
            result,
            "Password scan completed",
            project=project_name,
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("scan passwords "):
        project_name = command[
            len("scan passwords "):
        ].strip()

        return find_passwords(project_name)

    if command_lower.startswith("find hardcoded secrets "):
        project_name = command[
            len("find hardcoded secrets "):
        ].strip()

        hud_start_action(
            command,
            "Scanning hardcoded secrets",
            project=project_name,
            thinking=False
        )

        result = find_hardcoded_secrets(project_name)

        remember(
            "security_audit",
            f"Hardcoded secrets scan: {project_name}",
            result[:5000],
            ["security", "secrets", project_name]
        )

        hud_finish_action(
            command,
            result,
            "Hardcoded secrets scan completed",
            project=project_name,
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("find secrets "):
        project_name = command[
            len("find secrets "):
        ].strip()

        return find_hardcoded_secrets(project_name)

    if command_lower.startswith("scan hardcoded secrets "):
        project_name = command[
            len("scan hardcoded secrets "):
        ].strip()

        return find_hardcoded_secrets(project_name)

    if command_lower.startswith("scan secrets "):
        project_name = command[
            len("scan secrets "):
        ].strip()

        return find_hardcoded_secrets(project_name)

    if command_lower.startswith("find sql injection "):
        project_name = command[
            len("find sql injection "):
        ].strip()

        hud_start_action(
            command,
            "Scanning SQL injection risks",
            project=project_name,
            thinking=False
        )

        result = find_sql_injection(project_name)

        remember(
            "security_audit",
            f"SQL injection scan: {project_name}",
            result[:5000],
            ["security", "sql_injection", project_name]
        )

        hud_finish_action(
            command,
            result,
            "SQL injection scan completed",
            project=project_name,
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("scan sql injection "):
        project_name = command[
            len("scan sql injection "):
        ].strip()

        return find_sql_injection(project_name)

    if command_lower.startswith("find xss "):
        project_name = command[
            len("find xss "):
        ].strip()

        hud_start_action(
            command,
            "Scanning XSS risks",
            project=project_name,
            thinking=False
        )

        result = find_xss_risks(project_name)

        remember(
            "security_audit",
            f"XSS scan: {project_name}",
            result[:5000],
            ["security", "xss", project_name]
        )

        hud_finish_action(
            command,
            result,
            "XSS scan completed",
            project=project_name,
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("scan xss "):
        project_name = command[
            len("scan xss "):
        ].strip()

        return find_xss_risks(project_name)

    if command_lower.startswith("find dangerous imports "):
        project_name = command[
            len("find dangerous imports "):
        ].strip()

        hud_start_action(
            command,
            "Scanning dangerous imports",
            project=project_name,
            thinking=False
        )

        result = find_dangerous_imports(project_name)

        remember(
            "security_audit",
            f"Dangerous imports scan: {project_name}",
            result[:5000],
            ["security", "dangerous_imports", project_name]
        )

        hud_finish_action(
            command,
            result,
            "Dangerous imports scan completed",
            project=project_name,
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("scan dangerous imports "):
        project_name = command[
            len("scan dangerous imports "):
        ].strip()

        return find_dangerous_imports(project_name)

    if command_lower.startswith("find vulnerable files "):
        project_name = command[
            len("find vulnerable files "):
        ].strip()

        return full_security_audit(project_name)

    if command_lower.startswith("scan vulnerable files "):
        project_name = command[
            len("scan vulnerable files "):
        ].strip()

        return full_security_audit(project_name)

    if command_lower.startswith("full security audit "):
        project_name = command[
            len("full security audit "):
        ].strip()

        hud_start_action(
            command,
            "Running full security audit",
            project=project_name,
            thinking=False
        )

        result = full_security_audit(project_name)

        remember(
            "security_audit",
            f"Full security audit: {project_name}",
            result[:8000],
            ["security", "full_audit", project_name]
        )

        remember_project_event(
            project_name,
            "audit",
            "Full security audit",
            result[:12000],
            tags=["security", "audit", "full_audit"],
            metadata={"source": "jarvis_agent_step10"}
        )

        hud_finish_action(
            command,
            result,
            "Full security audit completed",
            project=project_name,
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("enterprise audit "):
        project_name = command[
            len("enterprise audit "):
        ].strip()

        hud_start_action(
            command,
            "Enterprise audit",
            project=project_name,
            thinking=False
        )

        result = full_security_audit(project_name)

        hud_finish_action(
            command,
            result,
            "Enterprise audit completed",
            project=project_name,
            security_score=extract_hud_security_score(result)
        )

        remember(
            "security_audit",
            f"Enterprise audit: {project_name}",
            result[:8000],
            ["security", "enterprise", "audit", project_name]
        )

        return result

    if command_lower.startswith("scan entire project "):
        project_name = command[
            len("scan entire project "):
        ].strip()

        return full_security_audit(project_name)

    if command_lower.startswith("generate security roadmap "):
        project_name = command[
            len("generate security roadmap "):
        ].strip()

        hud_start_action(
            command,
            "Generating security roadmap",
            project=project_name,
            thinking=False
        )

        result = generate_security_roadmap(project_name)

        remember(
            "security_audit",
            f"Security roadmap: {project_name}",
            result[:8000],
            ["security", "roadmap", project_name]
        )

        remember_project_event(
            project_name,
            "roadmap",
            "Security roadmap",
            result[:12000],
            tags=["security", "roadmap", "improvement"],
            metadata={"source": "jarvis_agent_step10"}
        )

        hud_finish_action(
            command,
            result,
            "Security roadmap completed",
            project=project_name,
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("security roadmap "):
        project_name = command[
            len("security roadmap "):
        ].strip()

        return generate_security_roadmap(project_name)

    # ==========================
    # SUGGEST FIXES
    # Rule-based recommendations. No automatic code changes.
    # ==========================
    if command_lower.startswith("suggest fixes for project "):
        project_name = command[
            len("suggest fixes for project "):
        ].strip()

        result = suggest_fixes_for_project(project_name)

        remember(
            "project_review",
            f"Suggested fixes for project: {project_name}",
            result[:5000],
            ["project", "suggest", "fixes", project_name]
        )

        return result

    if command_lower.startswith("suggest fixes "):
        project_name = command[
            len("suggest fixes "):
        ].strip()

        result = suggest_fixes_for_project(project_name)

        remember(
            "project_review",
            f"Suggested fixes: {project_name}",
            result[:5000],
            ["project", "suggest", "fixes", project_name]
        )

        return result

    if command_lower.startswith("project fixes "):
        project_name = command[
            len("project fixes "):
        ].strip()

        result = suggest_fixes_for_project(project_name)

        remember(
            "project_review",
            f"Project fixes: {project_name}",
            result[:5000],
            ["project", "fixes", project_name]
        )

        return result

    if command_lower.startswith("fix suggestions for project "):
        project_name = command[
            len("fix suggestions for project "):
        ].strip()

        result = suggest_fixes_for_project(project_name)

        remember(
            "project_review",
            f"Fix suggestions for project: {project_name}",
            result[:5000],
            ["project", "fix", "suggestions", project_name]
        )

        return result

    # ==========================
    # SMART DAILY PROJECT CHECK
    # Gives a daily portfolio/project status.
    # ==========================
    if command_lower in [
        "daily project check",
        "daily check",
        "jarvis daily check",
        "smart daily check"
    ]:
        result = daily_project_check()

        remember(
            "project_review",
            "Smart daily project check",
            result[:5000],
            ["project", "daily", "check"]
        )

        return result

    # ==========================
    # PROJECT REPORT EXPORT
    # Saves full Markdown report in reports/
    # ==========================
    if command_lower.startswith("export project report "):
        project_name = command[
            len("export project report "):
        ].strip()

        result = export_project_report(project_name)

        remember(
            "project_review",
            f"Exported project report: {project_name}",
            result[:5000],
            ["project", "export", "report", project_name]
        )

        return result

    if command_lower.startswith("export report "):
        project_name = command[
            len("export report "):
        ].strip()

        result = export_project_report(project_name)

        remember(
            "project_review",
            f"Exported report: {project_name}",
            result[:5000],
            ["project", "export", "report", project_name]
        )

        return result

    if command_lower.startswith("save project report "):
        project_name = command[
            len("save project report "):
        ].strip()

        result = export_project_report(project_name)

        remember(
            "project_review",
            f"Saved project report: {project_name}",
            result[:5000],
            ["project", "save", "report", project_name]
        )

        return result

    # ==========================
    # CROSS PROJECT INTELLIGENCE
    # Rule-based comparison/ranking, no LLM.
    # ==========================
    if command_lower.startswith("score project "):
        project_name = command[
            len("score project "):
        ].strip()

        hud_start_action(
            command,
            "Scoring project",
            project=project_name,
            thinking=False
        )

        result = score_project(project_name)

        remember(
            "project_review",
            f"Scored project: {project_name}",
            result[:5000],
            ["project", "score", project_name]
        )

        hud_finish_action(
            command,
            result,
            "Project score ready",
            project=project_name,
            project_score=extract_hud_project_score(result)
        )

        return result

    if command_lower.startswith("compare projects "):
        projects_text = command[
            len("compare projects "):
        ].strip()

        result = compare_projects(projects_text)

        remember(
            "project_review",
            f"Compared projects: {projects_text}",
            result[:5000],
            ["project", "compare", projects_text]
        )

        return result

    if command_lower.startswith("compare security "):
        projects_text = command[
            len("compare security "):
        ].strip()

        result = compare_security(projects_text)

        remember(
            "project_review",
            f"Compared project security: {projects_text}",
            result[:5000],
            ["project", "compare", "security", projects_text]
        )

        return result

    if command_lower.startswith("compare architecture "):
        projects_text = command[
            len("compare architecture "):
        ].strip()

        result = compare_architecture(projects_text)

        remember(
            "project_review",
            f"Compared project architecture: {projects_text}",
            result[:5000],
            ["project", "compare", "architecture", projects_text]
        )

        return result

    if command_lower == "rank projects by security":
        result = rank_projects_by_security()

        remember(
            "project_review",
            "Ranked projects by security",
            result[:5000],
            ["project", "rank", "security"]
        )

        return result

    if command_lower == "rank projects by architecture":
        result = rank_projects_by_architecture()

        remember(
            "project_review",
            "Ranked projects by architecture",
            result[:5000],
            ["project", "rank", "architecture"]
        )

        return result

    if command_lower == "rank projects by maintainability":
        result = rank_projects_by_maintainability()

        remember(
            "project_review",
            "Ranked projects by maintainability",
            result[:5000],
            ["project", "rank", "maintainability"]
        )

        return result

    if command_lower == "best project in memory":
        result = best_project_in_memory()

        remember(
            "project_review",
            "Best project in memory",
            result[:5000],
            ["project", "best", "memory"]
        )

        return result

    # ==========================
    # STRICT GROUNDED ANALYZER
    # Rule-based analyzer, no LLM, no speculation.
    # ==========================
    if command_lower.startswith("strict grounded analyzer project "):
        project_name = command[
            len("strict grounded analyzer project "):
        ].strip()

        result = strict_grounded_analyzer_project(project_name)

        remember(
            "project_review",
            f"Strict grounded analyzer project: {project_name}",
            result[:5000],
            ["project", "strict", "grounded", "analyzer", project_name]
        )

        return result

    if command_lower.startswith("strict security analyzer project "):
        project_name = command[
            len("strict security analyzer project "):
        ].strip()

        result = strict_security_analyzer_project(project_name)

        remember(
            "project_review",
            f"Strict security analyzer project: {project_name}",
            result[:5000],
            ["project", "strict", "security", "analyzer", project_name]
        )

        return result

    if command_lower.startswith("strict architecture analyzer project "):
        project_name = command[
            len("strict architecture analyzer project "):
        ].strip()

        result = strict_architecture_analyzer_project(project_name)

        remember(
            "project_review",
            f"Strict architecture analyzer project: {project_name}",
            result[:5000],
            ["project", "strict", "architecture", "analyzer", project_name]
        )

        return result

    # Short aliases
    if command_lower.startswith("strict analyzer project "):
        project_name = command[
            len("strict analyzer project "):
        ].strip()

        result = strict_grounded_analyzer_project(project_name)

        remember(
            "project_review",
            f"Strict analyzer project: {project_name}",
            result[:5000],
            ["project", "strict", "analyzer", project_name]
        )

        return result

    if command_lower.startswith("strict security project "):
        project_name = command[
            len("strict security project "):
        ].strip()

        result = strict_security_analyzer_project(project_name)

        remember(
            "project_review",
            f"Strict security project: {project_name}",
            result[:5000],
            ["project", "strict", "security", project_name]
        )

        return result

    if command_lower.startswith("strict architecture project "):
        project_name = command[
            len("strict architecture project "):
        ].strip()

        result = strict_architecture_analyzer_project(project_name)

        remember(
            "project_review",
            f"Strict architecture project: {project_name}",
            result[:5000],
            ["project", "strict", "architecture", project_name]
        )

        return result

    # ==========================
    # PROJECT EVIDENCE + GROUNDED REVIEW
    # These commands force project analysis to use real indexed files.
    # ==========================
    if command_lower.startswith("project evidence "):
        project_name = command[
            len("project evidence "):
        ].strip()

        result = project_evidence_report(project_name)

        remember(
            "project_review",
            f"Project evidence: {project_name}",
            result[:5000],
            ["project", "evidence", project_name]
        )

        return result

    if command_lower.startswith("evidence project "):
        project_name = command[
            len("evidence project "):
        ].strip()

        result = project_evidence_report(project_name)

        remember(
            "project_review",
            f"Evidence project: {project_name}",
            result[:5000],
            ["project", "evidence", project_name]
        )

        return result

    if command_lower.startswith("grounded review project "):
        project_name = command[
            len("grounded review project "):
        ].strip()

        result = grounded_review_project(project_name)

        remember(
            "project_review",
            f"Grounded review project: {project_name}",
            result[:5000],
            ["project", "grounded", "review", project_name]
        )

        return result

    if command_lower.startswith("grounded security review project "):
        project_name = command[
            len("grounded security review project "):
        ].strip()

        result = grounded_security_review_project(project_name)

        remember(
            "project_review",
            f"Grounded security review project: {project_name}",
            result[:5000],
            ["project", "grounded", "security", project_name]
        )

        return result

    if command_lower.startswith("grounded architecture review project "):
        project_name = command[
            len("grounded architecture review project "):
        ].strip()

        result = grounded_architecture_review_project(project_name)

        remember(
            "project_review",
            f"Grounded architecture review project: {project_name}",
            result[:5000],
            ["project", "grounded", "architecture", project_name]
        )

        return result

    if command_lower.startswith("grounded docs project "):
        project_name = command[
            len("grounded docs project "):
        ].strip()

        result = grounded_documentation_project(project_name)

        remember(
            "project_review",
            f"Grounded docs project: {project_name}",
            result[:5000],
            ["project", "grounded", "docs", project_name]
        )

        return result

    if command_lower.startswith("grounded documentation project "):
        project_name = command[
            len("grounded documentation project "):
        ].strip()

        result = grounded_documentation_project(project_name)

        remember(
            "project_review",
            f"Grounded documentation project: {project_name}",
            result[:5000],
            ["project", "grounded", "documentation", project_name]
        )

        return result

    # ==========================
    # PROJECT REVIEW ASSISTANT - EARLY ROUTING
    # IMPORTANT:
    # This block must stay before MEMORY FILE ASSISTANT because
    # "security review project ..." also starts with "security review ".
    # ==========================
    if command_lower.startswith("review project "):
        project_name = command[
            len("review project "):
        ].strip()

        hud_start_action(
            command,
            "Reviewing project",
            project=project_name,
            thinking=True
        )

        result = review_project(project_name)

        remember(
            "project_review",
            f"Reviewed project: {project_name}",
            result[:5000],
            ["project", "review", project_name]
        )

        hud_finish_action(
            command,
            result,
            "Project review completed",
            project=project_name,
            project_score=extract_hud_project_score(result)
        )

        return result

    if command_lower.startswith("security review project "):
        project_name = command[
            len("security review project "):
        ].strip()

        hud_start_action(
            command,
            "Security review project",
            project=project_name,
            thinking=True
        )

        result = find_security_issues(project_name)

        remember(
            "project_review",
            f"Security review project: {project_name}",
            result[:5000],
            ["project", "security", "review", project_name]
        )

        hud_finish_action(
            command,
            result,
            "Security review completed",
            project=project_name,
            security_score=extract_hud_security_score(result)
        )

        return result

    if command_lower.startswith("find security issues in project "):
        project_name = command[
            len("find security issues in project "):
        ].strip()

        result = find_security_issues(project_name)

        remember(
            "project_review",
            f"Security issues in project: {project_name}",
            result[:5000],
            ["project", "security", project_name]
        )

        return result

    if command_lower.startswith("find dead code in project "):
        project_name = command[
            len("find dead code in project "):
        ].strip()

        result = find_dead_code(project_name)

        remember(
            "project_review",
            f"Dead code scan project: {project_name}",
            result[:5000],
            ["project", "dead_code", project_name]
        )

        return result

    if command_lower.startswith("find duplicates in project "):
        project_name = command[
            len("find duplicates in project "):
        ].strip()

        result = find_duplicate_code(project_name)

        remember(
            "project_review",
            f"Duplicate code scan project: {project_name}",
            result[:5000],
            ["project", "duplicates", project_name]
        )

        return result

    if command_lower.startswith("generate architecture report for "):
        project_name = command[
            len("generate architecture report for "):
        ].strip()

        result = generate_architecture_report(project_name)

        remember(
            "project_review",
            f"Architecture report project: {project_name}",
            result[:5000],
            ["project", "architecture", project_name]
        )

        return result

    if command_lower.startswith("generate improvement roadmap for "):
        project_name = command[
            len("generate improvement roadmap for "):
        ].strip()

        result = generate_improvement_roadmap(project_name)

        remember(
            "project_review",
            f"Improvement roadmap project: {project_name}",
            result[:5000],
            ["project", "roadmap", project_name]
        )

        return result

    if command_lower.startswith("optimize project "):
        project_name = command[
            len("optimize project "):
        ].strip()

        result = optimize_project(project_name)

        remember(
            "project_review",
            f"Optimized project review: {project_name}",
            result[:5000],
            ["project", "optimization", project_name]
        )

        return result


    # ==========================
    # UNIVERSAL FOLDER ROUTING
    # ==========================
    if command_lower.startswith("rank folder "):
        folder_query = command[
            len("rank folder "):
        ].strip()

        result = rank_universal_folder(folder_query)

        remember(
            "universal_folder",
            f"Ranked folder matches: {folder_query}",
            result[:5000],
            ["folder", "rank", folder_query]
        )

        return result

    if command_lower.startswith("open folder "):
        folder_query = command[
            len("open folder "):
        ].strip()

        result = open_universal_folder(folder_query)

        remember(
            "universal_folder",
            f"Opened folder: {folder_query}",
            result[:3000],
            ["folder", "open", folder_query]
        )

        return result

    if command_lower.startswith("safe preview file "):
        file_query = command[
            len("safe preview file "):
        ].strip()

        result = safe_preview_file(file_query)

        remember(
            "universal_file",
            f"Safe preview file: {file_query}",
            result[:3000],
            ["file", "safe", "preview", file_query]
        )

        return result

    if command_lower.startswith("preview file "):
        file_query = command[
            len("preview file "):
        ].strip()

        result = safe_preview_file(file_query)

        remember(
            "universal_file",
            f"Preview file: {file_query}",
            result[:3000],
            ["file", "preview", file_query]
        )

        return result


    # ==========================
    # SAFE CODE EDIT + BACKUP ROUTING
    # These commands never modify code without backup support.
    # ==========================
    if command_lower in ["list backups", "list file backups"]:
        result = list_file_backups()

        remember(
            "safe_edit",
            "Listed file backups",
            result[:3000],
            ["file", "backup", "list"]
        )

        return result

    if command_lower.startswith("backup file "):
        file_query = command[
            len("backup file "):
        ].strip()

        result = create_file_backup(file_query)

        remember(
            "safe_edit",
            f"Backup file: {file_query}",
            result[:3000],
            ["file", "backup", file_query]
        )

        return result

    if command_lower.startswith("create backup file "):
        file_query = command[
            len("create backup file "):
        ].strip()

        result = create_file_backup(file_query)

        remember(
            "safe_edit",
            f"Create backup file: {file_query}",
            result[:3000],
            ["file", "backup", file_query]
        )

        return result

    if command_lower.startswith("create backup for file "):
        file_query = command[
            len("create backup for file "):
        ].strip()

        result = create_file_backup(file_query)

        remember(
            "safe_edit",
            f"Create backup for file: {file_query}",
            result[:3000],
            ["file", "backup", file_query]
        )

        return result

    if command_lower.startswith("restore latest backup file "):
        file_query = command[
            len("restore latest backup file "):
        ].strip()

        result = restore_latest_backup(file_query)

        remember(
            "safe_edit",
            f"Restore latest backup file: {file_query}",
            result[:3000],
            ["file", "restore", file_query]
        )

        return result

    if command_lower.startswith("restore backup file "):
        file_query = command[
            len("restore backup file "):
        ].strip()

        result = restore_latest_backup(file_query)

        remember(
            "safe_edit",
            f"Restore backup file: {file_query}",
            result[:3000],
            ["file", "restore", file_query]
        )

        return result

    if command_lower.startswith("restore file "):
        file_query = command[
            len("restore file "):
        ].strip()

        result = restore_latest_backup(file_query)

        remember(
            "safe_edit",
            f"Restore file: {file_query}",
            result[:3000],
            ["file", "restore", file_query]
        )

        return result

    if command_lower.startswith("suggest safe patch file "):
        file_query = command[
            len("suggest safe patch file "):
        ].strip()

        result = suggest_safe_patch_for_file(file_query)

        remember(
            "safe_edit",
            f"Suggest safe patch file: {file_query}",
            result[:5000],
            ["file", "safe", "patch", file_query]
        )

        return result

    if command_lower.startswith("suggest patch file "):
        file_query = command[
            len("suggest patch file "):
        ].strip()

        result = suggest_safe_patch_for_file(file_query)

        remember(
            "safe_edit",
            f"Suggest patch file: {file_query}",
            result[:5000],
            ["file", "patch", file_query]
        )

        return result

    if command_lower.startswith("safe patch file "):
        file_query = command[
            len("safe patch file "):
        ].strip()

        result = suggest_safe_patch_for_file(file_query)

        remember(
            "safe_edit",
            f"Safe patch file: {file_query}",
            result[:5000],
            ["file", "safe", "patch", file_query]
        )

        return result

    if command_lower == "list project file backups":
        result = list_project_file_backups()

        remember(
            "safe_edit",
            "Listed project file backups",
            result[:3000],
            ["project", "file", "backup", "list"]
        )

        return result

    if command_lower.startswith("preview project file "):
        args = command[
            len("preview project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        result = preview_project_file(project_name, file_query)

        remember(
            "project_file",
            f"Preview project file: {project_name} / {file_query}",
            result[:3000],
            ["project", "file", "preview", project_name, file_query]
        )

        return result

    if command_lower.startswith("backup project file "):
        args = command[
            len("backup project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        result = create_project_file_backup(project_name, file_query)

        remember(
            "safe_edit",
            f"Backup project file: {project_name} / {file_query}",
            result[:3000],
            ["project", "file", "backup", project_name, file_query]
        )

        return result

    if command_lower.startswith("create backup project file "):
        args = command[
            len("create backup project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        result = create_project_file_backup(project_name, file_query)

        remember(
            "safe_edit",
            f"Create backup project file: {project_name} / {file_query}",
            result[:3000],
            ["project", "file", "backup", project_name, file_query]
        )

        return result

    if command_lower.startswith("restore backup project file "):
        args = command[
            len("restore backup project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        result = restore_latest_project_file_backup(project_name, file_query)

        remember(
            "safe_edit",
            f"Restore backup project file: {project_name} / {file_query}",
            result[:3000],
            ["project", "file", "restore", project_name, file_query]
        )

        return result

    if command_lower.startswith("restore project file "):
        args = command[
            len("restore project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        result = restore_latest_project_file_backup(project_name, file_query)

        remember(
            "safe_edit",
            f"Restore project file: {project_name} / {file_query}",
            result[:3000],
            ["project", "file", "restore", project_name, file_query]
        )

        return result

    if command_lower.startswith("suggest safe patch project file "):
        args = command[
            len("suggest safe patch project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        result = suggest_safe_project_file_patch(project_name, file_query)

        remember(
            "safe_edit",
            f"Suggest safe patch project file: {project_name} / {file_query}",
            result[:5000],
            ["project", "file", "safe", "patch", project_name, file_query]
        )

        return result

    if command_lower.startswith("suggest patch project file "):
        args = command[
            len("suggest patch project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        result = suggest_safe_project_file_patch(project_name, file_query)

        remember(
            "safe_edit",
            f"Suggest patch project file: {project_name} / {file_query}",
            result[:5000],
            ["project", "file", "patch", project_name, file_query]
        )

        return result


    if command_lower.startswith("apply safe patch project file "):
        args = command[
            len("apply safe patch project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        update_hud_context(
            status="PROCESSING",
            command=command,
            project=project_name,
            file_path=file_query,
            action="Applying safe AI patch",
            ai_status="THINKING"
        )

        result = apply_ai_safe_project_file_patch(project_name, file_query)

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            project=project_name,
            file_path=file_query,
            action="Safe AI patch completed",
            ai_status="READY"
        )

        remember(
            "safe_edit",
            f"Applied safe AI patch project file: {project_name} / {file_query}",
            result[:5000],
            ["project", "file", "apply", "safe", "patch", project_name, file_query]
        )

        return result

    if command_lower.startswith("apply ai patch project file "):
        args = command[
            len("apply ai patch project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        result = apply_ai_safe_project_file_patch(project_name, file_query)

        remember(
            "safe_edit",
            f"Applied AI patch project file: {project_name} / {file_query}",
            result[:5000],
            ["project", "file", "apply", "ai", "patch", project_name, file_query]
        )

        return result

    if command_lower.startswith("auto improve project file "):
        args = command[
            len("auto improve project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        result = apply_ai_safe_project_file_patch(project_name, file_query)

        remember(
            "safe_edit",
            f"Auto improved project file: {project_name} / {file_query}",
            result[:5000],
            ["project", "file", "auto", "improve", project_name, file_query]
        )

        return result


    # ==========================
    # AI CODING ASSISTANT SHORTCUTS
    # Safer shortcuts mapped to existing backup + patch system.
    # ==========================
    if command_lower.startswith("fix project file "):
        args = command[
            len("fix project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        update_hud_context(
            status="PROCESSING",
            command=command,
            project=project_name,
            file_path=file_query,
            action="Applying safe AI fix",
            ai_status="THINKING",
            memory_status="SYNC",
            ollama_status="THINKING"
        )

        result = apply_ai_safe_project_file_patch(project_name, file_query)

        remember(
            "safe_edit",
            f"Fixed project file safely: {project_name} / {file_query}",
            result[:5000],
            ["project", "file", "fix", "safe", project_name, file_query]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            project=project_name,
            file_path=file_query,
            action="Safe AI fix completed",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower.startswith("fix file "):
        args = command[
            len("fix file "):
        ].strip()

        project_name, file_query, error = parse_file_from_project_text(args)

        if error:
            return error

        update_hud_context(
            status="PROCESSING",
            command=command,
            project=project_name,
            file_path=file_query,
            action="Applying safe AI fix",
            ai_status="THINKING",
            memory_status="SYNC",
            ollama_status="THINKING"
        )

        result = apply_ai_safe_project_file_patch(project_name, file_query)

        remember(
            "safe_edit",
            f"Fixed file safely: {project_name} / {file_query}",
            result[:5000],
            ["project", "file", "fix", "safe", project_name, file_query]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            project=project_name,
            file_path=file_query,
            action="Safe AI fix completed",
            ai_status="READY",
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower.startswith("secure project file "):
        args = command[
            len("secure project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        update_hud_context(
            status="PROCESSING",
            command=command,
            project=project_name,
            file_path=file_query,
            action="Applying safe security patch",
            ai_status="THINKING",
            memory_status="SYNC",
            ollama_status="THINKING"
        )

        result = apply_ai_safe_project_file_patch(project_name, file_query)

        remember(
            "safe_edit",
            f"Secured project file safely: {project_name} / {file_query}",
            result[:5000],
            ["project", "file", "security", "safe", project_name, file_query]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            project=project_name,
            file_path=file_query,
            action="Safe security patch completed",
            ai_status="READY",
            security_score=extract_hud_security_score(result),
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    if command_lower.startswith("secure file "):
        args = command[
            len("secure file "):
        ].strip()

        project_name, file_query, error = parse_file_from_project_text(args)

        if error:
            return error

        update_hud_context(
            status="PROCESSING",
            command=command,
            project=project_name,
            file_path=file_query,
            action="Applying safe security patch",
            ai_status="THINKING",
            memory_status="SYNC",
            ollama_status="THINKING"
        )

        result = apply_ai_safe_project_file_patch(project_name, file_query)

        remember(
            "safe_edit",
            f"Secured file safely: {project_name} / {file_query}",
            result[:5000],
            ["project", "file", "security", "safe", project_name, file_query]
        )

        update_hud_context(
            status="SUCCESS",
            command=command,
            result=result,
            project=project_name,
            file_path=file_query,
            action="Safe security patch completed",
            ai_status="READY",
            security_score=extract_hud_security_score(result),
            memory_status="SYNC",
            ollama_status="READY"
        )

        return result

    # ==========================
    # RANKED FILE NAVIGATION
    # Run:
    # rank file auth.py
    # Then:
    # open file #2
    # read file #2
    # review file #2
    # security review file #2
    # ==========================
    if command_lower.startswith("rank file "):
        file_query = command[
            len("rank file "):
        ].strip()

        result = rank_universal_file(file_query)

        remember(
            "universal_file",
            f"Ranked file matches: {file_query}",
            result[:5000],
            ["file", "rank", file_query]
        )

        return result

    if command_lower.startswith("open file #"):
        number = command[
            len("open file #"):
        ].strip()

        result = open_numbered_file(number)

        remember(
            "universal_file",
            f"Opened numbered file: #{number}",
            result[:3000],
            ["file", "open", "numbered", number]
        )

        return result

    if command_lower.startswith("read file #"):
        number = command[
            len("read file #"):
        ].strip()

        result = read_numbered_file(number)

        remember(
            "universal_file",
            f"Read numbered file: #{number}",
            result[:3000],
            ["file", "read", "numbered", number]
        )

        return result

    if command_lower.startswith("analyze file #"):
        number = command[
            len("analyze file #"):
        ].strip()

        result = analyze_numbered_file(number)

        remember(
            "universal_file",
            f"Analyzed numbered file: #{number}",
            result[:5000],
            ["file", "analyze", "numbered", number]
        )

        return result

    if command_lower.startswith("review file #"):
        number = command[
            len("review file #"):
        ].strip()

        result = review_numbered_file(number)

        remember(
            "universal_file",
            f"Reviewed numbered file: #{number}",
            result[:5000],
            ["file", "review", "numbered", number]
        )

        return result

    if command_lower.startswith("improve file #"):
        number = command[
            len("improve file #"):
        ].strip()

        result = improve_numbered_file(number)

        remember(
            "universal_file",
            f"Improved numbered file: #{number}",
            result[:5000],
            ["file", "improve", "numbered", number]
        )

        return result

    if command_lower.startswith("optimize file #"):
        number = command[
            len("optimize file #"):
        ].strip()

        result = optimize_numbered_file(number)

        remember(
            "universal_file",
            f"Optimized numbered file: #{number}",
            result[:5000],
            ["file", "optimize", "numbered", number]
        )

        return result

    if command_lower.startswith("security review file #"):
        number = command[
            len("security review file #"):
        ].strip()

        result = security_review_numbered_file(number)

        remember(
            "universal_file",
            f"Security reviewed numbered file: #{number}",
            result[:5000],
            ["file", "security", "review", "numbered", number]
        )

        return result

    # ==========================
    # BEST FILE MATCH COMMANDS
    # These commands automatically select the highest-scoring file match.
    # Examples:
    # open best auth.py from cyber
    # read best auth.py from cyber
    # review best auth.py from cyber
    # security review best auth.py from cyber
    # ==========================
    if command_lower.startswith("open best "):
        file_query = command[
            len("open best "):
        ].strip()

        result = open_best_universal_file(file_query)

        remember(
            "universal_file",
            f"Opened best file: {file_query}",
            result[:3000],
            ["file", "open", "best", file_query]
        )

        return result

    if command_lower.startswith("read best "):
        file_query = command[
            len("read best "):
        ].strip()

        result = read_best_universal_file(file_query)

        remember(
            "universal_file",
            f"Read best file: {file_query}",
            result[:3000],
            ["file", "read", "best", file_query]
        )

        return result

    if command_lower.startswith("analyze best "):
        file_query = command[
            len("analyze best "):
        ].strip()

        result = analyze_best_universal_file(file_query)

        remember(
            "universal_file",
            f"Analyzed best file: {file_query}",
            result[:5000],
            ["file", "analyze", "best", file_query]
        )

        return result

    if command_lower.startswith("review best "):
        file_query = command[
            len("review best "):
        ].strip()

        result = review_best_universal_file(file_query)

        remember(
            "universal_file",
            f"Reviewed best file: {file_query}",
            result[:5000],
            ["file", "review", "best", file_query]
        )

        return result

    if command_lower.startswith("security review best "):
        file_query = command[
            len("security review best "):
        ].strip()

        result = security_review_best_universal_file(file_query)

        remember(
            "universal_file",
            f"Security reviewed best file: {file_query}",
            result[:5000],
            ["file", "security", "review", "best", file_query]
        )

        return result

    # ==========================
    # UNIVERSAL FILE RESOLVER
    # Searches files in all indexed projects.
    # Supports:
    # find file auth.py
    # find file auth.py from cyber
    # open file routes/auth.py from cyber
    # read file app.py from jarvis
    # ==========================
    if command_lower.startswith("find file "):
        file_query = command[
            len("find file "):
        ].strip()

        result = find_universal_file(file_query)

        remember(
            "universal_file",
            f"Found file: {file_query}",
            result[:5000],
            ["file", "find", file_query]
        )

        return result

    # ==========================
    # FILE ASSISTANT
    # Universal resolver first, so it works across indexed projects.
    # ==========================
    if command_lower.startswith("open file "):
        file_query = command[
            len("open file "):
        ].strip()

        result = open_universal_file(file_query)

        remember(
            "memory_file",
            f"Opened file: {file_query}",
            result,
            ["file", "open", file_query]
        )

        return result

    if (
        command_lower.startswith("read file ")
        or command_lower.startswith("show file ")
    ):
        if command_lower.startswith("read file "):
            file_query = command[
                len("read file "):
            ].strip()
        else:
            file_query = command[
                len("show file "):
            ].strip()

        result = read_universal_file(file_query)

        remember(
            "memory_file",
            f"Read file: {file_query}",
            result[:3000],
            ["file", "read", file_query]
        )

        return result

    if command_lower.startswith("analyze file "):
        file_query = command[
            len("analyze file "):
        ].strip()

        result = analyze_universal_file(file_query)

        remember(
            "memory_file",
            f"Analyzed file: {file_query}",
            result[:5000],
            ["file", "analysis", file_query]
        )

        return result

    if command_lower.startswith("review file "):
        file_query = command[
            len("review file "):
        ].strip()

        result = review_universal_file(file_query)

        remember(
            "memory_file",
            f"Reviewed file: {file_query}",
            result[:5000],
            ["file", "review", file_query]
        )

        return result

    if command_lower.startswith("improve file "):
        file_query = command[
            len("improve file "):
        ].strip()

        result = improve_universal_file(file_query)

        remember(
            "memory_file",
            f"Improved file: {file_query}",
            result[:5000],
            ["file", "improve", file_query]
        )

        return result

    if command_lower.startswith("optimize file "):
        file_query = command[
            len("optimize file "):
        ].strip()

        result = optimize_universal_file(file_query)

        remember(
            "memory_file",
            f"Optimized file: {file_query}",
            result[:5000],
            ["file", "optimize", file_query]
        )

        return result

    if command_lower.startswith("security review "):
        file_query = command[
            len("security review "):
        ].strip()

        result = security_review_universal_file(file_query)

        remember(
            "memory_file",
            f"Security reviewed file: {file_query}",
            result[:5000],
            ["file", "security", "review", file_query]
        )

        return result

    # ==========================
    # PROJECT FILE ASSISTANT
    # ==========================
    if command_lower.startswith("open project file "):
        args = command[
            len("open project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        result = open_project_file(project_name, file_query)

        remember(
            "project_file",
            f"Opened project file: {project_name} / {file_query}",
            result,
            ["project", "file", "open", project_name, file_query]
        )

        return result

    if command_lower.startswith("read project file "):
        args = command[
            len("read project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        result = read_project_file(project_name, file_query)

        remember(
            "project_file",
            f"Read project file: {project_name} / {file_query}",
            result[:3000],
            ["project", "file", "read", project_name, file_query]
        )

        return result

    if command_lower.startswith("search project files "):
        args = command[
            len("search project files "):
        ].strip()

        project_name, keyword, error = parse_project_file_args(args)

        if error:
            return error

        result = search_project_files(project_name, keyword)

        remember(
            "project_file",
            f"Searched project files: {project_name} / {keyword}",
            result[:3000],
            ["project", "file", "search", project_name, keyword]
        )

        return result

    if command_lower.startswith("analyze project file "):
        args = command[
            len("analyze project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        result = analyze_project_file(project_name, file_query)

        remember(
            "project_file",
            f"Analyzed project file: {project_name} / {file_query}",
            result[:5000],
            ["project", "file", "analysis", project_name, file_query]
        )

        return result

    if command_lower.startswith("improve project file "):
        args = command[
            len("improve project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        result = improve_project_file(project_name, file_query)

        remember(
            "project_file",
            f"Improved project file: {project_name} / {file_query}",
            result[:5000],
            ["project", "file", "improve", project_name, file_query]
        )

        return result

    if command_lower.startswith("optimize project file "):
        args = command[
            len("optimize project file "):
        ].strip()

        project_name, file_query, error = parse_project_file_args(args)

        if error:
            return error

        result = optimize_project_file(project_name, file_query)

        remember(
            "project_file",
            f"Optimized project file: {project_name} / {file_query}",
            result[:5000],
            ["project", "file", "optimize", project_name, file_query]
        )

        return result

    # ==========================
    # OPEN PROJECT IN VS CODE
    # ==========================
    if (
        command_lower.startswith("open project ")
        and command_lower.endswith(" in vscode")
    ):
        project_name = command[
            len("open project "):-len(" in vscode")
        ].strip()

        result = open_project_in_vscode(project_name)

        remember(
            "project",
            f"Opened project in VS Code: {project_name}",
            result,
            ["project", "vscode", project_name]
        )

        return result

    if command_lower.startswith("edit project "):
        project_name = command[len("edit project "):].strip()

        result = open_project_in_vscode(project_name)

        remember(
            "project",
            f"Edited project: {project_name}",
            result,
            ["project", "edit", "vscode", project_name]
        )

        return result

    if command_lower.startswith("open code "):
        project_name = command[len("open code "):].strip()

        result = open_project_in_vscode(project_name)

        remember(
            "project",
            f"Opened code project: {project_name}",
            result,
            ["project", "code", "vscode", project_name]
        )

        return result

    # ==========================
    # OPEN WEBSITE
    # ==========================
    if command_lower.startswith("open website "):
        site = command[len("open website "):].strip()

        result = open_website(site)

        remember(
            "website",
            f"Opened website: {site}",
            result,
            ["website", site]
        )

        return result

    # ==========================
    # SEARCH FILE
    # ==========================
    if command_lower.startswith("search file "):
        file_name = command[len("search file "):].strip()

        files = search_files(file_name)

        if files:
            result = "\n".join(files)
        else:
            result = "No files found."

        remember(
            "file",
            f"Searched file: {file_name}",
            result,
            ["file", "search", file_name]
        )

        return result

    # ==========================
    # OPEN PROJECT
    # ==========================
    if command_lower.startswith("open project "):
        project_name = command[len("open project "):].strip()

        result = open_project(project_name)

        remember(
            "project",
            f"Opened project: {project_name}",
            result,
            ["project", "open", project_name]
        )

        return result

    # ==========================
    # ANALYZE PROJECT
    # ==========================
    if command_lower.startswith("analyze project "):
        project_name = command[len("analyze project "):].strip()

        result = analyze_project_by_name(project_name)

        remember(
            "project",
            f"Analyzed project: {project_name}",
            result[:5000],
            ["project", "analysis", project_name]
        )

        remember_project(
            project_name,
            result[:5000]
        )

        return result

    # ==========================
    # OPEN APPLICATION
    # ==========================
    if command_lower.startswith("open "):
        app_name = command[len("open "):].strip()

        result = open_installed_app(app_name)

        remember(
            "application",
            f"Opened application: {app_name}",
            result,
            ["application", app_name]
        )

        return result

    # ==========================
    # STRICT FALLBACK
    # ==========================
    return (
        "I did not execute this because it did not match a strict JARVIS command.\n"
        "Try one of these:\n"
        "- open calculator\n"
        "- open website wikipedia.org\n"
        "- open folder downloads\n"
        "- open project CyberShield AI\n"
        "- score project CyberShield AI\n"
        "- find file app.py from CyberShield AI\n"
        "- preview file app.py from CyberShield AI\n"
        "- backup project file CyberShield AI app.py\n"
        "- apply safe patch project file CyberShield AI app.py\n- show project structure CyberShield AI\n- show project files CyberShield AI\n- show project statistics CyberShield AI\n- show largest files CyberShield AI\n- show security report CyberShield AI\n- show backups\n- fix file app.py from CyberShield AI\n- review code on screen\n- explain error on screen\n"
    )


def main():
    ensure_indexes_ready()
    print("\nJ.A.R.V.I.S Agent Ready\n")

    print("[JARVIS] Checking Ollama local AI engine...")
    ollama_ready, ollama_message = startup_ollama_auto_repair()
    print(f"[JARVIS] Ollama ready: {ollama_ready}. {ollama_message}")


    while True:
        command = input("\nYou: ").strip()

        result = handle_command(command)

        if result == "exit":
            break

        print("\nJARVIS:\n")
        print(result)


if __name__ == "__main__":
    main()



# ==========================================================
# J.A.R.V.I.S MARK XLVII ENTERPRISE AUTONOMY UPGRADE
# Added safely at the end of jarvis_agent.py.
#
# Features:
# - advanced conversational memory
# - multi-step autonomous planning
# - Commander Mode
# - Architect Mode
# - IDE/project/file workflow control
# - autonomous debugging suggestions
# - safe patch generation
# - task/workflow management
# - unified high-level command router
# ==========================================================

JARVIS_MARK47_VERSION = "J.A.R.V.I.S Mark XLVII Enterprise Autonomy"
JARVIS_MEMORY_FILE = "jarvis_conversation_memory.json"
JARVIS_TASKS_FILE = "jarvis_tasks.json"
JARVIS_WORKFLOWS_FILE = "jarvis_workflows.json"
JARVIS_PATCH_DIR = "jarvis_patch_suggestions"
JARVIS_COMMAND_LIMIT = 500


def _j47_now():
    try:
        from datetime import datetime
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(time.time())


def _j47_safe_load(path, default=None):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as file:
            data = json.load(file)

        return data
    except Exception:
        return default if default is not None else {}


def _j47_safe_save(path, data):
    try:
        parent = os.path.dirname(path)

        if parent:
            os.makedirs(parent, exist_ok=True)

        with open(path, "w", encoding="utf-8") as file:
            json.dump(data, file, indent=2, ensure_ascii=False)

        return True
    except Exception:
        return False


def _j47_clean(text):
    text = str(text or "").strip()
    text = text.replace("—", "-").replace("–", "-")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _j47_lower(text):
    return _j47_clean(text).lower()


def _j47_safe_name(name):
    name = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(name or "").strip())
    name = name.strip("_")
    return name or "jarvis"


def _j47_project_alias(name):
    lower = _j47_lower(name)

    aliases = {
        "cyber": "CyberShield AI",
        "cyber shield": "CyberShield AI",
        "cyber shield ai": "CyberShield AI",
        "cybershield": "CyberShield AI",
        "cybershield ai": "CyberShield AI",
        "cybers in the": "CyberShield AI",
        "jarvis": "J.A.R.V.I.S",
        "jervis": "J.A.R.V.I.S",
        "j a r v i s": "J.A.R.V.I.S",
        "manager app": "ManagerApp",
        "managerapp": "ManagerApp",
    }

    return aliases.get(lower, str(name).strip())


def _j47_extract_project(command, default_last=True):
    text = _j47_clean(command)
    lower = text.lower()

    for alias in [
        "cyber shield ai",
        "cybershield ai",
        "cyber shield",
        "cyber",
        "j.a.r.v.i.s",
        "j a r v i s",
        "jarvis",
        "manager app",
        "managerapp",
    ]:
        if re.search(rf"\b{re.escape(alias)}\b", lower):
            return _j47_project_alias(alias)

    match = re.search(
        r"\b(?:project|for|from|in|of)\s+(.+?)(?:\s+and\s+|\s+then\s+|\s+as\s+|\s+with\s+|$)",
        text,
        flags=re.IGNORECASE
    )

    if match:
        candidate = match.group(1).strip(" .,:;")
        candidate = re.sub(
            r"\b(pdf|word|docx|excel|xlsx|ppt|pptx|powerpoint|report|review|audit|security|architecture|roadmap|task|workflow|patch|fix|debug)\b",
            "",
            candidate,
            flags=re.IGNORECASE
        ).strip()

        if candidate:
            return _j47_project_alias(candidate)

    if default_last:
        memory = _j47_safe_load(JARVIS_MEMORY_FILE, {})
        last_project = memory.get("last_project")

        if last_project:
            return last_project

    return "CyberShield AI"


def _j47_remember(command=None, response=None, project=None, file_path=None, mode=None, extra=None):
    memory = _j47_safe_load(JARVIS_MEMORY_FILE, {})

    history = memory.get("history", [])

    if not isinstance(history, list):
        history = []

    entry = {
        "time": _j47_now(),
        "command": str(command or ""),
        "response": str(response or "")[:2000],
        "project": project,
        "file": file_path,
        "mode": mode,
        "extra": extra or {},
    }

    history.append(entry)
    history = history[-JARVIS_COMMAND_LIMIT:]

    memory["history"] = history
    memory["updated_at"] = _j47_now()

    if project:
        memory["last_project"] = project

    if file_path:
        memory["last_file"] = file_path

    if mode:
        memory["last_mode"] = mode

    _j47_safe_save(JARVIS_MEMORY_FILE, memory)


def j47_memory_summary(limit=12):
    memory = _j47_safe_load(JARVIS_MEMORY_FILE, {})
    history = memory.get("history", [])

    output = [
        "J.A.R.V.I.S MARK XLVII MEMORY",
        "",
        f"Last project: {memory.get('last_project', 'None')}",
        f"Last file: {memory.get('last_file', 'None')}",
        f"Last mode: {memory.get('last_mode', 'None')}",
        f"Entries: {len(history)}",
        "",
        "Recent activity:"
    ]

    for item in history[-limit:]:
        output.append(
            f"- {item.get('time')} | {item.get('mode') or 'general'} | "
            f"{item.get('command', '')[:110]}"
        )

    return "\n".join(output)


def j47_remember_current_context(project=None, file_path=None, mode=None):
    project = project or _j47_extract_project("", default_last=True)

    _j47_remember(
        command="remember current context",
        response="Context saved.",
        project=project,
        file_path=file_path,
        mode=mode or "context"
    )

    return f"Context remembered: project={project}, file={file_path or 'None'}, mode={mode or 'context'}"


# ==========================================================
# TASK MANAGEMENT
# ==========================================================
def _j47_load_tasks():
    data = _j47_safe_load(JARVIS_TASKS_FILE, [])

    if isinstance(data, list):
        return data

    return []


def _j47_save_tasks(tasks):
    return _j47_safe_save(JARVIS_TASKS_FILE, tasks)


def j47_add_task(title, project=None, priority="normal", status="todo", notes=""):
    tasks = _j47_load_tasks()

    task = {
        "id": f"TASK-{int(time.time())}-{len(tasks)+1}",
        "created": _j47_now(),
        "updated": _j47_now(),
        "title": str(title).strip(),
        "project": project or _j47_extract_project(title, default_last=True),
        "priority": priority,
        "status": status,
        "notes": notes,
    }

    tasks.append(task)
    _j47_save_tasks(tasks)

    return f"Task added: {task['id']} - {task['title']}"


def j47_list_tasks(project=None, status=None):
    tasks = _j47_load_tasks()

    if project:
        project = _j47_project_alias(project)
        tasks = [task for task in tasks if _j47_lower(task.get("project")) == _j47_lower(project)]

    if status:
        tasks = [task for task in tasks if _j47_lower(task.get("status")) == _j47_lower(status)]

    if not tasks:
        return "No tasks found."

    output = ["J.A.R.V.I.S TASKS", ""]

    for task in tasks[-50:]:
        output.append(
            f"- {task.get('id')} | {task.get('status')} | {task.get('priority')} | "
            f"{task.get('project')} | {task.get('title')}"
        )

    return "\n".join(output)


def j47_update_task(task_id, status=None, notes=None, priority=None):
    tasks = _j47_load_tasks()
    found = False

    for task in tasks:
        if str(task.get("id")).lower() == str(task_id).lower():
            if status:
                task["status"] = status
            if notes is not None:
                task["notes"] = notes
            if priority:
                task["priority"] = priority
            task["updated"] = _j47_now()
            found = True
            break

    if not found:
        return f"Task not found: {task_id}"

    _j47_save_tasks(tasks)
    return f"Task updated: {task_id}"


def j47_task_command(command):
    text = _j47_clean(command)
    lower = text.lower()

    if lower in {"tasks", "show tasks", "list tasks", "task list"}:
        return j47_list_tasks()

    match = re.match(r"^(?:add|create)\s+task\s+(.+)$", text, flags=re.IGNORECASE)
    if match:
        title = match.group(1).strip()
        project = _j47_extract_project(title, default_last=True)
        return j47_add_task(title, project=project)

    match = re.match(r"^show\s+tasks\s+(?:for|from|in)\s+project\s+(.+)$", text, flags=re.IGNORECASE)
    if match:
        return j47_list_tasks(project=match.group(1).strip())

    match = re.match(r"^(?:mark|set)\s+task\s+(.+?)\s+(?:as|to)\s+(todo|doing|done|blocked)$", text, flags=re.IGNORECASE)
    if match:
        return j47_update_task(match.group(1).strip(), status=match.group(2).strip())

    return None


# ==========================================================
# AUTONOMOUS PLANNER
# ==========================================================
def j47_make_plan(goal, project=None):
    project = project or _j47_extract_project(goal, default_last=True)
    goal_lower = _j47_lower(goal)

    steps = []

    if any(word in goal_lower for word in ["security", "audit", "secure", "vulnerab"]):
        steps = [
            f"Find project {project}",
            f"Run full security audit for {project}",
            f"Find hardcoded secrets in {project}",
            f"Find SQL injection risks in {project}",
            f"Find XSS risks in {project}",
            f"Generate security roadmap for {project}",
            f"Create PDF security report for {project}",
        ]
    elif any(word in goal_lower for word in ["release", "deploy", "production", "go live"]):
        steps = [
            f"Review project {project}",
            f"Estimate production readiness for {project}",
            f"Prepare release checklist for {project}",
            f"Prepare deployment checklist for {project}",
            f"Find highest risk vulnerabilities for {project}",
            f"Create PDF release report for {project}",
        ]
    elif any(word in goal_lower for word in ["architecture", "architect", "design"]):
        steps = [
            f"Review project structure for {project}",
            f"Run architecture analyzer for {project}",
            f"Find duplicate code in {project}",
            f"Find dead code in {project}",
            f"Generate project roadmap for {project}",
            f"Create PowerPoint architecture report for {project}",
        ]
    elif any(word in goal_lower for word in ["debug", "error", "bug", "fix"]):
        steps = [
            f"Read current error or target file for {project}",
            f"Find related code files in {project}",
            f"Review suspicious code",
            f"Generate safe patch suggestion",
            f"Create backup before changes",
            f"Apply patch only after approval",
        ]
    else:
        steps = [
            f"Find project {project}",
            f"Review project {project}",
            f"Score project {project}",
            f"Find dead code in {project}",
            f"Find duplicate code in {project}",
            f"Generate project roadmap for {project}",
            f"Create PDF project report for {project}",
        ]

    plan = {
        "id": f"PLAN-{int(time.time())}",
        "created": _j47_now(),
        "goal": goal,
        "project": project,
        "steps": steps,
        "status": "planned",
    }

    workflows = _j47_safe_load(JARVIS_WORKFLOWS_FILE, [])

    if not isinstance(workflows, list):
        workflows = []

    workflows.append(plan)
    _j47_safe_save(JARVIS_WORKFLOWS_FILE, workflows)

    return (
        f"Autonomous plan created: {plan['id']}\n"
        f"Project: {project}\n"
        f"Goal: {goal}\n\n"
        + "\n".join(f"{index}. {step}" for index, step in enumerate(steps, start=1))
    )


def j47_show_workflows():
    workflows = _j47_safe_load(JARVIS_WORKFLOWS_FILE, [])

    if not isinstance(workflows, list) or not workflows:
        return "No workflows found."

    output = ["J.A.R.V.I.S WORKFLOWS", ""]

    for workflow in workflows[-20:]:
        output.append(
            f"- {workflow.get('id')} | {workflow.get('status')} | "
            f"{workflow.get('project')} | {workflow.get('goal')}"
        )

    return "\n".join(output)


def j47_run_plan_step_text(step):
    lower = _j47_lower(step)

    if lower.startswith("find project "):
        return handle_project_lookup_command(step.lower())

    if lower.startswith("review project "):
        project = step[len("review project "):].strip()
        return review_project(project)

    if lower.startswith("score project "):
        project = step[len("score project "):].strip()
        return score_project(project)

    if "full security audit" in lower:
        project = _j47_extract_project(step)
        return full_security_audit(project)

    if "hardcoded secrets" in lower:
        project = _j47_extract_project(step)
        return find_hardcoded_secrets(project)

    if "sql injection" in lower:
        project = _j47_extract_project(step)
        return find_sql_injection(project)

    if "xss" in lower:
        project = _j47_extract_project(step)
        return find_xss_risks(project)

    if "security roadmap" in lower:
        project = _j47_extract_project(step)
        return generate_security_roadmap(project)

    if "dead code" in lower:
        project = _j47_extract_project(step)
        return find_dead_code(project)

    if "duplicate code" in lower:
        project = _j47_extract_project(step)
        return find_duplicate_code(project)

    if "architecture analyzer" in lower:
        project = _j47_extract_project(step)
        return strict_architecture_analyzer_project(project)

    if "production readiness" in lower:
        project = _j47_extract_project(step)
        return production_readiness(project)

    if "release checklist" in lower:
        project = _j47_extract_project(step)
        return release_checklist(project)

    if "deployment checklist" in lower:
        project = _j47_extract_project(step)
        return deployment_checklist(project)

    if "roadmap" in lower:
        project = _j47_extract_project(step)
        return generate_project_roadmap(project)

    if "pdf" in lower and "report" in lower:
        project = _j47_extract_project(step)
        return _jarvis_export_report_any(project, "pdf", report_kind="project_review")

    if "powerpoint" in lower and "report" in lower:
        project = _j47_extract_project(step)
        return _jarvis_export_report_any(project, "powerpoint", report_kind="architecture")

    return f"Step noted but not executed automatically: {step}"


def j47_execute_last_workflow(max_steps=3):
    workflows = _j47_safe_load(JARVIS_WORKFLOWS_FILE, [])

    if not isinstance(workflows, list) or not workflows:
        return "No workflow available."

    workflow = workflows[-1]
    steps = workflow.get("steps", [])

    if not steps:
        return "Workflow has no steps."

    results = [
        f"Executing workflow: {workflow.get('id')}",
        f"Goal: {workflow.get('goal')}",
        "",
    ]

    for index, step in enumerate(steps[:max_steps], start=1):
        try:
            result = j47_run_plan_step_text(step)
        except Exception as error:
            result = f"ERROR: {error}"

        results.append(f"STEP {index}: {step}")
        results.append(str(result)[:3000])
        results.append("")

    workflow["status"] = "partially_executed"
    workflow["last_executed"] = _j47_now()
    _j47_safe_save(JARVIS_WORKFLOWS_FILE, workflows)

    return "\n".join(results)


# ==========================================================
# COMMANDER / ARCHITECT MODES
# ==========================================================
def j47_commander_mode(project_name=None):
    project = project_name or _j47_extract_project("", default_last=True)

    result_parts = [
        "J.A.R.V.I.S COMMANDER MODE",
        f"Project: {project}",
        "",
    ]

    try:
        result_parts.append("PROJECT SCORE")
        result_parts.append(score_project(project))
    except Exception as error:
        result_parts.append(f"Score unavailable: {error}")

    try:
        result_parts.append("\nNEXT BEST IMPROVEMENTS")
        result_parts.append(next_best_improvements(project))
    except Exception as error:
        result_parts.append(f"Next improvements unavailable: {error}")

    try:
        result_parts.append("\nHIGHEST RISK VULNERABILITIES")
        result_parts.append(highest_risk_vulnerabilities(project))
    except Exception as error:
        result_parts.append(f"Risk analysis unavailable: {error}")

    try:
        result_parts.append("\nPRODUCTION READINESS")
        result_parts.append(production_readiness(project))
    except Exception as error:
        result_parts.append(f"Production readiness unavailable: {error}")

    response = "\n".join(str(item) for item in result_parts)
    _j47_remember(command="commander mode", response=response, project=project, mode="commander")
    return response


def j47_architect_mode(project_name=None):
    project = project_name or _j47_extract_project("", default_last=True)

    result_parts = [
        "J.A.R.V.I.S ARCHITECT MODE",
        f"Project: {project}",
        "",
    ]

    try:
        result_parts.append("ARCHITECTURE ANALYSIS")
        result_parts.append(strict_architecture_analyzer_project(project))
    except Exception as error:
        result_parts.append(f"Architecture analyzer unavailable: {error}")

    try:
        result_parts.append("\nPROJECT ROADMAP")
        result_parts.append(generate_project_roadmap(project))
    except Exception as error:
        result_parts.append(f"Roadmap unavailable: {error}")

    try:
        result_parts.append("\nDUPLICATE CODE")
        result_parts.append(find_duplicate_code(project))
    except Exception as error:
        result_parts.append(f"Duplicate code scan unavailable: {error}")

    try:
        result_parts.append("\nDEAD CODE")
        result_parts.append(find_dead_code(project))
    except Exception as error:
        result_parts.append(f"Dead code scan unavailable: {error}")

    response = "\n".join(str(item) for item in result_parts)
    _j47_remember(command="architect mode", response=response, project=project, mode="architect")
    return response


# ==========================================================
# AUTONOMOUS DEBUGGING AND PATCHES
# ==========================================================
def _j47_extract_code_block(text):
    text = str(text or "")

    match = re.search(r"```(?:[a-zA-Z0-9_+-]+)?\n(.*?)```", text, flags=re.S)

    if match:
        return match.group(1).strip()

    return ""


def j47_debug_error(error_text, project_name=None):
    project = project_name or _j47_extract_project(str(error_text), default_last=True)

    prompt = f"""
You are J.A.R.V.I.S Mark XLVII, a senior debugging assistant.

Project: {project}

Analyze this error and return:
1. Most likely cause
2. Files to inspect
3. Step-by-step fix plan
4. Safe patch suggestion if possible
5. Commands to test after fix

ERROR:
{error_text}
"""

    try:
        response = ask_llm(prompt)
    except Exception as error:
        response = f"AI debugging unavailable: {error}"

    _j47_remember(command="debug error", response=response, project=project, mode="debug")
    return response


def j47_generate_safe_patch(project_name, file_query, instruction):
    project = _j47_project_alias(project_name)

    try:
        file_path, error = _jarvis_parse_project_file(project, file_query)
    except Exception as err:
        return f"Could not resolve file: {err}"

    if error:
        return error

    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as file:
            original_code = file.read()
    except Exception as err:
        return f"Could not read file: {err}"

    prompt = f"""
You are J.A.R.V.I.S Mark XLVII.

Create a SAFE patch suggestion only.
Do NOT claim the patch was applied.
Keep the same behavior unless the instruction requires otherwise.

Project: {project}
File: {file_path}
Instruction: {instruction}

Return:
1. Summary
2. Risks
3. Full improved code in one code block

CURRENT CODE:
```text
{original_code[:60000]}
```
"""

    try:
        response = ask_llm(prompt)
    except Exception as error:
        return f"AI patch generation unavailable: {error}"

    os.makedirs(JARVIS_PATCH_DIR, exist_ok=True)

    safe_project = _j47_safe_name(project)
    safe_file = _j47_safe_name(os.path.basename(file_path))
    patch_path = os.path.join(
        JARVIS_PATCH_DIR,
        f"{safe_project}_{safe_file}_{int(time.time())}.patch.md"
    )

    with open(patch_path, "w", encoding="utf-8", errors="ignore") as file:
        file.write(response)

    _j47_remember(
        command=f"generate safe patch {file_query}",
        response=response,
        project=project,
        file_path=file_path,
        mode="patch",
        extra={"patch_path": patch_path}
    )

    return f"Safe patch suggestion created:\n{patch_path}\n\n{response[:4000]}"


def j47_apply_last_patch_preview_only():
    memory = _j47_safe_load(JARVIS_MEMORY_FILE, {})
    history = memory.get("history", [])

    for item in reversed(history):
        extra = item.get("extra", {})

        if extra.get("patch_path"):
            return (
                "Last patch is only a suggestion and was NOT applied automatically.\n"
                f"Patch file: {extra.get('patch_path')}\n"
                "Review it manually before applying."
            )

    return "No patch suggestion found."


# ==========================================================
# IDE AND FILE CONTROL
# ==========================================================
def j47_open_project_in_ide(command):
    text = _j47_clean(command)

    match = re.search(
        r"(?:open|launch|start)\s+(?:project\s+)?(.+?)\s+(?:in|with)\s+(vs code|vscode|visual studio code|visual studio|visual studio community|vs community|intellij|intellij idea|pycharm|android studio|eclipse)$",
        text,
        flags=re.IGNORECASE
    )

    if not match:
        return None

    project = _j47_project_alias(match.group(1).strip())
    ide = match.group(2).strip()

    try:
        result = open_project_in_app(project, ide)
    except Exception:
        try:
            result = open_project_in_vscode(project)
        except Exception as error:
            result = f"Could not open project in IDE: {error}"

    _j47_remember(command=command, response=result, project=project, mode="ide")
    return result


def j47_open_last_project_in_ide(ide="vs code"):
    project = _j47_extract_project("", default_last=True)

    try:
        result = open_project_in_app(project, ide)
    except Exception as error:
        result = f"Could not open last project in {ide}: {error}"

    _j47_remember(command=f"open last project in {ide}", response=result, project=project, mode="ide")
    return result


# ==========================================================
# ROUTER
# ==========================================================
def handle_mark47_command(command):
    original = _j47_clean(command)

    if not original:
        return None

    lower = original.lower()

    # Health / status
    if lower in {
        "mark 47 status",
        "mark xlvii status",
        "jarvis status",
        "jarvis health",
        "health check",
    }:
        try:
            base = jarvis_health()
        except Exception:
            base = "Base health unavailable."

        return base + "\n\n" + j47_memory_summary(limit=6)

    if lower in {"memory", "memory summary", "show memory", "conversation memory"}:
        return j47_memory_summary()

    # Tasks
    task_result = j47_task_command(original)
    if task_result is not None:
        _j47_remember(command=original, response=task_result, mode="tasks")
        return task_result

    # Workflows/plans
    if lower in {"show workflows", "list workflows", "workflows"}:
        return j47_show_workflows()

    if lower in {"execute workflow", "run workflow", "execute last workflow", "run last workflow"}:
        return j47_execute_last_workflow(max_steps=3)

    match = re.match(r"^(?:plan|create plan|make plan)\s+(.+)$", original, flags=re.IGNORECASE)
    if match:
        goal = match.group(1).strip()
        project = _j47_extract_project(goal, default_last=True)
        result = j47_make_plan(goal, project)
        _j47_remember(command=original, response=result, project=project, mode="planner")
        return result

    # Modes
    if lower in {"commander mode", "open commander mode", "start commander mode"}:
        return j47_commander_mode()

    match = re.match(r"^(?:commander mode|start commander mode|open commander mode)\s+(?:for\s+)?(?:project\s+)?(.+)$", original, flags=re.IGNORECASE)
    if match:
        return j47_commander_mode(_j47_project_alias(match.group(1).strip()))

    if lower in {"architect mode", "open architect mode", "start architect mode"}:
        return j47_architect_mode()

    match = re.match(r"^(?:architect mode|start architect mode|open architect mode)\s+(?:for\s+)?(?:project\s+)?(.+)$", original, flags=re.IGNORECASE)
    if match:
        return j47_architect_mode(_j47_project_alias(match.group(1).strip()))

    # Debugging
    match = re.match(r"^(?:debug|explain error|analyze error)\s+(.+)$", original, flags=re.IGNORECASE)
    if match:
        return j47_debug_error(match.group(1).strip())

    match = re.match(
        r"^(?:generate|create|suggest)\s+(?:safe\s+)?patch\s+(?:for\s+)?(?:file\s+)?(.+?)\s+(?:in|from)\s+project\s+(.+?)\s+(?:to|that|which|where)\s+(.+)$",
        original,
        flags=re.IGNORECASE
    )
    if match:
        file_query = match.group(1).strip()
        project = _j47_project_alias(match.group(2).strip())
        instruction = match.group(3).strip()
        return j47_generate_safe_patch(project, file_query, instruction)

    if lower in {"apply last patch", "apply patch", "use last patch"}:
        return j47_apply_last_patch_preview_only()

    # IDE control
    ide_result = j47_open_project_in_ide(original)
    if ide_result is not None:
        return ide_result

    match = re.match(r"^open\s+last\s+project\s+in\s+(.+)$", original, flags=re.IGNORECASE)
    if match:
        return j47_open_last_project_in_ide(match.group(1).strip())

    # Report shortcut variants
    match = re.match(
        r"^(?:create|generate|make|export)\s+(pdf|word|docx|excel|xlsx|powerpoint|pptx|ppt|html|json|csv)\s+(?:report|review|presentation)\s+(?:for\s+)?(?:project\s+)?(.+)$",
        original,
        flags=re.IGNORECASE
    )
    if match:
        fmt = match.group(1).strip()
        project = _j47_project_alias(match.group(2).strip())
        result = _jarvis_export_report_any(project, fmt, report_kind="project_review")
        _j47_remember(command=original, response=result, project=project, mode="report")
        return result

    # Natural "do full enterprise analysis"
    if any(phrase in lower for phrase in [
        "full enterprise analysis",
        "analyze everything",
        "complete project analysis",
        "full jarvis analysis",
    ]):
        project = _j47_extract_project(original, default_last=True)
        result = j47_make_plan("full enterprise analysis", project)
        result += "\n\n" + j47_execute_last_workflow(max_steps=4)
        _j47_remember(command=original, response=result, project=project, mode="autonomous")
        return result

    return None


# ==========================================================
# FINAL SUPER ROUTER WRAPPER
# This wraps the existing handle_command if it exists.
# ==========================================================
try:
    _JARVIS_PRE_MARK47_HANDLE_COMMAND = handle_command
except Exception:
    _JARVIS_PRE_MARK47_HANDLE_COMMAND = None


def handle_command(command):
    command = _j47_clean(command)

    try:
        hud_start_action(
            command,
            action="Mark XLVII router",
            project=_j47_extract_project(command, default_last=True),
            thinking=False
        )
    except Exception:
        pass

    # 1. Mark XLVII advanced router.
    try:
        result = handle_mark47_command(command)

        if result is not None:
            _j47_remember(
                command=command,
                response=result,
                project=_j47_extract_project(command, default_last=True),
                mode="mark47"
            )

            try:
                hud_finish_action(
                    command,
                    result,
                    action="Mark XLVII command",
                    project=_j47_extract_project(command, default_last=True)
                )
            except Exception:
                pass

            return result
    except Exception as error:
        return f"Mark XLVII router error: {error}"

    # 2. Existing router from your original file.
    if _JARVIS_PRE_MARK47_HANDLE_COMMAND is not None:
        try:
            result = _JARVIS_PRE_MARK47_HANDLE_COMMAND(command)

            _j47_remember(
                command=command,
                response=result,
                project=_j47_extract_project(command, default_last=True),
                mode="legacy"
            )

            return result
        except Exception as error:
            return f"JARVIS command error: {error}"

    # 3. LLM fallback.
    try:
        result = ask_llm(command)
        _j47_remember(command=command, response=result, mode="llm")
        return result
    except Exception as error:
        return f"JARVIS fallback error: {error}"


def process_command(command):
    return handle_command(command)


def ask_jarvis(command):
    return handle_command(command)


def jarvis_mark47_capabilities():
    return """
J.A.R.V.I.S Mark XLVII capabilities:

- Advanced conversation memory
- Multi-step autonomous planning
- Commander Mode
- Architect Mode
- Task and workflow management
- IDE control for VS Code, Visual Studio, IntelliJ, PyCharm, Android Studio
- Project/file/code navigation
- Debugging assistant
- Safe patch suggestions
- Report generation: PDF, Word, PowerPoint, Excel, HTML, JSON, CSV
- Context-aware project commands
- Legacy command compatibility
""".strip()



# ==========================================================
# J.A.R.V.I.S PROJECT + IDE DIRECT ROUTER FIX
# Added at the end so it overrides the previous handle_command safely.
#
# Fixes:
# - open jarvis in vs code
# - open project jarvis in vs code
# - open cyber shield ai in vs code
# - open manager app in intellij
#
# Problem fixed:
# The old router sometimes sent "jarvis in vs code" to Windows
# as a literal app/file name. This router detects project + IDE first.
# ==========================================================

PROJECT_IDE_DIRECT_ROUTER_VERSION = "J.A.R.V.I.S Project IDE Direct Router Fix"

_PROJECT_IDE_ALIASES = {
    "vs code": "VS Code",
    "vscode": "VS Code",
    "visual studio code": "VS Code",
    "code": "VS Code",

    "visual studio": "Visual Studio Community",
    "visual studio community": "Visual Studio Community",
    "vs community": "Visual Studio Community",
    "visual studio professional": "Visual Studio Professional",
    "visual studio enterprise": "Visual Studio Enterprise",

    "intellij": "IntelliJ",
    "intellij idea": "IntelliJ",
    "intelli j": "IntelliJ",
    "idea": "IntelliJ",

    "eclipse": "Eclipse",
    "eclips": "Eclipse",
    "pycharm": "PyCharm",
    "android studio": "Android Studio",
    "webstorm": "WebStorm",
    "rider": "Rider",
    "clion": "CLion",
    "cursor": "Cursor",
    "windsurf": "Windsurf",
}

_PROJECT_NAME_ALIASES = {
    "cyber": "CyberShield AI",
    "cyber shield": "CyberShield AI",
    "cyber shield ai": "CyberShield AI",
    "cybershield": "CyberShield AI",
    "cybershield ai": "CyberShield AI",
    "cybershiel ai": "CyberShield AI",
    "cyber shiel ai": "CyberShield AI",
    "cybers in the": "CyberShield AI",

    "jarvis": "J.A.R.V.I.S",
    "jervis": "J.A.R.V.I.S",
    "j a r v i s": "J.A.R.V.I.S",
    "jarvis agent": "J.A.R.V.I.S",

    "manager app": "ManagerApp",
    "managerapp": "ManagerApp",
    "manager application": "ManagerApp",
}

_PROJECT_IDE_NOISE_WORDS = {
    "studio",
    "student",
    "students",
    "study",
    "studies",
    "steady",
    "status",
    "stereo",
    "audio",
    "video",
}


def _project_ide_router_clean(text):
    text = str(text or "").lower().strip()
    text = text.replace("-", " ").replace("_", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _project_ide_router_normalize_project(project_name):
    lower = _project_ide_router_clean(project_name)

    words = [
        word
        for word in lower.split()
        if word not in _PROJECT_IDE_NOISE_WORDS
    ]

    lower = " ".join(words).strip()
    lower = re.sub(r"^(the|my|a|an)\s+", "", lower).strip()
    lower = re.sub(r"\s+(project|application|app)$", "", lower).strip()

    if lower in _PROJECT_NAME_ALIASES:
        return _PROJECT_NAME_ALIASES[lower]

    best_name = None
    best_score = 0.0

    for alias, canonical in _PROJECT_NAME_ALIASES.items():
        score = 0.0

        if alias in lower or lower in alias:
            score = 0.90
        else:
            try:
                import difflib
                score = difflib.SequenceMatcher(None, lower, alias).ratio()
            except Exception:
                score = 0.0

        if score > best_score:
            best_score = score
            best_name = canonical

    if best_name and best_score >= 0.72:
        return best_name

    return str(project_name or "").strip()


def _project_ide_router_normalize_ide(ide_name):
    lower = _project_ide_router_clean(ide_name)

    if lower in _PROJECT_IDE_ALIASES:
        return _PROJECT_IDE_ALIASES[lower]

    best_ide = None
    best_score = 0.0

    for alias, canonical in _PROJECT_IDE_ALIASES.items():
        try:
            import difflib
            score = difflib.SequenceMatcher(None, lower, alias).ratio()
        except Exception:
            score = 0.0

        if score > best_score:
            best_score = score
            best_ide = canonical

    if best_ide and best_score >= 0.72:
        return best_ide

    return str(ide_name or "").strip()


def _project_ide_router_strip_noise(command):
    text = _project_ide_router_clean(command)

    ide_pattern = "|".join(
        re.escape(alias)
        for alias in sorted(_PROJECT_IDE_ALIASES.keys(), key=len, reverse=True)
    )

    noise_pattern = "|".join(
        re.escape(word)
        for word in sorted(_PROJECT_IDE_NOISE_WORDS, key=len, reverse=True)
    )

    text = re.sub(
        rf"\s+({noise_pattern})\s+(?=(?:in|with|using|on)\s+(?:{ide_pattern})\b)",
        " ",
        text,
        flags=re.IGNORECASE
    )

    text = re.sub(
        rf"\s+({noise_pattern})\s+(?=(?:{ide_pattern})\b)",
        " in ",
        text,
        flags=re.IGNORECASE
    )

    return re.sub(r"\s+", " ", text).strip()


def parse_project_ide_direct_command(command):
    original = str(command or "").strip()

    if not original:
        return None

    text = _project_ide_router_strip_noise(original)

    ide_pattern = "|".join(
        re.escape(alias)
        for alias in sorted(_PROJECT_IDE_ALIASES.keys(), key=len, reverse=True)
    )

    patterns = [
        # open project jarvis in vs code
        rf"^(?:open|launch|start|edit|load)\s+(?:the\s+|my\s+)?(?:project\s+)?(.+?)\s+(?:in|with|using|on)\s+({ide_pattern})$",

        # open jarvis project in vs code
        rf"^(?:open|launch|start|edit|load)\s+(?:the\s+|my\s+)?(.+?)\s+project\s+(?:in|with|using|on)\s+({ide_pattern})$",
    ]

    for pattern in patterns:
        match = re.match(pattern, text, flags=re.IGNORECASE)

        if match:
            raw_project = match.group(1).strip()
            raw_ide = match.group(2).strip()

            project = _project_ide_router_normalize_project(raw_project)
            ide = _project_ide_router_normalize_ide(raw_ide)

            if not project:
                return None

            return {
                "project": project,
                "ide": ide,
                "command": f"open project {project} in {ide}",
            }

    return None


def handle_project_ide_direct_command(command):
    parsed = parse_project_ide_direct_command(command)

    if not parsed:
        return None

    project = parsed["project"]
    ide = parsed["ide"]

    try:
        hud_start_action(
            command,
            action="Open project in IDE",
            project=project,
            thinking=False
        )
    except Exception:
        pass

    result = None

    # Prefer tools.py launcher because it has the most complete project/app index.
    try:
        result = open_project_in_app(project, ide)
    except Exception as error:
        result = f"Could not open {project} in {ide}: {error}"

    # Fallback to embedded developer assistant if tools failed.
    if (
        not result
        or str(result).lower().startswith("project not found")
        or "application/ide not found" in str(result).lower()
        or "could not find ide" in str(result).lower()
        or "could not open" in str(result).lower()
    ):
        try:
            ns = globals().get("_EMBEDDED_DEVELOPER_ASSISTANT_NAMESPACE", {})
            opener = ns.get("open_project_in_ide")

            if opener:
                fallback_result = opener(project, ide)

                if fallback_result and "could not" not in str(fallback_result).lower():
                    result = fallback_result
        except Exception:
            pass

    try:
        remember_project(project)
    except Exception:
        pass

    try:
        hud_finish_action(
            command,
            result,
            action="Open project in IDE",
            project=project
        )
    except Exception:
        pass

    return result or f"Could not open {project} in {ide}."


try:
    _PRE_PROJECT_IDE_DIRECT_HANDLE_COMMAND = handle_command
except Exception:
    _PRE_PROJECT_IDE_DIRECT_HANDLE_COMMAND = None


def handle_command(command):
    command = str(command or "").strip()

    if not command:
        return "No command received."

    # Highest priority: project + IDE commands.
    result = handle_project_ide_direct_command(command)

    if result is not None:
        return result

    # Also try after a small normalization for the exact bug:
    # "jarvis in vs code" must never be sent directly to Windows.
    lower = _project_ide_router_clean(command)

    if re.match(r"^(?:open|launch|start|edit|load)\s+.+\s+(?:in|with|using|on)\s+.+$", lower):
        parsed = parse_project_ide_direct_command(lower)

        if parsed:
            result = handle_project_ide_direct_command(parsed["command"])

            if result is not None:
                return result

    if _PRE_PROJECT_IDE_DIRECT_HANDLE_COMMAND is not None:
        return _PRE_PROJECT_IDE_DIRECT_HANDLE_COMMAND(command)

    return "JARVIS base command handler is not available."


def process_command(command):
    return handle_command(command)


def ask_jarvis(command):
    return handle_command(command)


def project_ide_direct_router_self_test():
    tests = [
        "open jarvis in vs code",
        "open project jarvis in vs code",
        "open jervis in vscode",
        "open cyber shield ai in vs code",
        "open project CyberShield AI studio in VS Code",
        "open manager app in intellij",
        "open project managerapp in eclipse",
    ]

    lines = [
        "PROJECT IDE DIRECT ROUTER SELF TEST",
        f"Version: {PROJECT_IDE_DIRECT_ROUTER_VERSION}",
        "",
    ]

    for item in tests:
        parsed = parse_project_ide_direct_command(item)
        lines.append(f"RAW: {item}")
        lines.append(f"PARSED: {parsed}")
        lines.append("")

    return "\n".join(lines)

